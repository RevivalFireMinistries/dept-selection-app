from fastapi import APIRouter, Depends, HTTPException, Query, Body, Response, Request
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, or_, and_
from typing import Optional, Dict, Any, List, Tuple
from io import BytesIO
from datetime import datetime, date, timedelta
import os
import re
import uuid
import json

from database import get_db
from models import Category, Department, Member, MemberDepartment, Settings, Appeal, Meeting, MeetingRSVP, NotificationConfig, NotificationLog, PosterRequest, ServiceProgram, ProgramTemplate, ServiceSchedule, AdminAuditLog, HomeChurch, HomeChurchProgramType, HomeChurchRoster, HomeChurchAttendance, Survey, SurveyQuestion, SurveyResponse, SurveyAnswer, MemberChangeRequest, Announcement
from schemas import (
    CategoryCreate, CategoryUpdate, CategoryResponse,
    DepartmentCreate, DepartmentUpdate, DepartmentResponse, DepartmentInCategory,
    MemberSubmission, MemberResponse,
    SettingUpdate, DepartmentsGroupedResponse,
    ReviewStatusUpdate, ReplaceDepartmentRequest, AssignDepartmentRequest,
    AppealCreate, AppealResolve,
    SetHODRequest,
    MeetingCreate, MeetingUpdate, RSVPRequest,
    SMTPSettingsUpdate, NotificationConfigUpdate, TestEmailRequest,
    PosterRequestCreate, PosterRequestResponse,
    ServiceProgramCreate, ServiceProgramUpdate, ServiceProgramResponse,
    ProgramTemplateCreate, ProgramTemplateUpdate,
    ServiceScheduleCreate, ServiceScheduleUpdate
)

router = APIRouter()


def validate_phone(phone: str) -> bool:
    """Validate phone number is exactly 10 digits"""
    digits = re.sub(r'\D', '', phone)
    return len(digits) == 10


def check_selections_locked(db: Session):
    """Raise 403 if results are published (selections are locked for non-admins)"""
    setting = db.query(Settings).filter(Settings.key == "resultsPublished").first()
    if setting and setting.value == "true":
        raise HTTPException(
            status_code=403,
            detail="Department selections are locked. Results have been published."
        )


# ============ AUTHENTICATION ============

import bcrypt
import secrets

def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def _verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))


def _mask_email(email: str) -> str:
    """Mask email for display: j***n@example.com"""
    if not email or '@' not in email:
        return email or ""
    local, domain = email.split('@', 1)
    if len(local) <= 2:
        masked = local[0] + '***'
    else:
        masked = local[0] + '***' + local[-1]
    return f"{masked}@{domain}"


def _send_password_setup_email(db, member, token: str):
    """Send a password setup email to a member who doesn't have a password yet."""
    try:
        from notifications.channels.rfm_notify import RfmNotifyChannel

        channel = RfmNotifyChannel()
        if not channel.is_configured():
            print("rfm-notify not configured — cannot send password setup email")
            return

        base_url = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
        if base_url and not base_url.startswith("http"):
            base_url = f"https://{base_url}"
        if not base_url:
            base_url = os.environ.get("BASE_URL", "http://localhost:8000")
        reset_url = f"{base_url}/reset-password?token={token}"

        FONT = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif"
        html = f"""
        <div style="max-width:480px;margin:0 auto;font-family:{FONT};color:#111827;">
            <div style="border-top:3px solid #6366f1;padding:32px 0 16px;">
                <h1 style="font-size:20px;margin:0 0 8px;">Set Up Your Password</h1>
                <p style="color:#6b7280;font-size:14px;margin:0 0 8px;">
                    Hi {member.full_name},
                </p>
                <p style="color:#6b7280;font-size:14px;margin:0 0 24px;">
                    The RFM Stellenbosch Portal now requires a password to sign in. Click below to set up your password.
                </p>
                <a href="{reset_url}" style="display:inline-block;padding:12px 32px;background:#6366f1;color:#ffffff;text-decoration:none;border-radius:8px;font-weight:600;font-size:14px;">
                    Set My Password
                </a>
                <p style="color:#9ca3af;font-size:12px;margin:24px 0 0;">
                    This link expires in 24 hours. If you didn't try to log in, you can safely ignore this email.
                </p>
            </div>
            <div style="border-top:1px solid #e5e7eb;padding:16px 0;margin-top:24px;">
                <p style="color:#9ca3af;font-size:11px;margin:0;">RFM Stellenbosch Portal</p>
            </div>
        </div>
        """

        success, error = channel.send(
            member.email,
            "Set Up Your Password - RFM Stellenbosch Portal",
            html,
            event_code="member.password_setup",
            recipient_id=getattr(member, "id", None),
            recipient_name=getattr(member, "full_name", None),
            idempotency_key=f"password_setup:{getattr(member, 'id', member.email)}:{token}",
            priority="high",  # transactional, not marketing
        )
        if not success:
            print(f"Failed to send password setup email: {error}")
        else:
            print(f"Sent password setup email to {member.email}")

    except Exception as e:
        print(f"Error sending password setup email: {e}")


@router.post("/auth/register")
def register_member(
    request: Request,
    data: dict = Body(...),
    db: Session = Depends(get_db)
):
    """Register a new member account (requires admin approval)"""
    full_name = (data.get("full_name") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    address = (data.get("address") or "").strip()
    password = data.get("password") or ""

    if not full_name:
        raise HTTPException(status_code=400, detail="Full name is required")
    if not phone or not validate_phone(phone):
        raise HTTPException(status_code=400, detail="Valid 10-digit phone number is required")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    if len(password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    # Check if phone already exists
    normalized = phone.strip().replace(" ", "").replace("-", "")
    existing = db.query(Member).all()
    for m in existing:
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized:
            raise HTTPException(status_code=400, detail="An account with this phone number already exists. Try logging in or use Forgot Password.")

    member = Member(
        full_name=full_name,
        phone=phone,
        email=email,
        address=address or "",
        password_hash=_hash_password(password),
        is_active=False  # Requires admin approval
    )
    db.add(member)
    db.flush()
    _log_member_action(request, db, member, "register", "member", member.id, f"New registration (pending approval)")
    db.commit()

    return {"success": True, "message": "Your account has been created and is pending admin approval. You will be notified once approved."}


@router.post("/auth/login")
def login_member(
    data: dict = Body(...),
    db: Session = Depends(get_db)
):
    """Login with phone + password"""
    phone = (data.get("phone") or "").strip()
    password = data.get("password") or ""

    if not phone:
        raise HTTPException(status_code=400, detail="Phone number is required")

    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        raise HTTPException(status_code=401, detail="No account found with this phone number")

    # First-time login: member has no password yet — phone number is the default password
    if not member.password_hash:
        # Verify they entered their phone number as the password
        phone_digits = member.phone.strip().replace(" ", "").replace("-", "")
        password_digits = password.strip().replace(" ", "").replace("-", "")
        if not password or password_digits != phone_digits:
            raise HTTPException(status_code=401, detail="First time signing in? Use your phone number as your password")

        # Generate token and redirect to set-password page
        token = secrets.token_urlsafe(32)
        member.reset_token = token
        member.reset_token_expires = datetime.utcnow() + timedelta(hours=2)
        db.commit()

        has_email = bool(member.email and member.email.strip())

        return JSONResponse(
            status_code=200,
            content={
                "needs_password": True,
                "token": token,
                "has_email": has_email,
                "full_name": member.full_name
            }
        )

    # Password provided — verify it
    if not password:
        raise HTTPException(status_code=401, detail="Please enter your password")

    if not _verify_password(password, member.password_hash):
        raise HTTPException(status_code=401, detail="Incorrect password")

    # Check if account is active (admin approved)
    if not member.is_active:
        raise HTTPException(status_code=403, detail="Your account is pending admin approval. You will be notified once approved.")

    # Create response with session cookie
    from routers.pages import _sign_member_session, MEMBER_COOKIE_NAME, SESSION_MAX_AGE
    token = _sign_member_session(member.id)
    response = JSONResponse(content={
        "success": True,
        "member_id": member.id,
        "phone": member.phone,
        "full_name": member.full_name
    })
    response.set_cookie(
        key=MEMBER_COOKIE_NAME,
        value=token,
        httponly=True,
        max_age=SESSION_MAX_AGE,
        samesite="lax"
    )
    return response


@router.post("/auth/forgot-password")
def forgot_password(
    data: dict = Body(...),
    db: Session = Depends(get_db)
):
    """Send password reset email"""
    phone = (data.get("phone") or "").strip()

    if not phone:
        raise HTTPException(status_code=400, detail="Phone number is required")

    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        # Don't reveal whether the phone exists
        return {"success": True, "message": "If an account exists with this phone number, a reset link has been sent to the registered email."}

    if not member.email:
        raise HTTPException(status_code=400, detail="No email address on file. Please contact an admin.")

    # Generate reset token
    token = secrets.token_urlsafe(32)
    member.reset_token = token
    member.reset_token_expires = datetime.utcnow() + timedelta(hours=2)
    db.commit()

    # Send email
    try:
        from notifications.dispatcher import get_email_settings

        settings_dict = {}
        for s in db.query(Settings).all():
            settings_dict[s.key] = s.value

        email_settings = get_email_settings(settings_dict)

        from notifications.channels.rfm_notify import RfmNotifyChannel
        channel = RfmNotifyChannel()
        if not channel.is_configured():
            print("rfm-notify not configured for password reset")
            return {"success": True, "message": "If an account exists with this phone number, a reset link has been sent to the registered email."}

        # Build reset URL
        base_url = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
        if base_url and not base_url.startswith("http"):
            base_url = f"https://{base_url}"
        if not base_url:
            base_url = os.environ.get("BASE_URL", "http://localhost:8000")
        reset_url = f"{base_url}/reset-password?token={token}"

        FONT = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif"
        html = f"""
        <div style="max-width:480px;margin:0 auto;font-family:{FONT};color:#111827;">
            <div style="border-top:3px solid #6366f1;padding:32px 0 16px;">
                <h1 style="font-size:20px;margin:0 0 8px;">Reset Your Password</h1>
                <p style="color:#6b7280;font-size:14px;margin:0 0 24px;">
                    We received a request to reset your password for the RFM Stellenbosch Portal.
                </p>
                <a href="{reset_url}" style="display:inline-block;padding:12px 32px;background:#6366f1;color:#ffffff;text-decoration:none;border-radius:8px;font-weight:600;font-size:14px;">
                    Reset Password
                </a>
                <p style="color:#9ca3af;font-size:12px;margin:24px 0 0;">
                    This link expires in 2 hours. If you didn't request this, you can safely ignore this email.
                </p>
            </div>
            <div style="border-top:1px solid #e5e7eb;padding:16px 0;margin-top:24px;">
                <p style="color:#9ca3af;font-size:11px;margin:0;">RFM Stellenbosch Portal</p>
            </div>
        </div>
        """

        success, error = channel.send(
            member.email,
            "Reset Your Password - RFM Stellenbosch Portal",
            html,
            event_code="member.password_reset",
            recipient_id=getattr(member, "id", None),
            recipient_name=getattr(member, "full_name", None),
            idempotency_key=f"password_reset:{getattr(member, 'id', member.email)}:{token}",
            priority="high",
        )
        if not success:
            print(f"Failed to send reset email: {error}")

    except Exception as e:
        print(f"Error sending reset email: {e}")

    return {"success": True, "message": "If an account exists with this phone number, a reset link has been sent to the registered email."}


@router.post("/auth/reset-password")
def reset_password(
    data: dict = Body(...),
    db: Session = Depends(get_db)
):
    """Reset password using token from email"""
    token = (data.get("token") or "").strip()
    new_password = data.get("password") or ""

    if not token:
        raise HTTPException(status_code=400, detail="Reset token is required")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    member = db.query(Member).filter(
        Member.reset_token == token,
        Member.reset_token_expires > datetime.utcnow()
    ).first()

    if not member:
        raise HTTPException(status_code=400, detail="Invalid or expired reset link. Please request a new one.")

    # Don't allow phone number as password
    phone_digits = member.phone.strip().replace(" ", "").replace("-", "")
    password_digits = new_password.strip().replace(" ", "").replace("-", "")
    if password_digits == phone_digits:
        raise HTTPException(status_code=400, detail="Your password cannot be your phone number. Please choose a different password.")

    member.password_hash = _hash_password(new_password)
    member.reset_token = None
    member.reset_token_expires = None

    # Allow updating email if provided (for members who didn't have one)
    new_email = (data.get("email") or "").strip()
    if new_email:
        member.email = new_email

    db.commit()

    return {"success": True, "message": "Password has been set successfully. You can now sign in."}


@router.get("/auth/me")
def get_current_user(request: Request, db: Session = Depends(get_db)):
    """Check if user is logged in and return their info including departments"""
    from routers.pages import get_current_member
    member = get_current_member(request, db)
    if not member:
        return {"logged_in": False}

    roles = []
    if member.leadership_roles:
        try:
            roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else member.leadership_roles
        except (ValueError, TypeError):
            roles = []

    # Get approved departments with category info
    member_with_depts = db.query(Member).options(
        joinedload(Member.departments).joinedload(MemberDepartment.department).joinedload(Department.category)
    ).filter(Member.id == member.id).first()

    departments = []
    if member_with_depts:
        for md in member_with_depts.departments:
            if md.status == "approved":
                dept_name = md.department.name if md.department else ""
                cat_name = md.department.category.name if md.department and md.department.category else ""
                departments.append({"id": md.department.id, "name": dept_name, "category": cat_name})

    # Check if member still needs to set up a password
    needs_password_setup = not member.password_hash

    return {
        "logged_in": True,
        "member_id": member.id,
        "full_name": member.full_name,
        "phone": member.phone,
        "email": member.email,
        "leadership_roles": roles,
        "departments": departments,
        "needs_password_setup": needs_password_setup,
        "can_create_surveys": bool(getattr(member, "can_create_surveys", False)),
        "is_hc_leader": db.query(HomeChurch).filter(HomeChurch.leader_member_id == member.id).count() > 0,
    }


# ============ AUDIT LOGGING ============

def _log_admin_action(request: Request, db: Session, action: str, entity_type: str = None, entity_id: int = None, details: str = None):
    """Log an admin action for audit trail"""
    from routers.pages import get_admin_identity
    admin = get_admin_identity(request)
    log = AdminAuditLog(
        admin_member_id=admin["member_id"] if admin else None,
        admin_name=admin["name"] if admin else "Unknown",
        actor_type="admin",
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        ip_address=request.client.host if request.client else None
    )
    db.add(log)


def _log_member_action(request: Request, db: Session, member, action: str, entity_type: str = None, entity_id: int = None, details: str = None):
    """Log a member/user action for audit trail"""
    log = AdminAuditLog(
        admin_member_id=member.id if member else None,
        admin_name=member.full_name if member else "Unknown",
        actor_type="member",
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        ip_address=request.client.host if request.client else None
    )
    db.add(log)


def _log_system_action(db: Session, action: str, entity_type: str = None, entity_id: int = None, details: str = None):
    """Log a system/automated action for audit trail"""
    log = AdminAuditLog(
        admin_member_id=None,
        admin_name="System",
        actor_type="system",
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        ip_address=None
    )
    db.add(log)


@router.get("/admin/identity")
def get_admin_identity_endpoint(request: Request):
    """Get the current admin's identity from cookie"""
    from routers.pages import get_admin_identity
    admin = get_admin_identity(request)
    if not admin:
        return {"authenticated": False}
    return {"authenticated": True, "member_id": admin["member_id"], "name": admin["name"]}


@router.get("/admin/audit-log")
def get_audit_log(
    limit: int = Query(100),
    offset: int = Query(0),
    actor_type: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
    member_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Get audit log entries with filtering"""
    query = db.query(AdminAuditLog)

    if actor_type:
        query = query.filter(AdminAuditLog.actor_type == actor_type)
    if action:
        query = query.filter(AdminAuditLog.action == action)
    if member_id:
        query = query.filter(AdminAuditLog.admin_member_id == member_id)
    if search:
        query = query.filter(
            or_(
                AdminAuditLog.admin_name.ilike(f"%{search}%"),
                AdminAuditLog.details.ilike(f"%{search}%"),
                AdminAuditLog.action.ilike(f"%{search}%")
            )
        )
    if date_from:
        query = query.filter(AdminAuditLog.created_at >= date_from)
    if date_to:
        query = query.filter(AdminAuditLog.created_at <= date_to + "T23:59:59")

    total = query.count()
    logs = query.order_by(AdminAuditLog.created_at.desc()).offset(offset).limit(limit).all()

    # Get distinct actors and actions for filter dropdowns
    all_actors = db.query(AdminAuditLog.admin_name, AdminAuditLog.admin_member_id, AdminAuditLog.actor_type)\
        .distinct().order_by(AdminAuditLog.admin_name).all()
    all_actions = db.query(AdminAuditLog.action).distinct().order_by(AdminAuditLog.action).all()

    return {
        "total": total,
        "actors": [{"name": a[0], "id": a[1], "actor_type": a[2]} for a in all_actors],
        "actions": [a[0] for a in all_actions],
        "logs": [{
            "id": l.id,
            "admin_name": l.admin_name,
            "admin_member_id": l.admin_member_id,
            "actor_type": l.actor_type if hasattr(l, 'actor_type') and l.actor_type else "admin",
            "action": l.action,
            "entity_type": l.entity_type,
            "entity_id": l.entity_id,
            "details": l.details,
            "ip_address": l.ip_address,
            "created_at": l.created_at.isoformat() if l.created_at else None
        } for l in logs]
    }


# ============ ADMIN: MEMBER APPROVAL ============

@router.get("/admin/pending-registrations")
def get_pending_registrations(db: Session = Depends(get_db)):
    """Admin: get all inactive (pending approval) members"""
    pending = db.query(Member).filter(Member.is_active == False).order_by(Member.created_at.desc()).all()
    return [{
        "id": m.id,
        "full_name": m.full_name,
        "phone": m.phone,
        "email": m.email,
        "created_at": m.created_at.isoformat() if m.created_at else None
    } for m in pending]


@router.post("/admin/approve-member/{member_id}")
def approve_member(member_id: int, request: Request = None, db: Session = Depends(get_db)):
    """Admin: approve a pending member registration"""
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    member.is_active = True
    if request:
        _log_admin_action(request, db, "approve_member", "member", member_id, f"Approved registration for {member.full_name}")
    db.commit()

    # Send approval email
    try:
        from notifications.dispatcher import get_email_settings
        settings_dict = {}
        for s in db.query(Settings).all():
            settings_dict[s.key] = s.value
        email_settings = get_email_settings(settings_dict)

        from notifications.channels.rfm_notify import RfmNotifyChannel
        channel = RfmNotifyChannel()

        if channel.is_configured() and member.email:
            base_url = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
            if base_url and not base_url.startswith("http"):
                base_url = f"https://{base_url}"
            if not base_url:
                base_url = os.environ.get("BASE_URL", "http://localhost:8000")

            FONT = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif"
            html = f"""
            <div style="max-width:480px;margin:0 auto;font-family:{FONT};color:#111827;">
                <div style="border-top:3px solid #059669;padding:32px 0 16px;">
                    <h1 style="font-size:20px;margin:0 0 8px;">Account Approved!</h1>
                    <p style="color:#6b7280;font-size:14px;margin:0 0 24px;">
                        Hi {member.full_name}, your RFM Stellenbosch Portal account has been approved. You can now log in.
                    </p>
                    <a href="{base_url}" style="display:inline-block;padding:12px 32px;background:#059669;color:#ffffff;text-decoration:none;border-radius:8px;font-weight:600;font-size:14px;">
                        Log In Now
                    </a>
                </div>
                <div style="border-top:1px solid #e5e7eb;padding:16px 0;margin-top:24px;">
                    <p style="color:#9ca3af;font-size:11px;margin:0;">RFM Stellenbosch Portal</p>
                </div>
            </div>
            """
            channel.send(
                member.email,
                "Account Approved - RFM Stellenbosch Portal",
                html,
                event_code="member.account_approved",
                recipient_id=getattr(member, "id", None),
                recipient_name=getattr(member, "full_name", None),
                idempotency_key=f"account_approved:{getattr(member, 'id', member.email)}",
            )
    except Exception as e:
        print(f"Error sending approval email: {e}")

    return {"success": True}


@router.post("/admin/reject-member/{member_id}")
def reject_member(member_id: int, request: Request = None, db: Session = Depends(get_db)):
    """Admin: reject and delete a pending member registration"""
    member = db.query(Member).filter(Member.id == member_id, Member.is_active == False).first()
    if not member:
        raise HTTPException(status_code=404, detail="Pending member not found")
    member_name = member.full_name
    if request:
        _log_admin_action(request, db, "reject_member", "member", member_id, f"Rejected registration for {member_name}")
    db.delete(member)
    db.commit()
    return {"success": True}


# ============ DEPARTMENTS ============

@router.get("/departments")
def get_departments(db: Session = Depends(get_db)):
    """Get all departments grouped by category"""
    categories = db.query(Category).options(
        joinedload(Category.departments).joinedload(Department.hod)
    ).order_by(Category.name).all()

    uncategorized = db.query(Department).options(
        joinedload(Department.hod)
    ).filter(
        Department.category_id == None
    ).order_by(Department.name).all()

    def dept_dict(d):
        result = {"id": d.id, "name": d.name, "categoryId": d.category_id}
        if d.hod:
            result["hod"] = {"id": d.hod.id, "fullName": d.hod.full_name, "phone": d.hod.phone}
        else:
            result["hod"] = None
        return result

    return {
        "categories": [
            {
                "id": cat.id,
                "name": cat.name,
                "maxSelections": cat.max_selections,
                "createdAt": cat.created_at.isoformat() if cat.created_at else None,
                "departments": [
                    dept_dict(d)
                    for d in sorted(cat.departments, key=lambda x: x.name)
                ]
            }
            for cat in categories
        ],
        "uncategorized": [
            dept_dict(d)
            for d in uncategorized
        ]
    }


def _sync_member_departments_to_central(member_id: int, db: Session) -> dict:
    """Best-effort propagation of a member's APPROVED department assignments
    to the central API. Used after assign / approve / reject / replace so
    the central record's department_ids stays in lockstep with our local
    member_departments rows.

    Returns a status dict for inclusion in the calling endpoint's response.
    Failures never block the local commit (kill switch / unreachable / member
    not linked / departments not yet synced are all silently no-op'd)."""
    status = {"attempted": False, "ok": False, "department_count": 0, "skipped_unlinked": 0, "error": None}
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        return status

    member = db.query(Member).filter(Member.id == member_id).first()
    if not member or not member.external_member_id:
        return status

    # Approved (or admin-source) member_department rows
    rows = db.query(MemberDepartment).options(
        joinedload(MemberDepartment.department)
    ).filter(
        MemberDepartment.member_id == member_id,
        MemberDepartment.status == "approved",
    ).all()

    ext_ids = []
    skipped = 0
    for md in rows:
        dept = md.department
        if dept and dept.external_department_id:
            ext_ids.append(str(dept.external_department_id))
        else:
            skipped += 1

    status["attempted"] = True
    status["department_count"] = len(ext_ids)
    status["skipped_unlinked"] = skipped
    try:
        result = _rfm.update_member(
            member.external_member_id,
            {"department_ids": ext_ids},
            db=db,
        )
        status["ok"] = result.ok
        if not result.ok:
            status["error"] = result.error
    except Exception as exc:
        status["error"] = f"unexpected: {exc}"
    return status


def _push_department_to_central(department: Department, request: Request, db: Session, *, action: str) -> dict:
    """Best-effort propagation of a department CRUD change to the central API.
    Returns a small status dict the caller can include in its response.
    Failures are logged + ignored so the local commit always wins."""
    status = {"attempted": False, "ok": False, "fields": [], "error": None}
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        return status

    try:
        if action == "delete":
            if not department.external_department_id:
                return status
            status["attempted"] = True
            r = _rfm.delete_department(department.external_department_id, db=db)
            status["ok"] = r.ok
            if not r.ok:
                status["error"] = r.error
            return status

        # Resolve assembly id; same priority chain as members
        assembly_id = _resolve_default_assembly_id(db)
        if not assembly_id:
            status["error"] = "could not resolve assembly_id; central sync skipped"
            return status

        if action == "create" or not department.external_department_id:
            # Try create. If the API rejects "name already exists" because
            # someone created the same dept centrally, fall back to matching
            # by name so we still link.
            payload = {
                "assembly_id": str(assembly_id),
                "name": department.name,
                "sort_order": department.id,
            }
            status["attempted"] = True
            r = _rfm.create_department(payload, db=db)
            if r.ok and isinstance(r.data, dict):
                rec = r.data.get("data", r.data) if "data" in r.data else r.data
                ext_id = rec.get("id") if isinstance(rec, dict) else None
                if ext_id:
                    department.external_department_id = str(ext_id)
                    from datetime import datetime as _dt
                    department.external_synced_at = _dt.utcnow()
                    db.commit()
                    status["ok"] = True
                    status["fields"] = ["created"]
                    return status
            # Try matching by name
            ls = _rfm.list_departments(assembly_id=str(assembly_id), include_inactive=True, db=db)
            if ls.ok and ls.data:
                items = ls.data if isinstance(ls.data, list) else ls.data.get("data") or []
                for d in items:
                    if _rfm.names_match_dept(department.name, d.get("name") or ""):
                        department.external_department_id = str(d.get("id"))
                        from datetime import datetime as _dt
                        department.external_synced_at = _dt.utcnow()
                        db.commit()
                        status["ok"] = True
                        status["fields"] = ["linked-by-name"]
                        return status
            status["error"] = r.error if not r.ok else "could not create or match dept centrally"
            return status

        # Update path
        status["attempted"] = True
        r = _rfm.update_department(department.external_department_id, {"name": department.name}, db=db)
        status["ok"] = r.ok
        status["fields"] = ["name"] if r.ok else []
        if not r.ok:
            status["error"] = r.error
        return status
    except Exception as exc:
        status["error"] = f"unexpected: {exc}"
        return status


def _home_church_payload(hc: HomeChurch, db: Session, *, assembly_id: str) -> dict:
    """Map a local HomeChurch into the central schema. The local model
    splits suburb/address/meeting_day/meeting_time/whatsapp_link, none of
    which map 1:1 to central — we put the structured bits into metadata
    and combine address+suburb into the human-facing 'location' string."""
    location_parts = [p for p in [hc.address, hc.suburb] if p]
    location = ", ".join(location_parts) if location_parts else None

    leader_id = None
    leader_name = None
    if hc.leader_member_id:
        leader = db.query(Member).filter(Member.id == hc.leader_member_id).first()
        if leader:
            leader_name = leader.full_name
            if leader.external_member_id:
                leader_id = str(leader.external_member_id)

    metadata = {
        "meeting_day": int(hc.meeting_day) if hc.meeting_day is not None else 0,
        "meeting_time": hc.meeting_time or "19:00",
        "whatsapp_link": hc.whatsapp_link or "",
        "suburb": hc.suburb or "",
        "address": hc.address or "",
        "local_id": hc.id,
    }
    payload = {
        "assembly_id": str(assembly_id),
        "name": hc.name,
        "leader_name": leader_name,
        "location": location,
        "notes": hc.notes,
        "sort_order": hc.id,
        "metadata": metadata,
    }
    if leader_id:
        payload["leader_id"] = leader_id
    return payload


def _push_home_church_to_central(hc: HomeChurch, request: Request, db: Session, *, action: str) -> dict:
    """Best-effort push of a HomeChurch CRUD change to the central API.
    Same pattern as departments: failures are logged + ignored, local commit wins."""
    status = {"attempted": False, "ok": False, "fields": [], "error": None}
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        return status

    try:
        if action == "delete":
            if not hc.external_home_church_id:
                return status
            status["attempted"] = True
            r = _rfm.delete_home_church(hc.external_home_church_id, db=db)
            status["ok"] = r.ok
            if not r.ok:
                status["error"] = r.error
            return status

        assembly_id = _resolve_default_assembly_id(db)
        if not assembly_id:
            status["error"] = "could not resolve assembly_id; central sync skipped"
            return status

        if action == "create" or not hc.external_home_church_id:
            payload = _home_church_payload(hc, db, assembly_id=assembly_id)
            status["attempted"] = True
            r = _rfm.create_home_church(payload, db=db)
            if r.ok and isinstance(r.data, dict):
                rec = r.data.get("data", r.data) if "data" in r.data else r.data
                ext_id = rec.get("id") if isinstance(rec, dict) else None
                if ext_id:
                    hc.external_home_church_id = str(ext_id)
                    from datetime import datetime as _dt
                    hc.external_synced_at = _dt.utcnow()
                    db.commit()
                    status["ok"] = True
                    status["fields"] = ["created"]
                    return status
            # Fall back: link by name if a record already exists centrally
            ls = _rfm.list_home_churches(assembly_id=str(assembly_id), include_inactive=True, db=db)
            if ls.ok and ls.data:
                items = ls.data if isinstance(ls.data, list) else ls.data.get("data") or []
                for d in items:
                    if _rfm.names_match_home_church(hc.name, d.get("name") or ""):
                        hc.external_home_church_id = str(d.get("id"))
                        from datetime import datetime as _dt
                        hc.external_synced_at = _dt.utcnow()
                        db.commit()
                        status["ok"] = True
                        status["fields"] = ["linked-by-name"]
                        return status
            status["error"] = r.error if not r.ok else "could not create or match home church centrally"
            return status

        # Update path
        payload = _home_church_payload(hc, db, assembly_id=assembly_id)
        # central PUT only accepts mutable fields (no assembly_id)
        update_fields = {k: v for k, v in payload.items() if k != "assembly_id"}
        update_fields["is_active"] = bool(hc.is_active)
        status["attempted"] = True
        r = _rfm.update_home_church(hc.external_home_church_id, update_fields, db=db)
        status["ok"] = r.ok
        status["fields"] = list(update_fields.keys()) if r.ok else []
        if not r.ok:
            status["error"] = r.error
        if r.ok:
            from datetime import datetime as _dt
            hc.external_synced_at = _dt.utcnow()
            db.commit()
        return status
    except Exception as exc:
        status["error"] = f"unexpected: {exc}"
        return status


@router.post("/departments")
def create_department(data: DepartmentCreate, request: Request, db: Session = Depends(get_db)):
    """Create a new department"""
    if not data.name:
        raise HTTPException(status_code=400, detail="Name is required")

    department = Department(name=data.name, category_id=data.category_id)
    db.add(department)
    db.commit()
    db.refresh(department)

    api_sync = _push_department_to_central(department, request, db, action="create")

    log_msg = f"Created department '{department.name}'"
    if api_sync["attempted"]:
        log_msg += f" (central: {'synced' if api_sync['ok'] else 'failed - ' + (api_sync['error'] or '?')})"
    _log_admin_action(request, db, "create_department", "department", department.id, log_msg)
    db.commit()

    return {
        "id": department.id, "name": department.name, "categoryId": department.category_id,
        "external_department_id": department.external_department_id,
        "api_sync": api_sync,
    }


@router.put("/departments")
def update_department(data: DepartmentUpdate, request: Request, db: Session = Depends(get_db)):
    """Update an existing department"""
    if not data.id or not data.name:
        raise HTTPException(status_code=400, detail="ID and name are required")

    department = db.query(Department).filter(Department.id == data.id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    old_name = department.name
    department.name = data.name
    department.category_id = data.category_id
    db.commit()

    api_sync = {"attempted": False, "ok": False, "fields": [], "error": None}
    if old_name != department.name or not department.external_department_id:
        api_sync = _push_department_to_central(department, request, db, action="update")

    log_msg = f"Updated department '{old_name}' → '{department.name}'"
    if api_sync["attempted"]:
        log_msg += f" (central: {'synced' if api_sync['ok'] else 'failed - ' + (api_sync['error'] or '?')})"
    _log_admin_action(request, db, "update_department", "department", department.id, log_msg)
    db.commit()

    return {
        "id": department.id, "name": department.name, "categoryId": department.category_id,
        "external_department_id": department.external_department_id,
        "api_sync": api_sync,
    }


@router.delete("/departments")
def delete_department(id: int = Query(...), request: Request = None, db: Session = Depends(get_db)):
    """Delete a department"""
    department = db.query(Department).filter(Department.id == id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    dept_name = department.name
    api_sync = _push_department_to_central(department, request, db, action="delete")

    db.delete(department)
    db.commit()

    if request:
        log_msg = f"Deleted department '{dept_name}'"
        if api_sync["attempted"]:
            log_msg += f" (central: {'soft-deleted' if api_sync['ok'] else 'failed - ' + (api_sync['error'] or '?')})"
        _log_admin_action(request, db, "delete_department", "department", id, log_msg)
        db.commit()

    return {"success": True, "api_sync": api_sync}


# ============ CATEGORIES ============

@router.get("/categories")
def get_categories(db: Session = Depends(get_db)):
    """Get all categories with their departments"""
    categories = db.query(Category).options(
        joinedload(Category.departments)
    ).order_by(Category.name).all()

    return [
        {
            "id": cat.id,
            "name": cat.name,
            "maxSelections": cat.max_selections,
            "createdAt": cat.created_at.isoformat() if cat.created_at else None,
            "departments": [
                {"id": d.id, "name": d.name, "categoryId": d.category_id}
                for d in sorted(cat.departments, key=lambda x: x.name)
            ]
        }
        for cat in categories
    ]


@router.post("/categories")
def create_category(data: CategoryCreate, request: Request, db: Session = Depends(get_db)):
    """Create a new category"""
    if not data.name:
        raise HTTPException(status_code=400, detail="Name is required")

    category = Category(name=data.name, max_selections=data.max_selections)
    db.add(category)
    db.commit()
    db.refresh(category)

    _log_admin_action(request, db, "create_category", "category", category.id, f"Created category '{category.name}' (max: {category.max_selections})")
    db.commit()

    return {"id": category.id, "name": category.name, "maxSelections": category.max_selections}


@router.put("/categories")
def update_category(data: CategoryUpdate, request: Request, db: Session = Depends(get_db)):
    """Update an existing category"""
    if not data.id or not data.name:
        raise HTTPException(status_code=400, detail="ID and name are required")

    category = db.query(Category).filter(Category.id == data.id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    old_name = category.name
    category.name = data.name
    category.max_selections = data.max_selections
    db.commit()

    _log_admin_action(request, db, "update_category", "category", category.id, f"Updated category '{old_name}' → '{category.name}'")
    db.commit()

    return {"id": category.id, "name": category.name, "maxSelections": category.max_selections}


@router.delete("/categories")
def delete_category(id: int = Query(...), request: Request = None, db: Session = Depends(get_db)):
    """Delete a category (departments become uncategorized)"""
    category = db.query(Category).filter(Category.id == id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    cat_name = category.name
    db.delete(category)
    db.commit()

    if request:
        _log_admin_action(request, db, "delete_category", "category", id, f"Deleted category '{cat_name}'")
        db.commit()

    return {"success": True}


# ============ MEMBERS ============

@router.get("/members")
def get_members(db: Session = Depends(get_db)):
    """Get all members with their departments"""
    members = db.query(Member).options(
        joinedload(Member.departments).joinedload(MemberDepartment.department).joinedload(Department.category)
    ).order_by(Member.created_at.desc()).all()

    return [
        {
            "id": m.id,
            "fullName": _title_case_name(m.full_name),
            "titledName": _get_titled_name(m),
            "phone": m.phone,
            "email": m.email,
            "address": m.address,
            "createdAt": m.created_at.isoformat() if m.created_at else None,
            "departments": [
                {
                    "id": md.id,
                    "memberId": md.member_id,
                    "departmentId": md.department_id,
                    "createdAt": md.created_at.isoformat() if md.created_at else None,
                    "department": {
                        "id": md.department.id,
                        "name": md.department.name,
                        "categoryId": md.department.category_id,
                        "category": {
                            "id": md.department.category.id,
                            "name": md.department.category.name
                        } if md.department.category else None
                    }
                }
                for md in m.departments
            ]
        }
        for m in members
    ]


@router.delete("/members")
def delete_member(id: int = Query(None), db: Session = Depends(get_db)):
    """Delete a member or all members if id=all"""
    if id is None:
        raise HTTPException(status_code=400, detail="Member ID is required")

    member = db.query(Member).filter(Member.id == id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    db.delete(member)
    db.commit()

    return {"success": True}


@router.delete("/members/purge")
def purge_all_members(db: Session = Depends(get_db)):
    """Delete all member submissions"""
    count = db.query(Member).count()
    db.query(MemberDepartment).delete()
    db.query(Member).delete()
    db.commit()

    return {"success": True, "deleted": count}


@router.get("/members/lookup")
def lookup_member_by_phone(phone: str = Query(...), db: Session = Depends(get_db)):
    """Lookup a member by phone number"""
    # Normalize phone - remove spaces and common formatting
    normalized = phone.strip().replace(" ", "").replace("-", "")

    # Try exact match first
    member = db.query(Member).filter(Member.phone == phone).first()

    # Try normalized match
    if not member:
        members = db.query(Member).all()
        for m in members:
            m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
            if m_normalized == normalized:
                member = m
                break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    return {
        "id": member.id,
        "fullName": member.full_name,
        "phone": member.phone
    }


@router.get("/members/{member_id}")
def get_member_by_id(member_id: int, db: Session = Depends(get_db)):
    """Get a single member by ID with their departments"""
    member = db.query(Member).options(
        joinedload(Member.departments).joinedload(MemberDepartment.department)
    ).filter(Member.id == member_id).first()

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    return {
        "id": member.id,
        "fullName": member.full_name,
        "phone": member.phone,
        "email": member.email,
        "address": member.address,
        "createdAt": member.created_at.isoformat() if member.created_at else None,
        "departments": [
            {
                "id": md.id,
                "departmentId": md.department_id,
                "department": {
                    "id": md.department.id,
                    "name": md.department.name
                }
            }
            for md in member.departments
        ]
    }


@router.put("/members/{member_id}")
def update_member(member_id: int, data: dict, source: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """Update a member's information and department selections"""
    if source != "desk":
        check_selections_locked(db)

    member = db.query(Member).filter(Member.id == member_id).first()

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Update basic info
    if "full_name" in data:
        member.full_name = _title_case_name(data["full_name"])
    if "email" in data:
        member.email = data["email"]
    if "address" in data:
        member.address = data["address"]

    # Update departments if provided
    if "selected_departments" in data:
        selected = data["selected_departments"]

        # Validate max departments
        max_setting = db.query(Settings).filter(Settings.key == "maxDepartments").first()
        max_departments = int(max_setting.value) if max_setting else 3

        if len(selected) > max_departments:
            raise HTTPException(
                status_code=400,
                detail=f"You can only select up to {max_departments} departments"
            )

        # Validate per-category limits
        departments = db.query(Department).options(
            joinedload(Department.category)
        ).filter(Department.id.in_(selected)).all()

        category_selections: Dict[int, list] = {}
        for dept in departments:
            if dept.category_id:
                if dept.category_id not in category_selections:
                    category_selections[dept.category_id] = []
                category_selections[dept.category_id].append(dept.id)

        for category_id, selected_dept_ids in category_selections.items():
            category = db.query(Category).filter(Category.id == category_id).first()
            max_allowed = category.max_selections if category else 1
            if len(selected_dept_ids) > max_allowed:
                raise HTTPException(
                    status_code=400,
                    detail=f"You can only select up to {max_allowed} department(s) from '{category.name}'"
                )

        # Get existing department associations with their statuses
        existing_mds = db.query(MemberDepartment).filter(
            MemberDepartment.member_id == member_id
        ).all()

        # Build a map of existing department_id -> MemberDepartment for status preservation
        existing_map = {md.department_id: md for md in existing_mds}
        existing_dept_ids = set(existing_map.keys())
        new_dept_ids = set(selected)

        # Departments to remove (in existing but not in new selection)
        to_remove = existing_dept_ids - new_dept_ids
        # Departments to add (in new selection but not existing)
        to_add = new_dept_ids - existing_dept_ids
        # Departments to keep (in both) - preserve their status
        to_keep = existing_dept_ids & new_dept_ids

        # Delete only the removed departments
        if to_remove:
            db.query(MemberDepartment).filter(
                MemberDepartment.member_id == member_id,
                MemberDepartment.department_id.in_(to_remove)
            ).delete(synchronize_session=False)

        # Create new associations only for newly added departments (with pending status)
        for dept_id in to_add:
            md = MemberDepartment(
                member_id=member_id,
                department_id=dept_id,
                source="member",
                status="pending"
            )
            db.add(md)

    db.commit()

    return {"success": True, "memberId": member_id}


# ============ SETTINGS ============

@router.get("/settings")
def get_settings(db: Session = Depends(get_db)):
    """Get all settings as key-value pairs"""
    settings = db.query(Settings).all()
    return {s.key: s.value for s in settings}


@router.put("/settings")
def update_setting(data: SettingUpdate, request: Request = None, db: Session = Depends(get_db)):
    """Update or create a setting"""
    if not data.key or data.value is None:
        raise HTTPException(status_code=400, detail="Key and value are required")

    setting = db.query(Settings).filter(Settings.key == data.key).first()
    if setting:
        setting.value = str(data.value)
    else:
        setting = Settings(key=data.key, value=str(data.value))
        db.add(setting)

    # Don't log password values
    safe_value = "****" if "password" in data.key.lower() else str(data.value)
    if request:
        _log_admin_action(request, db, "update_settings", "settings", None, f"Updated {data.key} = {safe_value}")
    db.commit()

    return {"success": True}


# ============ SUBMIT ============

@router.post("/submit")
def submit_form(data: MemberSubmission, request: Request, source: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """Submit member department selection form"""
    if source != "desk":
        check_selections_locked(db)

    # Validate required fields
    if not data.full_name or not data.phone or not data.address:
        raise HTTPException(status_code=400, detail="Full name, phone, and address are required")

    # Validate phone format (10 digits)
    if not validate_phone(data.phone):
        raise HTTPException(status_code=400, detail="Phone number must be 10 digits (e.g., 0711234456)")

    if not data.selected_departments or len(data.selected_departments) == 0:
        raise HTTPException(status_code=400, detail="Please select at least one department")

    # Check max departments limit
    max_setting = db.query(Settings).filter(Settings.key == "maxDepartments").first()
    max_departments = int(max_setting.value) if max_setting else 3

    if len(data.selected_departments) > max_departments:
        raise HTTPException(
            status_code=400,
            detail=f"You can only select up to {max_departments} departments"
        )

    # Check per-category selection limits
    departments = db.query(Department).options(
        joinedload(Department.category)
    ).filter(
        Department.id.in_(data.selected_departments)
    ).all()

    # Count selections per category
    category_selections: Dict[int, list] = {}
    for dept in departments:
        if dept.category_id:
            if dept.category_id not in category_selections:
                category_selections[dept.category_id] = []
            category_selections[dept.category_id].append(dept.id)

    # Validate against each category's max_selections
    for category_id, selected_dept_ids in category_selections.items():
        category = db.query(Category).filter(Category.id == category_id).first()
        max_allowed = category.max_selections if category else 1
        if len(selected_dept_ids) > max_allowed:
            raise HTTPException(
                status_code=400,
                detail=f"You can only select up to {max_allowed} department(s) from '{category.name}'"
            )

    # Create member
    member = Member(
        full_name=_title_case_name(data.full_name),
        phone=data.phone,
        email=data.email or "",
        address=data.address
    )
    db.add(member)
    db.flush()

    # Create member-department associations
    dept_names = [d.name for d in departments]
    for dept_id in data.selected_departments:
        md = MemberDepartment(member_id=member.id, department_id=dept_id)
        db.add(md)

    _log_member_action(request, db, member, "submit_selection", "member", member.id,
        f"Selected {len(data.selected_departments)} dept(s): {', '.join(dept_names)}")
    db.commit()
    db.refresh(member)

    return {"success": True, "memberId": member.id}


# ============ SEED ============

@router.api_route("/seed", methods=["GET", "POST"])
def seed_database(db: Session = Depends(get_db)):
    """Initialize database with default data"""
    # Check if already seeded
    existing = db.query(Settings).first()
    if existing:
        return {"message": "Database already seeded", "seeded": False}

    # Create default settings
    db.add(Settings(key="maxDepartments", value="3"))
    db.add(Settings(key="adminPassword", value="admin123"))
    db.add(Settings(key="deskPassword", value="desk123"))
    db.add(Settings(key="resultsPublished", value="false"))
    db.add(Settings(key="publishedAt", value=""))
    db.add(Settings(key="appealWindowOpen", value="false"))
    db.add(Settings(key="selectionYear", value="2026"))

    # Create categories with departments
    music = Category(name="Music Ministry")
    db.add(music)
    db.flush()
    db.add(Department(name="Choir", category_id=music.id))
    db.add(Department(name="Praise Team", category_id=music.id))
    db.add(Department(name="Sound & Media", category_id=music.id))

    children = Category(name="Children's Ministry")
    db.add(children)
    db.flush()
    db.add(Department(name="Sunday School Teachers", category_id=children.id))
    db.add(Department(name="Nursery", category_id=children.id))

    outreach = Category(name="Outreach")
    db.add(outreach)
    db.flush()
    db.add(Department(name="Evangelism Team", category_id=outreach.id))
    db.add(Department(name="Community Service", category_id=outreach.id))

    # Create uncategorized departments
    db.add(Department(name="Ushering"))
    db.add(Department(name="Prayer Team"))
    db.add(Department(name="Hospitality"))

    db.commit()

    return {"message": "Database seeded successfully", "seeded": True}


# ============ STATS ============

@router.get("/stats/departments")
def get_department_stats(db: Session = Depends(get_db)):
    """Get member count per department, grouped by category"""
    # Get all departments with their member counts
    departments = db.query(Department).options(
        joinedload(Department.category)
    ).all()

    # Count members per department
    dept_counts = {}
    member_depts = db.query(MemberDepartment).all()
    for md in member_depts:
        dept_counts[md.department_id] = dept_counts.get(md.department_id, 0) + 1

    # Group by category
    categories_data = {}
    uncategorized = []

    for dept in departments:
        count = dept_counts.get(dept.id, 0)
        dept_info = {
            "id": dept.id,
            "name": dept.name,
            "memberCount": count
        }

        if dept.category_id:
            if dept.category_id not in categories_data:
                categories_data[dept.category_id] = {
                    "id": dept.category.id,
                    "name": dept.category.name,
                    "departments": []
                }
            categories_data[dept.category_id]["departments"].append(dept_info)
        else:
            uncategorized.append(dept_info)

    # Sort departments by member count (descending)
    for cat_id in categories_data:
        categories_data[cat_id]["departments"].sort(key=lambda x: x["memberCount"], reverse=True)
    uncategorized.sort(key=lambda x: x["memberCount"], reverse=True)

    return {
        "categories": list(categories_data.values()),
        "uncategorized": uncategorized
    }


@router.get("/stats/departments/{department_id}")
def get_department_members(department_id: int, db: Session = Depends(get_db)):
    """Get all members who selected a specific department"""
    department = db.query(Department).filter(Department.id == department_id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    # Get members who selected this department
    member_depts = db.query(MemberDepartment).filter(
        MemberDepartment.department_id == department_id
    ).all()

    member_ids = [md.member_id for md in member_depts]
    members = db.query(Member).filter(Member.id.in_(member_ids)).order_by(Member.full_name).all()

    return {
        "department": {
            "id": department.id,
            "name": department.name,
            "categoryId": department.category_id
        },
        "members": [
            {
                "id": m.id,
                "fullName": m.full_name,
                "phone": m.phone,
                "email": m.email,
                "address": m.address
            }
            for m in members
        ],
        "totalCount": len(members)
    }


# ============ EXPORT ============

def sanitize_sheet_name(name: str) -> str:
    """Sanitize string for Excel sheet name"""
    # Remove invalid characters
    sanitized = re.sub(r'[\[\]:*?/\\]', '', name)
    # Limit to 31 characters
    return sanitized[:31]


@router.get("/export")
def export_data(
    type: str = Query("department"),
    approved_only: bool = Query(False),
    db: Session = Depends(get_db)
):
    """Export data to Excel file. Use approved_only=true to export only approved selections."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    today = datetime.now().strftime("%Y-%m-%d")
    suffix = "-approved" if approved_only else ""

    if type == "member":
        # Export by member
        ws = wb.active
        ws.title = "Members"

        # Get all departments for headers
        all_departments = db.query(Department).order_by(Department.name).all()
        dept_names = [d.name for d in all_departments]
        dept_ids = [d.id for d in all_departments]

        # Headers
        headers = ["Full Name", "Phone", "Email", "Address", "Submitted On"] + dept_names
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Get all members
        members = db.query(Member).options(
            joinedload(Member.departments)
        ).order_by(Member.created_at.desc()).all()

        for row, member in enumerate(members, 2):
            ws.cell(row=row, column=1, value=member.full_name)
            ws.cell(row=row, column=2, value=member.phone)
            ws.cell(row=row, column=3, value=member.email)
            ws.cell(row=row, column=4, value=member.address)
            ws.cell(row=row, column=5, value=member.created_at.strftime("%Y-%m-%d %H:%M") if member.created_at else "")

            # Filter by status if approved_only
            if approved_only:
                member_dept_ids = {md.department_id for md in member.departments if md.status == "approved"}
            else:
                member_dept_ids = {md.department_id for md in member.departments}

            for i, dept_id in enumerate(dept_ids):
                ws.cell(row=row, column=6 + i, value="Yes" if dept_id in member_dept_ids else "")

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="Department")
        ws_summary.cell(row=1, column=2, value="Member Count")

        for row, dept in enumerate(all_departments, 2):
            if approved_only:
                count = db.query(MemberDepartment).filter(
                    MemberDepartment.department_id == dept.id,
                    MemberDepartment.status == "approved"
                ).count()
            else:
                count = db.query(MemberDepartment).filter(MemberDepartment.department_id == dept.id).count()
            ws_summary.cell(row=row, column=1, value=dept.name)
            ws_summary.cell(row=row, column=2, value=count)

        filename = f"members-report{suffix}-{today}.xlsx"

    else:
        # Export by department
        departments = db.query(Department).options(
            joinedload(Department.category),
            joinedload(Department.member_departments).joinedload(MemberDepartment.member)
        ).order_by(Department.name).all()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        ws.cell(row=1, column=1, value="Department")
        ws.cell(row=1, column=2, value="Category")
        ws.cell(row=1, column=3, value="Member Count")

        for row, dept in enumerate(departments, 2):
            if approved_only:
                filtered_mds = [md for md in dept.member_departments if md.status == "approved"]
            else:
                filtered_mds = dept.member_departments
            ws.cell(row=row, column=1, value=dept.name)
            ws.cell(row=row, column=2, value=dept.category.name if dept.category else "Uncategorized")
            ws.cell(row=row, column=3, value=len(filtered_mds))

        # One sheet per department
        for dept in departments:
            if approved_only:
                filtered_mds = [md for md in dept.member_departments if md.status == "approved"]
            else:
                filtered_mds = dept.member_departments

            sheet_name = sanitize_sheet_name(dept.name)
            ws_dept = wb.create_sheet(sheet_name)

            # Header info
            ws_dept.cell(row=1, column=1, value="Department:")
            ws_dept.cell(row=1, column=2, value=dept.name)
            ws_dept.cell(row=2, column=1, value="Category:")
            ws_dept.cell(row=2, column=2, value=dept.category.name if dept.category else "Uncategorized")
            ws_dept.cell(row=3, column=1, value="Total Members:")
            ws_dept.cell(row=3, column=2, value=len(filtered_mds))

            # Column headers
            ws_dept.cell(row=5, column=1, value="Full Name")
            ws_dept.cell(row=5, column=2, value="Phone")
            ws_dept.cell(row=5, column=3, value="Email")
            ws_dept.cell(row=5, column=4, value="Address")
            ws_dept.cell(row=5, column=5, value="Submitted On")

            # Members
            for row, md in enumerate(filtered_mds, 6):
                ws_dept.cell(row=row, column=1, value=md.member.full_name)
                ws_dept.cell(row=row, column=2, value=md.member.phone)
                ws_dept.cell(row=row, column=3, value=md.member.email)
                ws_dept.cell(row=row, column=4, value=md.member.address)
                ws_dept.cell(row=row, column=5, value=md.member.created_at.strftime("%Y-%m-%d %H:%M") if md.member.created_at else "")

            # Set column widths
            for col in range(1, 6):
                ws_dept.column_dimensions[get_column_letter(col)].width = 20

        filename = f"departments-report{suffix}-{today}.xlsx"

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ ADMIN APPROVAL ENDPOINTS ============

@router.get("/admin/reviews")
def get_all_reviews(db: Session = Depends(get_db)):
    """Get all members with their department selections for review"""
    members = db.query(Member).options(
        joinedload(Member.departments).joinedload(MemberDepartment.department).joinedload(Department.category)
    ).order_by(Member.created_at.desc()).all()

    result = []
    for m in members:
        selections = []
        for md in m.departments:
            selections.append({
                "id": md.id,
                "member_id": md.member_id,
                "department_id": md.department_id,
                "department_name": md.department.name,
                "category_name": md.department.category.name if md.department.category else None,
                "source": md.source or "member",
                "status": md.status or "pending",
                "admin_note": md.admin_note,
                "created_at": md.created_at.isoformat() if md.created_at else None,
                "status_changed_at": md.status_changed_at.isoformat() if md.status_changed_at else None,
                "replaced_by_id": md.replaced_by_id
            })

        result.append({
            "id": m.id,
            "full_name": m.full_name,
            "phone": m.phone,
            "email": m.email,
            "address": m.address,
            "created_at": m.created_at.isoformat() if m.created_at else None,
            "selections": selections
        })

    return result


@router.put("/admin/reviews/{member_department_id}")
def update_review_status(
    member_department_id: int,
    data: ReviewStatusUpdate,
    request: Request = None,
    db: Session = Depends(get_db)
):
    """Approve or reject a single department selection"""
    md = db.query(MemberDepartment).options(
        joinedload(MemberDepartment.member),
        joinedload(MemberDepartment.department).joinedload(Department.category)
    ).filter(MemberDepartment.id == member_department_id).first()
    if not md:
        raise HTTPException(status_code=404, detail="Selection not found")

    if data.status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Status must be 'approved' or 'rejected'")

    md.status = data.status
    md.admin_note = data.admin_note
    md.status_changed_at = datetime.now()
    action = "review_approve" if data.status == "approved" else "review_reject"
    db.commit()

    # Status change might add or remove from approved bucket; resync.
    api_sync = _sync_member_departments_to_central(md.member_id, db)
    if request:
        log_msg = f"{'Approved' if data.status == 'approved' else 'Rejected'} {md.member.full_name} for {md.department.name}"
        if api_sync["attempted"]:
            log_msg += f" (central: {'synced ' + str(api_sync['department_count']) + ' depts' if api_sync['ok'] else 'failed - ' + (api_sync['error'] or '?')})"
        _log_admin_action(request, db, action, "member_department", member_department_id, log_msg)
        db.commit()

    # Dispatch notification
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        event_type = EventType.MEMBER_APPROVED if data.status == "approved" else EventType.MEMBER_REJECTED
        dispatch_event(db, event_type, {
            "member_id": md.member.id,
            "member_name": md.member.full_name,
            "member_email": md.member.email,
            "department_name": md.department.name,
            "category_name": md.department.category.name if md.department.category else None,
            "admin_note": data.admin_note
        })
    except Exception as e:
        print(f"Failed to dispatch notification: {e}")

    return {"success": True, "id": member_department_id, "status": data.status, "api_sync": api_sync}


@router.post("/admin/reviews/{member_department_id}/replace")
def replace_department(
    member_department_id: int,
    data: ReplaceDepartmentRequest,
    request: Request,
    db: Session = Depends(get_db)
):
    """Replace a selection with a different department (reject original, create admin-assigned)"""
    md = db.query(MemberDepartment).filter(MemberDepartment.id == member_department_id).first()
    if not md:
        raise HTTPException(status_code=404, detail="Selection not found")

    # Check new department exists
    new_dept = db.query(Department).filter(Department.id == data.new_department_id).first()
    if not new_dept:
        raise HTTPException(status_code=404, detail="New department not found")

    # Reject the original
    md.status = "rejected"
    md.admin_note = data.admin_note or f"Replaced with {new_dept.name}"
    md.status_changed_at = datetime.now()

    # Create new admin-assigned selection
    new_md = MemberDepartment(
        member_id=md.member_id,
        department_id=data.new_department_id,
        source="admin",
        status="approved",
        admin_note=f"Replacement for {md.department.name}",
        status_changed_at=datetime.now()
    )
    db.add(new_md)
    db.flush()

    # Link original to replacement
    md.replaced_by_id = new_md.id
    db.commit()

    _log_admin_action(request, db, "replace_department", "member_department", member_department_id,
                      f"Replaced '{md.department.name}' with '{new_dept.name}' for member #{md.member_id}")
    db.commit()

    return {"success": True, "original_id": member_department_id, "new_id": new_md.id}


@router.post("/admin/members")
def admin_create_member(data: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    """Admin: create a new member without requiring department selections"""
    full_name = _title_case_name((data.get("full_name") or "").strip())
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    address = (data.get("address") or "").strip()

    if not full_name or not phone:
        raise HTTPException(status_code=400, detail="Full name and phone are required")

    # Validate phone format (10 digits)
    if not validate_phone(phone):
        raise HTTPException(status_code=400, detail="Phone number must be 10 digits (e.g., 0711234456)")

    # Check if member with same phone and name already exists
    existing = db.query(Member).filter(
        Member.phone == phone,
        func.lower(Member.full_name) == full_name.lower()
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="A member with this name and phone already exists")

    member = Member(
        full_name=full_name,
        phone=phone,
        email=email,
        address=address
    )
    db.add(member)
    db.commit()
    db.refresh(member)

    if request:
        _log_admin_action(request, db, "create_member", "member", member.id, f"Created member '{member.full_name}' ({member.phone})")
        db.commit()

    return {
        "id": member.id,
        "fullName": member.full_name,
        "phone": member.phone,
        "email": member.email,
        "address": member.address
    }


@router.put("/admin/members/{member_id}")
def admin_update_member(member_id: int, data: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    """Admin: update a member's personal details (name, phone, email, address).
    If the member is linked to a central rfm-database record, the same changes
    are propagated up so the central directory stays in sync. Local saves
    always succeed even if the API push fails (e.g. central DB unreachable);
    failure is reported in the response so admin knows."""
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Capture before-state so we know what actually changed
    before = {
        "full_name": member.full_name,
        "phone": member.phone,
        "email": member.email or "",
        "address": member.address or "",
    }

    if "full_name" in data and data["full_name"].strip():
        member.full_name = _title_case_name(data["full_name"].strip())
    if "phone" in data and data["phone"].strip():
        # Check for phone conflict
        existing = db.query(Member).filter(Member.phone == data["phone"].strip(), Member.id != member_id).first()
        if existing:
            # Allow shared phones (family members)
            pass
        member.phone = data["phone"].strip()
    if "email" in data:
        member.email = (data["email"] or "").strip()
    if "address" in data:
        member.address = (data["address"] or "").strip()

    db.commit()
    db.refresh(member)

    # ---- Central DB propagation ----
    # Only attempt when the member is linked AND integration is on. Build the
    # PATCH payload from fields that actually changed; fall through with
    # api_push details so the UI can show "✓ also synced to central" or
    # surface failures without blocking the local save.
    api_push = {"attempted": False, "ok": False, "fields": [], "error": None}
    try:
        if member.external_member_id and _rfm.is_enabled(db) and _rfm.is_configured(db):
            payload = {}
            if before["full_name"] != member.full_name:
                first, last = _rfm.derive_first_last_from_full(member.full_name)
                if first:
                    payload["first_name"] = first
                if last:
                    payload["last_name"] = last
            if before["phone"] != member.phone:
                canonical = _rfm.to_sa_canonical_mobile(member.phone or "")
                if canonical:
                    payload["phone"] = canonical
                # If new phone isn't a valid SA mobile we silently skip the push
                # for that field — local save still happened.
            if before["email"] != (member.email or ""):
                v = (member.email or "").strip()
                if not v:
                    # Allow clearing — but only push if the request actually
                    # sent an empty string (vs not sending the field at all).
                    if "email" in data:
                        payload["email"] = ""
                elif "@" in v and "." in v.split("@")[-1]:
                    payload["email"] = v
            if before["address"] != (member.address or ""):
                # Local has a single string; the API splits across 4 fields.
                # We just push the whole thing into physical_address since we
                # don't have structured suburb/city/postal from the local UI.
                payload["physical_address"] = member.address or ""

            if payload:
                api_push["attempted"] = True
                api_push["fields"] = list(payload.keys())
                push_result = _rfm.update_member(member.external_member_id, payload, db=db)
                api_push["ok"] = push_result.ok
                if not push_result.ok:
                    api_push["error"] = push_result.error
    except Exception as exc:
        api_push["error"] = f"unexpected: {exc}"

    if request:
        log_msg = f"Updated member '{member.full_name}'"
        if api_push["attempted"]:
            if api_push["ok"]:
                log_msg += f" (synced to central: {', '.join(api_push['fields'])})"
            else:
                log_msg += f" (central sync failed: {api_push['error']})"
        _log_admin_action(request, db, "update_member", "member", member_id, log_msg)
        db.commit()

    return {
        "id": member.id,
        "full_name": member.full_name,
        "phone": member.phone,
        "email": member.email,
        "address": member.address,
        "api_sync": api_push,
    }


@router.delete("/admin/members/{member_id}")
def admin_delete_member(member_id: int, request: Request = None, db: Session = Depends(get_db)):
    """Admin permanently deletes a member and all related records"""
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    member_name = member.full_name
    if request:
        _log_admin_action(request, db, "delete_member", "member", member_id, f"Permanently deleted member {member_name}")

    # Clear HOD references on departments
    db.query(Department).filter(Department.hod_member_id == member_id).update(
        {"hod_member_id": None}, synchronize_session=False
    )

    # Clear SET NULL references to avoid FK issues with SQLAlchemy
    from models import ServiceProgram, ServiceSchedule, Meeting, NotificationLog, PosterRequest
    db.query(ServiceProgram).filter(ServiceProgram.created_by_member_id == member_id).update(
        {"created_by_member_id": None}, synchronize_session=False
    )
    db.query(ServiceSchedule).filter(ServiceSchedule.service_manager_id == member_id).update(
        {"service_manager_id": None}, synchronize_session=False
    )
    db.query(Meeting).filter(Meeting.created_by_id == member_id).update(
        {"created_by_id": None}, synchronize_session=False
    )
    db.query(NotificationLog).filter(NotificationLog.recipient_id == member_id).update(
        {"recipient_id": None}, synchronize_session=False
    )
    db.query(PosterRequest).filter(PosterRequest.acknowledged_by_id == member_id).update(
        {"acknowledged_by_id": None}, synchronize_session=False
    )

    # Delete the member (CASCADE handles member_departments, appeals, meeting_rsvps, poster_requests)
    db.delete(member)
    db.commit()

    return {"success": True, "message": f"Member '{member_name}' has been permanently deleted"}


@router.post("/admin/members/{member_id}/assign")
def assign_department(
    member_id: int,
    data: AssignDepartmentRequest,
    request: Request = None,
    db: Session = Depends(get_db)
):
    """Admin assigns an additional department to a member"""
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    dept = db.query(Department).options(
        joinedload(Department.category)
    ).filter(Department.id == data.department_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    # Check if already assigned
    existing = db.query(MemberDepartment).filter(
        MemberDepartment.member_id == member_id,
        MemberDepartment.department_id == data.department_id
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Department already assigned to this member")

    # Create admin-assigned selection
    md = MemberDepartment(
        member_id=member_id,
        department_id=data.department_id,
        source="admin",
        status="approved",
        admin_note=data.admin_note,
        status_changed_at=datetime.now()
    )
    db.add(md)
    db.commit()
    db.refresh(md)

    # Propagate to central (best-effort) — admin-assigned == approved status
    api_sync = _sync_member_departments_to_central(member_id, db)
    if request:
        log_msg = f"Assigned {dept.name} to {member.full_name}"
        if api_sync["attempted"]:
            log_msg += f" (central: {'synced ' + str(api_sync['department_count']) + ' depts' if api_sync['ok'] else 'failed - ' + (api_sync['error'] or '?')})"
            if api_sync["skipped_unlinked"]:
                log_msg += f"; {api_sync['skipped_unlinked']} dept(s) not yet linked"
        _log_admin_action(request, db, "assign_department", "member", member_id, log_msg)
        db.commit()

    # Dispatch notification
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        dispatch_event(db, EventType.DEPARTMENT_ASSIGNED, {
            "member_id": member.id,
            "member_name": member.full_name,
            "member_email": member.email,
            "department_name": dept.name,
            "category_name": dept.category.name if dept.category else None,
            "admin_note": data.admin_note
        })
    except Exception as e:
        print(f"Failed to dispatch notification: {e}")

    return {"success": True, "id": md.id, "api_sync": api_sync}


@router.post("/admin/reviews/bulk-approve")
def bulk_approve_pending(request: Request = None, db: Session = Depends(get_db)):
    """Approve all pending selections (including null status from before workflow)"""
    count = db.query(MemberDepartment).filter(
        or_(MemberDepartment.status == "pending", MemberDepartment.status.is_(None))
    ).update({
        MemberDepartment.status: "approved",
        MemberDepartment.status_changed_at: datetime.now()
    }, synchronize_session='fetch')
    if request:
        _log_admin_action(request, db, "bulk_approve", "member_department", None, f"Bulk approved {count} pending selections")
    db.commit()

    return {"success": True, "approved_count": count}


# ============ LEADERSHIP ROLES ============

@router.get("/admin/members/{member_id}/leadership-roles")
def get_member_leadership_roles(member_id: int, db: Session = Depends(get_db)):
    """Get leadership roles for a member"""
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    roles = []
    if member.leadership_roles:
        try:
            roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else member.leadership_roles
        except (ValueError, TypeError):
            roles = []

    # Check if member is HOD of any department
    hod_depts = db.query(Department).filter(Department.hod_member_id == member_id).all()
    is_hod = len(hod_depts) > 0

    return {
        "member_id": member_id,
        "roles": roles,
        "is_hod": is_hod,
        "hod_departments": [{"id": d.id, "name": d.name} for d in hod_depts]
    }


@router.put("/admin/members/{member_id}/leadership-roles")
def update_member_leadership_roles(
    member_id: int,
    roles: List[str] = Body(..., embed=True),
    request: Request = None,
    db: Session = Depends(get_db)
):
    """Update leadership roles for a member (deacon, elder, etc.)"""
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Validate roles
    valid_roles = ["pastor", "deacon", "elder", "service_manager", "admin", "dr", "mr", "mrs"]  # HOD is derived from departments
    invalid = [r for r in roles if r not in valid_roles]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Invalid roles: {invalid}. Valid roles are: {valid_roles}")

    if request:
        _log_admin_action(request, db, "update_leadership_roles", "member", member_id,
            f"Updated roles for {member.full_name}: {', '.join(roles) if roles else 'none'}")
    member.leadership_roles = json.dumps(roles) if roles else None
    db.commit()

    return {"success": True, "roles": roles}


@router.get("/admin/leadership-roles")
def get_all_leadership_roles(db: Session = Depends(get_db)):
    """Get all members with their leadership roles for management"""
    members = db.query(Member).all()

    result = []
    for member in members:
        roles = []
        if member.leadership_roles:
            try:
                roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else member.leadership_roles
            except (ValueError, TypeError):
                roles = []

        # Check if member is HOD
        hod_depts = db.query(Department).filter(Department.hod_member_id == member.id).all()
        if hod_depts:
            roles = ["hod"] + roles  # HOD comes first

        result.append({
            "id": member.id,
            "full_name": member.full_name,
            "phone": member.phone,
            "email": member.email,
            "roles": roles,
            "hod_departments": [{"id": d.id, "name": d.name} for d in hod_depts]
        })

    # Sort by name
    result.sort(key=lambda x: x["full_name"])

    return {"members": result}


# ============ PUBLISH ENDPOINTS ============

@router.get("/admin/preview")
def preview_publish(db: Session = Depends(get_db)):
    """Preview what members will see after publishing"""
    members = db.query(Member).options(
        joinedload(Member.departments).joinedload(MemberDepartment.department)
    ).all()

    # Include null status as "pending" (for records created before approval workflow)
    pending_count = db.query(MemberDepartment).filter(
        or_(MemberDepartment.status == "pending", MemberDepartment.status.is_(None))
    ).count()
    total_approved = db.query(MemberDepartment).filter(MemberDepartment.status == "approved").count()

    members_preview = []
    for m in members:
        approved_depts = [
            md.department.name
            for md in m.departments
            if md.status == "approved"
        ]
        if approved_depts:
            members_preview.append({
                "id": m.id,
                "full_name": m.full_name,
                "phone": m.phone,
                "approved_departments": approved_depts
            })

    return {
        "total_members": len(members_preview),
        "total_approved_assignments": total_approved,
        "pending_count": pending_count,
        "members_preview": members_preview
    }


@router.post("/admin/publish")
def publish_results(request: Request = None, db: Session = Depends(get_db)):
    """Publish results - make approved selections visible to members"""
    if request:
        _log_admin_action(request, db, "publish_results", "settings", None, "Published department selection results")
    now = datetime.now().isoformat()

    # Update resultsPublished setting
    setting = db.query(Settings).filter(Settings.key == "resultsPublished").first()
    if setting:
        setting.value = "true"
    else:
        db.add(Settings(key="resultsPublished", value="true"))

    # Update publishedAt
    pub_setting = db.query(Settings).filter(Settings.key == "publishedAt").first()
    if pub_setting:
        pub_setting.value = now
    else:
        db.add(Settings(key="publishedAt", value=now))

    db.commit()

    # Dispatch notification to all members with approved selections
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        # Get all members with at least one approved selection
        members_with_approved = db.query(Member).join(MemberDepartment).filter(
            MemberDepartment.status == "approved"
        ).distinct().all()

        # Get year
        year_setting = db.query(Settings).filter(Settings.key == "selectionYear").first()
        year = year_setting.value if year_setting else "2026"

        # Build recipients list
        recipients = [
            {"id": m.id, "email": m.email, "phone": m.phone}
            for m in members_with_approved
            if m.email  # Only include members with email
        ]

        if recipients:
            dispatch_event(db, EventType.RESULTS_PUBLISHED, {
                "year": year,
                "recipients": recipients
            })
    except Exception as e:
        print(f"Failed to dispatch notification: {e}")

    return {"success": True, "published_at": now}


@router.post("/admin/unpublish")
def unpublish_results(request: Request = None, db: Session = Depends(get_db)):
    """Unpublish results - hide from members"""
    if request:
        _log_admin_action(request, db, "unpublish_results", "settings", None, "Unpublished department selection results")
    setting = db.query(Settings).filter(Settings.key == "resultsPublished").first()
    if setting:
        setting.value = "false"
    db.commit()

    return {"success": True}


# ============ MEMBER RESULTS ENDPOINT ============

@router.get("/results")
def get_member_results(phone: str = Query(...), db: Session = Depends(get_db)):
    """Lookup results by phone number - returns all members with that phone (for families)"""
    # Check if results are published
    pub_setting = db.query(Settings).filter(Settings.key == "resultsPublished").first()
    is_published = pub_setting and pub_setting.value == "true"

    # Check appeal window
    appeal_setting = db.query(Settings).filter(Settings.key == "appealWindowOpen").first()
    appeal_open = appeal_setting and appeal_setting.value == "true"

    # Get year
    year_setting = db.query(Settings).filter(Settings.key == "selectionYear").first()
    year = year_setting.value if year_setting else "2026"

    # Normalize phone
    normalized = phone.strip().replace(" ", "").replace("-", "")

    # Find ALL members with this phone number (family members)
    all_members = db.query(Member).options(
        joinedload(Member.departments).joinedload(MemberDepartment.department).joinedload(Department.category)
    ).all()

    matching_members = []
    for m in all_members:
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            matching_members.append(m)

    if not matching_members:
        return {
            "published": is_published,
            "appeal_window_open": appeal_open,
            "year": year,
            "members": [],
            "message": "No submission found for this phone number."
        }

    # Build member data with all their selections
    members_data = []
    for member in matching_members:
        # Get all department selections with full status info
        selections = []
        for md in member.departments:
            selections.append({
                "id": md.id,
                "department_id": md.department.id,
                "department_name": md.department.name,
                "category_name": md.department.category.name if md.department.category else None,
                "status": md.status or "pending",
                "source": md.source or "member",  # "member" or "admin"
                "admin_note": md.admin_note,
                "created_at": md.created_at.isoformat() if md.created_at else None
            })

        # Separate by status for convenience
        approved = [s for s in selections if s["status"] == "approved"]
        pending = [s for s in selections if s["status"] == "pending" or s["status"] is None]
        rejected = [s for s in selections if s["status"] == "rejected"]

        # Check for admin-added departments that haven't been accepted (source=admin)
        admin_added = [s for s in selections if s["source"] == "admin"]

        # Parse leadership roles
        member_leadership_roles = []
        if member.leadership_roles:
            try:
                member_leadership_roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else member.leadership_roles
            except (ValueError, TypeError):
                member_leadership_roles = []

        is_hc_leader = db.query(HomeChurch).filter(HomeChurch.leader_member_id == member.id).count() > 0

        members_data.append({
            "id": member.id,
            "full_name": member.full_name,
            "email": member.email,
            "leadership_roles": member_leadership_roles,
            "all_selections": selections,
            "approved_departments": approved,
            "pending_departments": pending,
            "rejected_departments": rejected,
            "admin_added_departments": admin_added,
            "is_hc_leader": is_hc_leader,
        })

    return {
        "published": is_published,
        "appeal_window_open": appeal_open,
        "year": year,
        "members": members_data,
        "is_family": len(members_data) > 1
    }


@router.post("/results/accept/{member_department_id}")
def accept_admin_assignment(
    member_department_id: int,
    request: Request,
    phone: str = Query(...),
    db: Session = Depends(get_db)
):
    """Member accepts an admin-added department assignment"""
    # Find the member department
    md = db.query(MemberDepartment).options(
        joinedload(MemberDepartment.member)
    ).filter(MemberDepartment.id == member_department_id).first()

    if not md:
        raise HTTPException(status_code=404, detail="Selection not found")

    # Verify phone matches
    normalized_input = phone.strip().replace(" ", "").replace("-", "")
    normalized_member = md.member.phone.strip().replace(" ", "").replace("-", "")

    if normalized_input != normalized_member:
        raise HTTPException(status_code=403, detail="Phone number does not match")

    # Only allow accepting admin-added departments
    if md.source != "admin":
        raise HTTPException(status_code=400, detail="This selection was not added by admin")

    # Mark as accepted by updating source to "accepted" or adding a note
    md.admin_note = (md.admin_note or "") + " [Accepted by member]"
    md.status_changed_at = datetime.now()
    db.commit()

    _log_member_action(request, db, md.member, "accept_assignment", "member_department", member_department_id,
                       f"Accepted admin-assigned department")
    db.commit()

    return {"success": True, "message": "Department assignment accepted"}


# ============ APPEAL ENDPOINTS ============

@router.post("/appeals")
def submit_appeal(data: AppealCreate, request: Request, db: Session = Depends(get_db)):
    """Submit an appeal (public endpoint)"""
    # Check results are published
    pub_setting = db.query(Settings).filter(Settings.key == "resultsPublished").first()
    if not pub_setting or pub_setting.value != "true":
        raise HTTPException(status_code=400, detail="Results are not yet published")

    # Check appeal window is open
    appeal_setting = db.query(Settings).filter(Settings.key == "appealWindowOpen").first()
    if not appeal_setting or appeal_setting.value != "true":
        raise HTTPException(status_code=400, detail="Appeal window is currently closed")

    # Find member - by ID if provided (for families/info desk), otherwise by phone
    member = None
    if data.member_id:
        member = db.query(Member).filter(Member.id == data.member_id).first()
        # Verify phone matches for security
        if member:
            normalized_input = data.phone.strip().replace(" ", "").replace("-", "")
            normalized_member = member.phone.strip().replace(" ", "").replace("-", "")
            if normalized_input != normalized_member:
                raise HTTPException(status_code=403, detail="Phone number does not match member")
    else:
        # Find member by phone only
        normalized = data.phone.strip().replace(" ", "").replace("-", "")
        members = db.query(Member).all()
        for m in members:
            m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
            if m_normalized == normalized or m.phone == data.phone:
                member = m
                break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Get department names for notification
    unwanted_dept = None
    wanted_dept = None
    if data.unwanted_department_id:
        unwanted_dept = db.query(Department).filter(Department.id == data.unwanted_department_id).first()
    if data.wanted_department_id:
        wanted_dept = db.query(Department).filter(Department.id == data.wanted_department_id).first()

    # Create appeal
    appeal = Appeal(
        member_id=member.id,
        unwanted_department_id=data.unwanted_department_id,
        wanted_department_id=data.wanted_department_id,
        reason=data.reason,
        status="pending"
    )
    db.add(appeal)
    db.flush()
    unwanted_name = unwanted_dept.name if unwanted_dept else "none"
    wanted_name = wanted_dept.name if wanted_dept else "none"
    _log_member_action(request, db, member, "submit_appeal", "appeal", appeal.id,
        f"Appeal: remove {unwanted_name}, want {wanted_name}")
    db.commit()
    db.refresh(appeal)

    # Dispatch notification to admin
    try:
        from notifications.dispatcher import dispatch_to_admins
        from notifications.events import EventType

        dispatch_to_admins(db, EventType.APPEAL_SUBMITTED, {
            "appeal_id": appeal.id,
            "member_name": member.full_name,
            "member_phone": member.phone,
            "unwanted_department": unwanted_dept.name if unwanted_dept else None,
            "wanted_department": wanted_dept.name if wanted_dept else None,
            "reason": data.reason
        })
    except Exception as e:
        print(f"Failed to dispatch notification: {e}")

    return {"success": True, "appeal_id": appeal.id}


@router.get("/admin/appeals")
def get_all_appeals(db: Session = Depends(get_db)):
    """Get all appeals for admin review"""
    appeals = db.query(Appeal).options(
        joinedload(Appeal.member),
        joinedload(Appeal.unwanted_department),
        joinedload(Appeal.wanted_department)
    ).order_by(Appeal.created_at.desc()).all()

    return [
        {
            "id": a.id,
            "member_id": a.member_id,
            "member_name": a.member.full_name,
            "member_phone": a.member.phone,
            "unwanted_department_id": a.unwanted_department_id,
            "unwanted_department_name": a.unwanted_department.name if a.unwanted_department else None,
            "wanted_department_id": a.wanted_department_id,
            "wanted_department_name": a.wanted_department.name if a.wanted_department else None,
            "reason": a.reason,
            "status": a.status,
            "admin_response": a.admin_response,
            "created_at": a.created_at.isoformat() if a.created_at else None,
            "resolved_at": a.resolved_at.isoformat() if a.resolved_at else None
        }
        for a in appeals
    ]


@router.put("/admin/appeals/{appeal_id}")
def resolve_appeal(
    appeal_id: int,
    data: AppealResolve,
    request: Request = None,
    db: Session = Depends(get_db)
):
    """Resolve an appeal (approve or reject)"""
    appeal = db.query(Appeal).options(
        joinedload(Appeal.member),
        joinedload(Appeal.unwanted_department),
        joinedload(Appeal.wanted_department)
    ).filter(Appeal.id == appeal_id).first()
    if not appeal:
        raise HTTPException(status_code=404, detail="Appeal not found")

    if data.status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Status must be 'approved' or 'rejected'")

    appeal.status = data.status
    appeal.admin_response = data.admin_response
    appeal.resolved_at = datetime.now()
    if request:
        _log_admin_action(request, db, "resolve_appeal", "appeal", appeal_id,
            f"{'Approved' if data.status == 'approved' else 'Rejected'} appeal from {appeal.member.full_name}")

    # If approved, update the member's departments
    if data.status == "approved":
        # Remove unwanted department if specified
        if appeal.unwanted_department_id:
            md = db.query(MemberDepartment).filter(
                MemberDepartment.member_id == appeal.member_id,
                MemberDepartment.department_id == appeal.unwanted_department_id,
                MemberDepartment.status == "approved"
            ).first()
            if md:
                md.status = "rejected"
                md.admin_note = "Removed via approved appeal"
                md.status_changed_at = datetime.now()

        # Add wanted department if specified
        if appeal.wanted_department_id:
            # Check not already assigned
            existing = db.query(MemberDepartment).filter(
                MemberDepartment.member_id == appeal.member_id,
                MemberDepartment.department_id == appeal.wanted_department_id
            ).first()
            if not existing:
                new_md = MemberDepartment(
                    member_id=appeal.member_id,
                    department_id=appeal.wanted_department_id,
                    source="admin",
                    status="approved",
                    admin_note="Added via approved appeal",
                    status_changed_at=datetime.now()
                )
                db.add(new_md)

    db.commit()

    # Dispatch notification to member
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        dispatch_event(db, EventType.APPEAL_RESOLVED, {
            "member_id": appeal.member.id,
            "member_name": appeal.member.full_name,
            "member_email": appeal.member.email,
            "status": data.status,
            "unwanted_department": appeal.unwanted_department.name if appeal.unwanted_department else None,
            "wanted_department": appeal.wanted_department.name if appeal.wanted_department else None,
            "admin_response": data.admin_response
        })
    except Exception as e:
        print(f"Failed to dispatch notification: {e}")

    return {"success": True, "appeal_id": appeal_id, "status": data.status}


@router.post("/admin/appeals/window")
def toggle_appeal_window(open: bool = Query(...), request: Request = None, db: Session = Depends(get_db)):
    """Open or close the appeal window"""
    setting = db.query(Settings).filter(Settings.key == "appealWindowOpen").first()
    if setting:
        setting.value = "true" if open else "false"
    else:
        db.add(Settings(key="appealWindowOpen", value="true" if open else "false"))

    if request:
        _log_admin_action(request, db, "toggle_appeal_window", "settings", None,
            f"{'Opened' if open else 'Closed'} the appeal window")
    db.commit()

    return {"success": True, "appeal_window_open": open}


# ============ HOD / ELDER ENDPOINTS ============


def _find_member_by_phone(db: Session, phone: str):
    """Find a member by phone number (normalized)"""
    normalized = phone.strip().replace(" ", "").replace("-", "")
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            return m
    return None


def _is_elder(member: "Member") -> bool:
    """Check if a member has the elder or pastor leadership role"""
    roles = member.leadership_roles or []
    return "elder" in roles or "pastor" in roles


def _get_accessible_departments(db: Session, member: "Member"):
    """Get departments a member can manage meetings for (HOD depts + all depts if elder)"""
    if _is_elder(member):
        # Elders can access all departments
        return db.query(Department).options(
            joinedload(Department.category),
            joinedload(Department.member_departments).joinedload(MemberDepartment.member)
        ).order_by(Department.name).all()
    else:
        # HODs can only access their own departments
        return db.query(Department).options(
            joinedload(Department.category),
            joinedload(Department.member_departments).joinedload(MemberDepartment.member)
        ).filter(Department.hod_member_id == member.id).order_by(Department.name).all()


def _can_manage_department(db: Session, member: "Member", department_id: int) -> bool:
    """Check if member can manage a department (is HOD of it or is an Elder)"""
    if _is_elder(member):
        return True
    department = db.query(Department).filter(Department.id == department_id).first()
    return department and department.hod_member_id == member.id

@router.post("/admin/departments/{department_id}/set-hod")
def set_department_hod(
    department_id: int,
    data: SetHODRequest,
    request: Request,
    db: Session = Depends(get_db)
):
    """Assign a member as Head of Department"""
    department = db.query(Department).filter(Department.id == department_id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    member = db.query(Member).filter(Member.id == data.member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    department.hod_member_id = member.id
    db.commit()

    _log_admin_action(request, db, "set_hod", "department", department_id, f"Set {member.full_name} as HOD of '{department.name}'")
    db.commit()

    return {
        "success": True,
        "message": "HOD assigned",
        "department": department.name,
        "hod_name": member.full_name
    }


@router.delete("/admin/departments/{department_id}/remove-hod")
def remove_department_hod(
    department_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Remove the HOD assignment from a department"""
    department = db.query(Department).filter(Department.id == department_id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    department.hod_member_id = None
    db.commit()

    _log_admin_action(request, db, "remove_hod", "department", department_id, f"Removed HOD from '{department.name}'")
    db.commit()

    return {"success": True, "message": "HOD removed"}


@router.get("/hod/departments")
def get_hod_departments(phone: str = Query(...), db: Session = Depends(get_db)):
    """Get departments where this member is HOD, with member lists and statuses"""
    hod_member = _find_member_by_phone(db, phone)
    if not hod_member:
        raise HTTPException(status_code=404, detail="Member not found")

    is_elder = _is_elder(hod_member)

    # Always show only departments where member is actual HOD (not all for elders)
    # Elder access to all departments is only for meeting booking
    departments = db.query(Department).options(
        joinedload(Department.category),
        joinedload(Department.member_departments).joinedload(MemberDepartment.member)
    ).filter(Department.hod_member_id == hod_member.id).order_by(Department.name).all()

    if not departments:
        return {
            "hod_name": hod_member.full_name,
            "is_elder": is_elder,
            "role": "pastor" if "pastor" in (hod_member.leadership_roles or []) else ("elder" if is_elder else "hod"),
            "departments": []
        }

    dept_views = []
    for dept in departments:
        members_list = []
        approved_count = 0
        pending_count = 0
        rejected_count = 0

        for md in dept.member_departments:
            status = md.status or "pending"
            if status == "approved":
                approved_count += 1
            elif status == "rejected":
                rejected_count += 1
            else:
                pending_count += 1

            members_list.append({
                "id": md.member.id,
                "full_name": md.member.full_name,
                "phone": md.member.phone,
                "email": md.member.email,
                "status": status,
                "source": md.source or "member",
                "created_at": md.created_at.isoformat() if md.created_at else None
            })

        # Sort members: pending first, then approved, then rejected
        status_order = {"pending": 0, "approved": 1, "rejected": 2}
        members_list.sort(key=lambda x: (status_order.get(x["status"], 3), x["full_name"]))

        dept_views.append({
            "id": dept.id,
            "name": dept.name,
            "category_name": dept.category.name if dept.category else None,
            "members": members_list,
            "total_members": len(members_list),
            "approved_count": approved_count,
            "pending_count": pending_count,
            "rejected_count": rejected_count
        })

    return {
        "hod_name": hod_member.full_name,
        "is_elder": is_elder,
        "role": "pastor" if "pastor" in (hod_member.leadership_roles or []) else ("elder" if is_elder else "hod"),
        "departments": dept_views
    }


# ============ MEETING ENDPOINTS ============

def slot_to_time(slot: int) -> str:
    """Convert slot number (0-47) to HH:MM format"""
    hours = slot // 2
    minutes = (slot % 2) * 30
    return f"{hours:02d}:{minutes:02d}"


def get_week_dates(reference_date: date) -> Tuple[date, date]:
    """Get Monday and Sunday of the week containing reference_date"""
    # Get Monday of the week (weekday 0 = Monday)
    monday = reference_date - timedelta(days=reference_date.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def generate_recurring_dates(start_date: date, recurrence: str, end_date: date) -> List[date]:
    """
    Generate a list of dates for recurring meetings.

    Args:
        start_date: The first meeting date
        recurrence: One of 'daily', 'weekly', 'biweekly', 'monthly'
        end_date: The last possible date for meetings

    Returns:
        List of dates including the start_date
    """
    dates = [start_date]
    current = start_date

    # Limit to prevent runaway loops (max 52 occurrences)
    max_occurrences = 52

    while len(dates) < max_occurrences:
        if recurrence == 'daily':
            current = current + timedelta(days=1)
        elif recurrence == 'weekly':
            current = current + timedelta(weeks=1)
        elif recurrence == 'biweekly':
            current = current + timedelta(weeks=2)
        elif recurrence == 'monthly':
            # Add one month (handle month boundaries)
            month = current.month + 1
            year = current.year
            if month > 12:
                month = 1
                year += 1
            # Handle day overflow (e.g., Jan 31 -> Feb 28)
            day = min(current.day, 28)  # Safe for all months
            try:
                current = date(year, month, day)
            except ValueError:
                current = date(year, month, 28)
        else:
            break

        if current > end_date:
            break

        dates.append(current)

    return dates


def check_slot_availability(
    db: Session,
    department_id: int,
    meeting_date: date,
    start_slot: int,
    end_slot: int,
    exclude_meeting_id: Optional[int] = None
) -> Tuple[bool, Optional[str]]:
    """Check if time slots are available for a department's meeting.
    Returns (available, conflict_reason)"""

    # Find overlapping meetings on the same date
    query = db.query(Meeting).filter(
        Meeting.meeting_date == meeting_date,
        Meeting.start_slot < end_slot,
        Meeting.end_slot > start_slot
    )
    if exclude_meeting_id:
        query = query.filter(Meeting.id != exclude_meeting_id)

    conflicts = query.all()

    for meeting in conflicts:
        # Same department can't have overlapping meetings
        if meeting.department_id == department_id:
            return False, f"You already have a meeting scheduled at this time"

        # Get approved members of the booking department
        dept_members = set(
            md.member_id for md in db.query(MemberDepartment)
            .filter(
                MemberDepartment.department_id == department_id,
                MemberDepartment.status == "approved"
            ).all()
        )

        # Get approved members of the conflicting meeting's department
        other_members = set(
            md.member_id for md in db.query(MemberDepartment)
            .filter(
                MemberDepartment.department_id == meeting.department_id,
                MemberDepartment.status == "approved"
            ).all()
        )

        # Check for member intersection
        if dept_members & other_members:
            dept = db.query(Department).filter(Department.id == meeting.department_id).first()
            return False, f"Conflict with {dept.name if dept else 'another department'} - shared members"

    return True, None


def find_duplicate_meeting(
    db: Session,
    title: str,
    meeting_date: date,
    start_slot: int,
    end_slot: int,
    department_id: Optional[int] = None,
    target_department_ids: Optional[str] = None,
    target_member_ids: Optional[str] = None,
    target_leadership_roles: Optional[str] = None,
    is_general: int = 0,
    window_seconds: int = 120
) -> Optional[Meeting]:
    """Detect recent duplicate meeting.

    Returns an existing meeting created in the last `window_seconds` that matches
    on the dimensions that define a calendar event: title, date, slot range, and
    the target (department / multi-dept / individuals / roles / general). This
    catches double-click submissions and network retries without blocking
    legitimate identical meetings scheduled far apart.
    """
    from datetime import datetime, timezone, timedelta
    cutoff = datetime.now(timezone.utc) - timedelta(seconds=window_seconds)

    q = db.query(Meeting).filter(
        Meeting.title == title,
        Meeting.meeting_date == meeting_date,
        Meeting.start_slot == start_slot,
        Meeting.end_slot == end_slot,
        Meeting.created_at >= cutoff,
    )

    # Match on targeting dimension (all these fields are nullable in the schema)
    if department_id is not None:
        q = q.filter(Meeting.department_id == department_id)
    else:
        q = q.filter(Meeting.department_id.is_(None))

    if target_department_ids is not None:
        q = q.filter(Meeting.target_department_ids == target_department_ids)
    else:
        q = q.filter(Meeting.target_department_ids.is_(None))

    if target_member_ids is not None:
        q = q.filter(Meeting.target_member_ids == target_member_ids)
    else:
        q = q.filter(Meeting.target_member_ids.is_(None))

    if target_leadership_roles is not None:
        q = q.filter(Meeting.target_leadership_roles == target_leadership_roles)
    else:
        q = q.filter(Meeting.target_leadership_roles.is_(None))

    q = q.filter(Meeting.is_general == is_general)

    return q.order_by(Meeting.created_at.desc()).first()


def format_meeting_response(meeting: Meeting, db: Session, member_id: Optional[int] = None) -> dict:
    """Format a meeting for API response"""
    # Get RSVP counts
    rsvp_count = db.query(MeetingRSVP).filter(MeetingRSVP.meeting_id == meeting.id).count()
    attending_count = db.query(MeetingRSVP).filter(
        MeetingRSVP.meeting_id == meeting.id,
        MeetingRSVP.response == "attending"
    ).count()

    # Get member's RSVP if member_id provided
    my_rsvp = None
    if member_id:
        rsvp = db.query(MeetingRSVP).filter(
            MeetingRSVP.meeting_id == meeting.id,
            MeetingRSVP.member_id == member_id
        ).first()
        if rsvp:
            my_rsvp = rsvp.response

    # Parse target department IDs and get names
    target_dept_ids = None
    target_dept_names = None
    if meeting.target_department_ids:
        try:
            target_dept_ids = json.loads(meeting.target_department_ids) if isinstance(meeting.target_department_ids, str) else meeting.target_department_ids
            depts = db.query(Department).filter(Department.id.in_(target_dept_ids)).all()
            target_dept_names = [d.name for d in depts]
        except (ValueError, TypeError):
            pass

    # Parse target member IDs and get count
    target_member_ids = None
    target_member_count = 0
    if meeting.target_member_ids:
        try:
            target_member_ids = json.loads(meeting.target_member_ids) if isinstance(meeting.target_member_ids, str) else meeting.target_member_ids
            target_member_count = len(target_member_ids) if target_member_ids else 0
        except (ValueError, TypeError):
            pass

    # Parse target leadership roles
    target_leadership_roles = None
    if meeting.target_leadership_roles:
        try:
            target_leadership_roles = json.loads(meeting.target_leadership_roles) if isinstance(meeting.target_leadership_roles, str) else meeting.target_leadership_roles
        except (ValueError, TypeError):
            pass

    # Determine department name for display
    if meeting.is_general:
        dept_name = "All Leaders"
    elif target_leadership_roles:
        role_labels = {"hod": "HODs", "pastor": "Pastors", "deacon": "Deacons", "elder": "Elders", "service_manager": "Service Managers"}
        dept_name = ", ".join([role_labels.get(r, r.title()) for r in target_leadership_roles])
    elif target_member_ids:
        dept_name = f"{target_member_count} Selected Members"
    elif target_dept_ids:
        dept_name = ", ".join(target_dept_names) if target_dept_names else "Multiple Departments"
    elif meeting.department:
        dept_name = meeting.department.name
    else:
        dept_name = None

    return {
        "id": meeting.id,
        "department_id": meeting.department_id,
        "department_name": dept_name,
        "title": meeting.title,
        "description": meeting.description,
        "meeting_date": meeting.meeting_date.isoformat() if meeting.meeting_date else None,
        "start_slot": meeting.start_slot,
        "end_slot": meeting.end_slot,
        "start_time": slot_to_time(meeting.start_slot),
        "end_time": slot_to_time(meeting.end_slot),
        "location": meeting.location,
        "meeting_link": meeting.meeting_link,
        "created_by_id": meeting.created_by_id,
        "created_by_name": meeting.created_by.full_name if meeting.created_by else None,
        "created_at": meeting.created_at.isoformat() if meeting.created_at else None,
        "rsvp_count": rsvp_count,
        "attending_count": attending_count,
        "my_rsvp": my_rsvp,
        "is_general": bool(meeting.is_general),
        "target_department_ids": target_dept_ids,
        "target_department_names": target_dept_names,
        "target_leadership_roles": target_leadership_roles,
        "recurrence_group_id": meeting.recurrence_group_id
    }


# --- HOD Meeting Endpoints ---

@router.get("/hod/meetings")
def get_hod_meetings(phone: str = Query(...), db: Session = Depends(get_db)):
    """Get all meetings for departments where this member is HOD or Elder"""
    hod_member = _find_member_by_phone(db, phone)
    if not hod_member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Get accessible department IDs
    departments = _get_accessible_departments(db, hod_member)
    dept_ids = [d.id for d in departments]

    if not dept_ids:
        return {"meetings": [], "total": 0}

    # Get meetings for these departments
    meetings = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    ).filter(
        Meeting.department_id.in_(dept_ids),
        Meeting.meeting_date >= date.today()
    ).order_by(Meeting.meeting_date, Meeting.start_slot).all()

    return {
        "meetings": [format_meeting_response(m, db) for m in meetings],
        "total": len(meetings)
    }


@router.get("/hod/calendar")
def get_hod_calendar(
    phone: str = Query(...),
    department_id: int = Query(...),
    week_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Get weekly calendar with availability for a department"""
    hod_member = _find_member_by_phone(db, phone)
    if not hod_member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Verify HOD/Elder access to this department
    if not _can_manage_department(db, hod_member, department_id):
        raise HTTPException(status_code=403, detail="You do not have access to this department")

    department = db.query(Department).filter(Department.id == department_id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")

    # Parse week date or use today
    if week_date:
        try:
            ref_date = date.fromisoformat(week_date)
        except ValueError:
            ref_date = date.today()
    else:
        ref_date = date.today()

    week_start, week_end = get_week_dates(ref_date)

    # Get all meetings for this week
    all_meetings = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    ).filter(
        Meeting.meeting_date >= week_start,
        Meeting.meeting_date <= week_end
    ).all()

    # Build calendar
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    days = []

    for day_offset in range(7):
        current_date = week_start + timedelta(days=day_offset)
        day_meetings = [m for m in all_meetings if m.meeting_date == current_date]

        # Build slots (showing 6 AM to 10 PM = slots 12 to 44)
        slots = []
        for slot in range(12, 44):  # 6:00 to 22:00
            slot_time = slot_to_time(slot)

            # Check if this slot is taken by any meeting for this department
            dept_meeting = None
            for m in day_meetings:
                if m.department_id == department_id and m.start_slot <= slot < m.end_slot:
                    dept_meeting = m
                    break

            if dept_meeting:
                slots.append({
                    "slot": slot,
                    "time": slot_time,
                    "available": False,
                    "meeting": format_meeting_response(dept_meeting, db),
                    "conflict_reason": None
                })
            else:
                # Check availability (conflicts with other departments)
                available, conflict_reason = check_slot_availability(
                    db, department_id, current_date, slot, slot + 1
                )
                slots.append({
                    "slot": slot,
                    "time": slot_time,
                    "available": available,
                    "meeting": None,
                    "conflict_reason": conflict_reason
                })

        days.append({
            "date": current_date.isoformat(),
            "day_name": day_names[day_offset],
            "slots": slots
        })

    return {
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "department_id": department_id,
        "department_name": department.name,
        "days": days
    }


@router.post("/hod/meetings")
def create_hod_meeting(
    data: MeetingCreate,
    request: Request,
    phone: str = Query(...),
    db: Session = Depends(get_db)
):
    """Create a meeting (HOD or Elder) - supports recurring meetings"""
    hod_member = _find_member_by_phone(db, phone)
    if not hod_member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Verify HOD/Elder access
    department = db.query(Department).filter(Department.id == data.department_id).first()
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")
    if not _can_manage_department(db, hod_member, data.department_id):
        raise HTTPException(status_code=403, detail="You do not have access to this department")

    # Validate slots
    if data.start_slot < 0 or data.end_slot > 48 or data.start_slot >= data.end_slot:
        raise HTTPException(status_code=400, detail="Invalid time slots")

    # Generate meeting dates (single or recurring)
    meeting_dates = [data.meeting_date]
    recurrence_group_id = None

    if data.recurrence and data.recurrence != 'none' and data.recurrence_end_date:
        if data.recurrence_end_date <= data.meeting_date:
            raise HTTPException(status_code=400, detail="Recurrence end date must be after meeting date")

        meeting_dates = generate_recurring_dates(data.meeting_date, data.recurrence, data.recurrence_end_date)
        recurrence_group_id = str(uuid.uuid4())

    # Check availability for all dates
    conflicts = []
    for mtg_date in meeting_dates:
        available, conflict_reason = check_slot_availability(
            db, data.department_id, mtg_date, data.start_slot, data.end_slot
        )
        if not available:
            conflicts.append(f"{mtg_date.isoformat()}: {conflict_reason}")

    if conflicts:
        if len(conflicts) == 1:
            raise HTTPException(status_code=409, detail=conflicts[0])
        else:
            raise HTTPException(status_code=409, detail=f"Conflicts found: {'; '.join(conflicts[:3])}" +
                                (f" and {len(conflicts) - 3} more" if len(conflicts) > 3 else ""))

    # Idempotency: detect a recent duplicate of the anchor meeting (double-click,
    # network retry). Returns the existing series instead of creating another one.
    existing_dup = find_duplicate_meeting(
        db,
        title=data.title,
        meeting_date=data.meeting_date,
        start_slot=data.start_slot,
        end_slot=data.end_slot,
        department_id=data.department_id,
        is_general=0,
    )
    if existing_dup:
        existing_full = db.query(Meeting).options(
            joinedload(Meeting.department),
            joinedload(Meeting.created_by)
        ).filter(Meeting.id == existing_dup.id).first()
        occurrence_count = 1
        if existing_dup.recurrence_group_id:
            occurrence_count = db.query(Meeting).filter(
                Meeting.recurrence_group_id == existing_dup.recurrence_group_id
            ).count()
        return {
            "success": True,
            "meeting": format_meeting_response(existing_full, db, member_id=hod_member.id),
            "meetings_created": occurrence_count,
            "recurrence_group_id": existing_dup.recurrence_group_id,
            "duplicate_suppressed": True,
        }

    # Create meetings for all dates
    created_meetings = []
    for mtg_date in meeting_dates:
        meeting = Meeting(
            department_id=data.department_id,
            created_by_id=hod_member.id,
            title=data.title,
            description=data.description,
            meeting_date=mtg_date,
            start_slot=data.start_slot,
            end_slot=data.end_slot,
            location=data.location,
            meeting_link=data.meeting_link,
            recurrence_group_id=recurrence_group_id
        )
        db.add(meeting)
        created_meetings.append(meeting)

    db.commit()

    # Refresh all meetings
    for meeting in created_meetings:
        db.refresh(meeting)

    # Reload first meeting with relationships
    first_meeting = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    ).filter(Meeting.id == created_meetings[0].id).first()

    # Dispatch notification to department members and HOD
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        # Get approved members of this department
        dept_members = db.query(Member).join(MemberDepartment).filter(
            MemberDepartment.department_id == data.department_id,
            MemberDepartment.status == "approved"
        ).all()

        member_ids = set(m.id for m in dept_members)
        recipients = [{"id": m.id, "name": m.full_name, "email": m.email, "phone": m.phone} for m in dept_members if m.email]

        # Include HOD if not already in recipients
        if department.hod_member_id and department.hod_member_id not in member_ids:
            hod = db.query(Member).filter(Member.id == department.hod_member_id).first()
            if hod and hod.email:
                recipients.append({"id": hod.id, "name": hod.full_name, "email": hod.email, "phone": hod.phone})

        if recipients:
            recurrence_info = None
            if len(created_meetings) > 1:
                recurrence_info = f"This is a recurring meeting ({len(created_meetings)} occurrences until {meeting_dates[-1].isoformat()})"

            dispatch_event(db, EventType.MEETING_CREATED, {
                "meeting_id": first_meeting.id,
                "title": first_meeting.title,
                "description": first_meeting.description,
                "meeting_date": first_meeting.meeting_date.isoformat() if first_meeting.meeting_date else None,
                "start_time": slot_to_time(first_meeting.start_slot),
                "end_time": slot_to_time(first_meeting.end_slot),
                "location": first_meeting.location,
                "meeting_link": first_meeting.meeting_link,
                "department_name": department.name,
                "recurrence_info": recurrence_info,
                "recipients": recipients
            })
    except Exception as e:
        print(f"Failed to dispatch notification: {e}")

    _log_member_action(request, db, hod_member, "create_meeting", "meeting", created_meetings[0].id,
                       f"Created meeting '{data.title}' for {department.name} ({len(created_meetings)} occurrence(s))")
    db.commit()

    return {
        "success": True,
        "meeting": format_meeting_response(first_meeting, db),
        "meetings_created": len(created_meetings),
        "recurrence_group_id": recurrence_group_id
    }


@router.put("/hod/meetings/{meeting_id}")
def update_hod_meeting(
    meeting_id: int,
    data: MeetingUpdate,
    request: Request,
    phone: str = Query(...),
    db: Session = Depends(get_db)
):
    """Update a meeting (HOD who created it or Elder)"""
    hod_member = _find_member_by_phone(db, phone)
    if not hod_member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Find meeting
    meeting = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    ).filter(Meeting.id == meeting_id).first()

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    # Verify ownership (elders can edit any meeting, HODs only their own)
    if not _is_elder(hod_member) and meeting.created_by_id != hod_member.id:
        raise HTTPException(status_code=403, detail="You can only edit meetings you created")

    # Handle date/slot changes with conflict check
    new_date = data.meeting_date if data.meeting_date else meeting.meeting_date
    new_start = data.start_slot if data.start_slot is not None else meeting.start_slot
    new_end = data.end_slot if data.end_slot is not None else meeting.end_slot

    if new_start < 0 or new_end > 48 or new_start >= new_end:
        raise HTTPException(status_code=400, detail="Invalid time slots")

    # Check availability if date or slots changed
    if new_date != meeting.meeting_date or new_start != meeting.start_slot or new_end != meeting.end_slot:
        available, conflict_reason = check_slot_availability(
            db, meeting.department_id, new_date, new_start, new_end, exclude_meeting_id=meeting_id
        )
        if not available:
            raise HTTPException(status_code=409, detail=conflict_reason or "Time slot not available")

    # Update fields
    if data.title is not None:
        meeting.title = data.title
    if data.description is not None:
        meeting.description = data.description
    if data.meeting_date is not None:
        meeting.meeting_date = data.meeting_date
    if data.start_slot is not None:
        meeting.start_slot = data.start_slot
    if data.end_slot is not None:
        meeting.end_slot = data.end_slot
    if data.location is not None:
        meeting.location = data.location
    if data.meeting_link is not None:
        meeting.meeting_link = data.meeting_link

    db.commit()
    db.refresh(meeting)

    _log_member_action(request, db, hod_member, "update_meeting", "meeting", meeting_id,
                       f"Updated meeting '{meeting.title}'")
    db.commit()

    return {"success": True, "meeting": format_meeting_response(meeting, db)}


@router.delete("/hod/meetings/{meeting_id}")
def delete_hod_meeting(
    meeting_id: int,
    request: Request,
    phone: str = Query(...),
    delete_scope: str = Query("single", description="single, future, or all"),
    db: Session = Depends(get_db)
):
    """Delete meeting(s) - supports single, future (this and future), or all in recurring series"""
    hod_member = _find_member_by_phone(db, phone)
    if not hod_member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Find meeting
    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    # Verify ownership (elders can delete any meeting, HODs only their own)
    if not _is_elder(hod_member) and meeting.created_by_id != hod_member.id:
        raise HTTPException(status_code=403, detail="You can only delete meetings you created")

    deleted_count = 1

    if delete_scope in ("future", "all") and meeting.recurrence_group_id:
        # Delete multiple meetings in the recurring series
        # Elders can delete all; HODs only those they created
        query = db.query(Meeting).filter(
            Meeting.recurrence_group_id == meeting.recurrence_group_id
        )
        if not _is_elder(hod_member):
            query = query.filter(Meeting.created_by_id == hod_member.id)

        if delete_scope == "future":
            # Delete this meeting and all future ones in the series
            query = query.filter(Meeting.meeting_date >= meeting.meeting_date)

        meetings_to_delete = query.all()
        deleted_count = len(meetings_to_delete)

        for m in meetings_to_delete:
            db.delete(m)
    else:
        # Delete only this single meeting
        db.delete(meeting)

    db.commit()

    _log_member_action(request, db, hod_member, "delete_meeting", "meeting", meeting_id,
                       f"Deleted {deleted_count} meeting(s) (scope: {delete_scope})")
    db.commit()

    message = f"Deleted {deleted_count} meeting(s)" if deleted_count > 1 else "Meeting deleted"
    return {"success": True, "message": message, "deleted_count": deleted_count}


# --- Member Meeting Endpoints ---

@router.get("/meetings")
def get_member_meetings(phone: str = Query(...), db: Session = Depends(get_db)):
    """Get meetings for the current month for member's approved departments, plus general meetings"""
    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Get approved department IDs
    approved_dept_ids = set(
        md.department_id for md in db.query(MemberDepartment).filter(
            MemberDepartment.member_id == member.id,
            MemberDepartment.status == "approved"
        ).all()
    )

    # Also include departments where user is HOD
    hod_dept_ids = set(
        d.id for d in db.query(Department).filter(
            Department.hod_member_id == member.id
        ).all()
    )

    # Combine both sets
    member_dept_ids = approved_dept_ids | hod_dept_ids

    # Get member's leadership roles
    member_roles = []
    if member.leadership_roles:
        try:
            member_roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else member.leadership_roles
        except (ValueError, TypeError):
            member_roles = []
    # Add "hod" role if member is HOD of any department
    if hod_dept_ids:
        member_roles = ["hod"] + member_roles

    # Get remaining meetings for current month (from today onwards)
    today = date.today()
    # Calculate last day of month
    if today.month == 12:
        month_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    all_meetings = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    ).filter(
        Meeting.meeting_date >= today,
        Meeting.meeting_date <= month_end
    ).order_by(Meeting.meeting_date, Meeting.start_slot).all()

    # Filter meetings that are relevant to this member
    relevant_meetings = []
    for meeting in all_meetings:
        # General meetings are for all approved members or HODs
        if meeting.is_general:
            if member_dept_ids:  # Only show if member has at least one department
                relevant_meetings.append(meeting)
            continue

        # Leadership role meetings - check if member has any of the target roles
        if meeting.target_leadership_roles:
            try:
                target_roles = json.loads(meeting.target_leadership_roles) if isinstance(meeting.target_leadership_roles, str) else meeting.target_leadership_roles
                if set(target_roles) & set(member_roles):  # Member has at least one target role
                    relevant_meetings.append(meeting)
            except (ValueError, TypeError):
                pass
            continue

        # Individual member meetings - check if member is in the target list
        if meeting.target_member_ids:
            try:
                target_member_list = json.loads(meeting.target_member_ids) if isinstance(meeting.target_member_ids, str) else meeting.target_member_ids
                if member.id in target_member_list:
                    relevant_meetings.append(meeting)
            except (ValueError, TypeError):
                pass
            continue

        # Multi-department meetings
        if meeting.target_department_ids:
            try:
                target_ids = json.loads(meeting.target_department_ids) if isinstance(meeting.target_department_ids, str) else meeting.target_department_ids
                if set(target_ids) & member_dept_ids:  # Member has at least one target department
                    relevant_meetings.append(meeting)
            except (ValueError, TypeError):
                pass
            continue

        # Single department meetings
        if meeting.department_id and meeting.department_id in member_dept_ids:
            relevant_meetings.append(meeting)

    return {
        "meetings": [format_meeting_response(m, db, member.id) for m in relevant_meetings],
        "total": len(relevant_meetings),
        "month": today.strftime("%B %Y"),
        "month_end": month_end.isoformat()
    }


@router.post("/meetings/{meeting_id}/rsvp")
def submit_rsvp(
    meeting_id: int,
    data: RSVPRequest,
    request: Request,
    phone: str = Query(...),
    db: Session = Depends(get_db)
):
    """Submit or update RSVP for a meeting"""
    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Find meeting
    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    # Verify member is allowed to RSVP for this meeting
    is_allowed = False

    if meeting.is_general:
        # All Leaders meeting - member must have at least one approved department
        has_approved = db.query(MemberDepartment).filter(
            MemberDepartment.member_id == member.id,
            MemberDepartment.status == "approved"
        ).first()
        is_allowed = has_approved is not None
    elif meeting.target_leadership_roles:
        # Leadership role meeting - member must have one of the target roles
        try:
            target_roles = json.loads(meeting.target_leadership_roles) if isinstance(meeting.target_leadership_roles, str) else meeting.target_leadership_roles
        except:
            target_roles = []

        # Get member's leadership roles
        member_roles = []
        if member.leadership_roles:
            try:
                member_roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else member.leadership_roles
            except:
                member_roles = []

        # Check if member is HOD for "hod" role
        if "hod" in target_roles:
            is_hod = db.query(Department).filter(Department.hod_member_id == member.id).first()
            if is_hod:
                is_allowed = True

        # Check other roles
        if not is_allowed and set(target_roles) & set(member_roles):
            is_allowed = True
    elif meeting.target_member_ids:
        # Individual members meeting - member must be in the target list
        try:
            target_ids = json.loads(meeting.target_member_ids) if isinstance(meeting.target_member_ids, str) else meeting.target_member_ids
            is_allowed = member.id in target_ids
        except:
            is_allowed = False
    elif meeting.target_department_ids:
        # Multi-department meeting - member must be approved in one of the target departments
        try:
            target_ids = json.loads(meeting.target_department_ids) if isinstance(meeting.target_department_ids, str) else meeting.target_department_ids
        except:
            target_ids = []

        if target_ids:
            membership = db.query(MemberDepartment).filter(
                MemberDepartment.member_id == member.id,
                MemberDepartment.department_id.in_(target_ids),
                MemberDepartment.status == "approved"
            ).first()
            is_allowed = membership is not None
    elif meeting.department_id:
        # Single department meeting - member must be approved in that department OR be HOD
        membership = db.query(MemberDepartment).filter(
            MemberDepartment.member_id == member.id,
            MemberDepartment.department_id == meeting.department_id,
            MemberDepartment.status == "approved"
        ).first()

        # Also check if member is HOD of this department
        is_hod = db.query(Department).filter(
            Department.id == meeting.department_id,
            Department.hod_member_id == member.id
        ).first()

        is_allowed = membership is not None or is_hod is not None

    if not is_allowed:
        raise HTTPException(status_code=403, detail="You are not eligible to RSVP for this meeting")

    # Validate response
    if data.response not in ["attending", "not_attending"]:
        raise HTTPException(status_code=400, detail="Response must be 'attending' or 'not_attending'")

    # Find or create RSVP
    rsvp = db.query(MeetingRSVP).filter(
        MeetingRSVP.meeting_id == meeting_id,
        MeetingRSVP.member_id == member.id
    ).first()

    if rsvp:
        rsvp.response = data.response
        rsvp.updated_at = datetime.now()
    else:
        rsvp = MeetingRSVP(
            meeting_id=meeting_id,
            member_id=member.id,
            response=data.response
        )
        db.add(rsvp)

    _log_member_action(request, db, member, "meeting_rsvp", "meeting", meeting_id,
        f"RSVP: {data.response} for {meeting.title}")
    db.commit()

    return {"success": True, "response": data.response}


# --- Admin Meeting Endpoints ---

@router.get("/admin/meetings")
def get_all_meetings(
    department_id: Optional[int] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Get all meetings (admin)"""
    query = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    )

    if department_id:
        query = query.filter(Meeting.department_id == department_id)

    if from_date:
        try:
            from_dt = date.fromisoformat(from_date)
            query = query.filter(Meeting.meeting_date >= from_dt)
        except ValueError:
            pass

    if to_date:
        try:
            to_dt = date.fromisoformat(to_date)
            query = query.filter(Meeting.meeting_date <= to_dt)
        except ValueError:
            pass

    meetings = query.order_by(Meeting.meeting_date.desc(), Meeting.start_slot).all()

    return {
        "meetings": [format_meeting_response(m, db) for m in meetings],
        "total": len(meetings)
    }


@router.post("/admin/meetings")
def create_admin_meeting(data: MeetingCreate, request: Request = None, db: Session = Depends(get_db)):
    """Create a meeting (admin) - supports single dept, all leaders, multi-dept, individuals, and recurring"""

    if data.start_slot < 0 or data.end_slot > 48 or data.start_slot >= data.end_slot:
        raise HTTPException(status_code=400, detail="Invalid time slots")

    # Determine meeting type and validate
    is_general = data.is_general
    target_dept_ids_str = None
    target_member_ids_str = None
    target_leadership_roles_str = None
    department_id = None

    if is_general:
        # All leaders meeting - no specific department
        pass
    elif data.target_leadership_roles and len(data.target_leadership_roles) > 0:
        # Leadership roles meeting (HODs, Deacons, Elders)
        valid_roles = ["hod", "pastor", "deacon", "elder", "service_manager"]
        invalid = [r for r in data.target_leadership_roles if r not in valid_roles]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Invalid roles: {invalid}. Valid roles are: {valid_roles}")
        target_leadership_roles_str = json.dumps(data.target_leadership_roles)
    elif data.target_member_ids and len(data.target_member_ids) > 0:
        # Individual members meeting
        # Validate all members exist
        for member_id in data.target_member_ids:
            member = db.query(Member).filter(Member.id == member_id).first()
            if not member:
                raise HTTPException(status_code=404, detail=f"Member {member_id} not found")
        target_member_ids_str = json.dumps(data.target_member_ids)
    elif data.target_department_ids and len(data.target_department_ids) > 0:
        # Multi-department meeting
        # Validate all departments exist
        for dept_id in data.target_department_ids:
            dept = db.query(Department).filter(Department.id == dept_id).first()
            if not dept:
                raise HTTPException(status_code=404, detail=f"Department {dept_id} not found")
        target_dept_ids_str = json.dumps(data.target_department_ids)
    elif data.department_id:
        # Single department meeting
        department = db.query(Department).filter(Department.id == data.department_id).first()
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")

        department_id = data.department_id
    else:
        raise HTTPException(status_code=400, detail="Meeting must target a department, multiple departments, leadership roles, specific members, or all leaders")

    # Generate meeting dates (single or recurring)
    meeting_dates = [data.meeting_date]
    recurrence_group_id = None

    if data.recurrence and data.recurrence != 'none' and data.recurrence_end_date:
        if data.recurrence_end_date <= data.meeting_date:
            raise HTTPException(status_code=400, detail="Recurrence end date must be after meeting date")

        meeting_dates = generate_recurring_dates(data.meeting_date, data.recurrence, data.recurrence_end_date)
        recurrence_group_id = str(uuid.uuid4())

    # Check availability for all dates (only for department-based meetings)
    conflicts = []
    dept_ids_to_check = []
    if is_general:
        dept_ids_to_check = []  # No specific department to check
    elif target_member_ids_str:
        dept_ids_to_check = []  # Individual meeting - no department conflicts
    elif data.target_department_ids:
        dept_ids_to_check = data.target_department_ids
    elif department_id:
        dept_ids_to_check = [department_id]

    for meeting_date in meeting_dates:
        for dept_id in dept_ids_to_check:
            available, conflict_reason = check_slot_availability(
                db, dept_id, meeting_date, data.start_slot, data.end_slot
            )
            if not available:
                dept = db.query(Department).filter(Department.id == dept_id).first()
                conflicts.append(f"{meeting_date.isoformat()} ({dept.name if dept else 'Unknown'}): {conflict_reason}")

    if conflicts:
        if len(conflicts) == 1:
            raise HTTPException(status_code=409, detail=conflicts[0])
        else:
            raise HTTPException(status_code=409, detail=f"Conflicts found: {'; '.join(conflicts[:3])}" +
                                (f" and {len(conflicts) - 3} more" if len(conflicts) > 3 else ""))

    # Idempotency: detect a recent duplicate of the anchor meeting (double-click,
    # network retry, impatient admin). Returns the existing series instead of
    # creating another one.
    existing_dup = find_duplicate_meeting(
        db,
        title=data.title,
        meeting_date=data.meeting_date,
        start_slot=data.start_slot,
        end_slot=data.end_slot,
        department_id=department_id,
        target_department_ids=target_dept_ids_str,
        target_member_ids=target_member_ids_str,
        target_leadership_roles=target_leadership_roles_str,
        is_general=1 if is_general else 0,
    )
    if existing_dup:
        existing_full = db.query(Meeting).options(
            joinedload(Meeting.department),
            joinedload(Meeting.created_by)
        ).filter(Meeting.id == existing_dup.id).first()
        # Count occurrences in the series (if it was a recurring series)
        occurrence_count = 1
        if existing_dup.recurrence_group_id:
            occurrence_count = db.query(Meeting).filter(
                Meeting.recurrence_group_id == existing_dup.recurrence_group_id
            ).count()
        return {
            "success": True,
            "meeting": format_meeting_response(existing_full, db),
            "meetings_created": occurrence_count,
            "recurrence_group_id": existing_dup.recurrence_group_id,
            "duplicate_suppressed": True,
        }

    # Create meetings for all dates
    created_meetings = []
    for mtg_date in meeting_dates:
        meeting = Meeting(
            department_id=department_id,
            created_by_id=None,  # Admin-created
            title=data.title,
            description=data.description,
            meeting_date=mtg_date,
            start_slot=data.start_slot,
            end_slot=data.end_slot,
            location=data.location,
            meeting_link=data.meeting_link,
            is_general=1 if is_general else 0,
            target_department_ids=target_dept_ids_str,
            target_member_ids=target_member_ids_str,
            target_leadership_roles=target_leadership_roles_str,
            recurrence_group_id=recurrence_group_id
        )
        db.add(meeting)
        created_meetings.append(meeting)

    db.commit()

    # Refresh all meetings
    for meeting in created_meetings:
        db.refresh(meeting)

    # Get the first meeting with relationships for response
    first_meeting = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    ).filter(Meeting.id == created_meetings[0].id).first()

    # Dispatch notification for the first meeting only (or series)
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        # Determine recipients based on meeting type
        member_ids = set()
        recipients = []
        dept_name = None
        target_dept_ids = []

        if is_general:
            # All approved members
            members = db.query(Member).join(MemberDepartment).filter(
                MemberDepartment.status == "approved"
            ).distinct().all()
            member_ids = set(m.id for m in members)
            recipients = [{"id": m.id, "name": m.full_name, "email": m.email, "phone": m.phone} for m in members if m.email]
            dept_name = "All Leaders"
            # Get all department IDs for HOD lookup
            all_depts = db.query(Department.id).all()
            target_dept_ids = [d[0] for d in all_depts]
        elif data.target_leadership_roles:
            # Members with specific leadership roles
            role_labels = {"hod": "HODs", "pastor": "Pastors", "deacon": "Deacons", "elder": "Elders", "service_manager": "Service Managers"}
            dept_name = ", ".join([role_labels.get(r, r.title()) for r in data.target_leadership_roles])
            members_set = set()

            for role in data.target_leadership_roles:
                if role == "hod":
                    # Get all HODs
                    hod_depts = db.query(Department).filter(Department.hod_member_id.isnot(None)).all()
                    hod_ids = [d.hod_member_id for d in hod_depts]
                    hods = db.query(Member).filter(Member.id.in_(hod_ids)).all()
                    for m in hods:
                        members_set.add(m)
                else:
                    # Get members with this role in leadership_roles JSON
                    all_members = db.query(Member).filter(Member.leadership_roles.isnot(None)).all()
                    for m in all_members:
                        try:
                            m_roles = json.loads(m.leadership_roles) if isinstance(m.leadership_roles, str) else m.leadership_roles
                            if role in m_roles:
                                members_set.add(m)
                        except (ValueError, TypeError):
                            pass

            member_ids = set(m.id for m in members_set)
            recipients = [{"id": m.id, "name": m.full_name, "email": m.email, "phone": m.phone} for m in members_set if m.email]
        elif data.target_member_ids:
            # Specific individual members
            members = db.query(Member).filter(Member.id.in_(data.target_member_ids)).all()
            member_ids = set(m.id for m in members)
            recipients = [{"id": m.id, "name": m.full_name, "email": m.email, "phone": m.phone} for m in members if m.email]
            dept_name = f"{len(members)} Selected Members"
        elif data.target_department_ids:
            # Members from specified departments
            members = db.query(Member).join(MemberDepartment).filter(
                MemberDepartment.department_id.in_(data.target_department_ids),
                MemberDepartment.status == "approved"
            ).distinct().all()
            member_ids = set(m.id for m in members)
            recipients = [{"id": m.id, "name": m.full_name, "email": m.email, "phone": m.phone} for m in members if m.email]
            depts = db.query(Department).filter(Department.id.in_(data.target_department_ids)).all()
            dept_name = ", ".join([d.name for d in depts])
            target_dept_ids = data.target_department_ids
        elif department_id:
            # Single department
            members = db.query(Member).join(MemberDepartment).filter(
                MemberDepartment.department_id == department_id,
                MemberDepartment.status == "approved"
            ).all()
            member_ids = set(m.id for m in members)
            recipients = [{"id": m.id, "name": m.full_name, "email": m.email, "phone": m.phone} for m in members if m.email]
            dept = db.query(Department).filter(Department.id == department_id).first()
            dept_name = dept.name if dept else None
            target_dept_ids = [department_id]

        # Include HODs of relevant departments if not already in recipients
        if target_dept_ids:
            hod_depts = db.query(Department).filter(
                Department.id.in_(target_dept_ids),
                Department.hod_member_id.isnot(None)
            ).all()
            hod_ids = [d.hod_member_id for d in hod_depts if d.hod_member_id not in member_ids]
            if hod_ids:
                hods = db.query(Member).filter(Member.id.in_(hod_ids)).all()
                for hod in hods:
                    if hod.email:
                        recipients.append({"id": hod.id, "name": hod.full_name, "email": hod.email, "phone": hod.phone})

        if recipients:
            # Include recurrence info in notification
            recurrence_info = None
            if len(created_meetings) > 1:
                recurrence_info = f"This is a recurring meeting ({len(created_meetings)} occurrences until {meeting_dates[-1].isoformat()})"

            dispatch_event(db, EventType.MEETING_CREATED, {
                "meeting_id": first_meeting.id,
                "title": first_meeting.title,
                "description": first_meeting.description,
                "meeting_date": first_meeting.meeting_date.isoformat() if first_meeting.meeting_date else None,
                "start_time": slot_to_time(first_meeting.start_slot),
                "end_time": slot_to_time(first_meeting.end_slot),
                "location": first_meeting.location,
                "meeting_link": first_meeting.meeting_link,
                "department_name": dept_name,
                "is_general": is_general,
                "recurrence_info": recurrence_info,
                "recipients": recipients
            })
    except Exception as e:
        print(f"Failed to dispatch notification: {e}")

    if request:
        _log_admin_action(request, db, "create_meeting", "meeting", created_meetings[0].id,
                          f"Created meeting '{data.title}' ({len(created_meetings)} occurrence(s))")
        db.commit()

    return {
        "success": True,
        "meeting": format_meeting_response(first_meeting, db),
        "meetings_created": len(created_meetings),
        "recurrence_group_id": recurrence_group_id
    }


@router.put("/admin/meetings/{meeting_id}")
def update_admin_meeting(
    meeting_id: int,
    data: MeetingUpdate,
    request: Request,
    db: Session = Depends(get_db)
):
    """Update any meeting (admin)"""
    meeting = db.query(Meeting).options(
        joinedload(Meeting.department),
        joinedload(Meeting.created_by)
    ).filter(Meeting.id == meeting_id).first()

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    # Handle date/slot changes with conflict check
    new_date = data.meeting_date if data.meeting_date else meeting.meeting_date
    new_start = data.start_slot if data.start_slot is not None else meeting.start_slot
    new_end = data.end_slot if data.end_slot is not None else meeting.end_slot

    if new_start < 0 or new_end > 48 or new_start >= new_end:
        raise HTTPException(status_code=400, detail="Invalid time slots")

    if new_date != meeting.meeting_date or new_start != meeting.start_slot or new_end != meeting.end_slot:
        available, conflict_reason = check_slot_availability(
            db, meeting.department_id, new_date, new_start, new_end, exclude_meeting_id=meeting_id
        )
        if not available:
            raise HTTPException(status_code=409, detail=conflict_reason or "Time slot not available")

    if data.title is not None:
        meeting.title = data.title
    if data.description is not None:
        meeting.description = data.description
    if data.meeting_date is not None:
        meeting.meeting_date = data.meeting_date
    if data.start_slot is not None:
        meeting.start_slot = data.start_slot
    if data.end_slot is not None:
        meeting.end_slot = data.end_slot
    if data.location is not None:
        meeting.location = data.location
    if data.meeting_link is not None:
        meeting.meeting_link = data.meeting_link

    db.commit()
    db.refresh(meeting)

    _log_admin_action(request, db, "update_meeting", "meeting", meeting_id, f"Updated meeting '{meeting.title}'")
    db.commit()

    return {"success": True, "meeting": format_meeting_response(meeting, db)}


@router.delete("/admin/meetings/{meeting_id}")
def delete_admin_meeting(
    meeting_id: int,
    request: Request,
    delete_scope: str = Query("single", description="single, future, or all"),
    db: Session = Depends(get_db)
):
    """Delete meeting(s) - supports single, future (this and future), or all in recurring series"""
    meeting = db.query(Meeting).filter(Meeting.id == meeting_id).first()
    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    deleted_count = 1

    if delete_scope in ("future", "all") and meeting.recurrence_group_id:
        # Delete multiple meetings in the recurring series
        query = db.query(Meeting).filter(
            Meeting.recurrence_group_id == meeting.recurrence_group_id
        )

        if delete_scope == "future":
            # Delete this meeting and all future ones in the series
            query = query.filter(Meeting.meeting_date >= meeting.meeting_date)

        meetings_to_delete = query.all()
        deleted_count = len(meetings_to_delete)

        for m in meetings_to_delete:
            db.delete(m)
    else:
        # Delete only this single meeting
        db.delete(meeting)

    db.commit()

    _log_admin_action(request, db, "delete_meeting", "meeting", meeting_id,
                      f"Deleted {deleted_count} meeting(s) (scope: {delete_scope})")
    db.commit()

    message = f"Deleted {deleted_count} meeting(s)" if deleted_count > 1 else "Meeting deleted"
    return {"success": True, "message": message, "deleted_count": deleted_count}


@router.get("/admin/meetings/{meeting_id}/rsvps")
def get_meeting_rsvps(meeting_id: int, db: Session = Depends(get_db)):
    """Get all RSVPs for a meeting (admin)"""
    meeting = db.query(Meeting).options(
        joinedload(Meeting.department)
    ).filter(Meeting.id == meeting_id).first()

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    rsvps = db.query(MeetingRSVP).options(
        joinedload(MeetingRSVP.member)
    ).filter(MeetingRSVP.meeting_id == meeting_id).all()

    return {
        "meeting_id": meeting_id,
        "meeting_title": meeting.title,
        "department_name": meeting.department.name if meeting.department else None,
        "rsvps": [
            {
                "id": r.id,
                "member_id": r.member_id,
                "member_name": r.member.full_name if r.member else None,
                "member_phone": r.member.phone if r.member else None,
                "response": r.response,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None
            }
            for r in rsvps
        ],
        "total": len(rsvps),
        "attending": sum(1 for r in rsvps if r.response == "attending"),
        "not_attending": sum(1 for r in rsvps if r.response == "not_attending")
    }


# ============ POSTER REQUEST ENDPOINTS ============

def format_poster_request(pr: PosterRequest) -> dict:
    """Format a poster request for API response"""
    import json

    # Parse speakers from JSON
    speakers = None
    if pr.speakers:
        try:
            speakers = json.loads(pr.speakers)
        except (json.JSONDecodeError, TypeError):
            speakers = None

    # Parse output_formats from JSON
    output_formats = None
    if pr.output_formats:
        try:
            output_formats = json.loads(pr.output_formats)
        except (json.JSONDecodeError, TypeError):
            output_formats = None

    return {
        "id": pr.id,
        "requester_id": pr.requester_id,
        "requester_name": pr.requester.full_name if pr.requester else None,
        "requester_email": pr.requester.email if pr.requester else None,
        "event_name": pr.event_name,
        "ministry_department": pr.ministry_department,
        "event_date": pr.event_date.isoformat() if pr.event_date else None,
        "event_time": pr.event_time,
        "venue_platform": pr.venue_platform,
        "speakers": speakers,
        "theme_tagline": pr.theme_tagline,
        "scripture": pr.scripture,
        "target_audience": pr.target_audience,
        "purpose": pr.purpose,
        "output_formats": output_formats,
        "additional_notes": pr.additional_notes,
        "status": pr.status,
        "acknowledged_by_id": pr.acknowledged_by_id,
        "acknowledged_by_name": pr.acknowledged_by.full_name if pr.acknowledged_by else None,
        "acknowledged_at": pr.acknowledged_at.isoformat() if pr.acknowledged_at else None,
        "created_at": pr.created_at.isoformat() if pr.created_at else None
    }


@router.post("/poster-requests")
def create_poster_request(
    data: PosterRequestCreate,
    request: Request,
    phone: str = Query(...),
    db: Session = Depends(get_db)
):
    """Submit a new poster request (requires member login)"""
    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Serialize speakers and output_formats to JSON
    import json
    speakers_json = None
    if data.speakers:
        speakers_json = json.dumps([s.model_dump() for s in data.speakers])

    output_formats_json = None
    if data.output_formats:
        output_formats_json = json.dumps(data.output_formats)

    # Create the poster request
    pr = PosterRequest(
        requester_id=member.id,
        event_name=data.event_name,
        ministry_department=data.ministry_department,
        event_date=data.event_date,
        event_time=data.event_time,
        venue_platform=data.venue_platform,
        speakers=speakers_json,
        theme_tagline=data.theme_tagline,
        scripture=data.scripture,
        target_audience=data.target_audience,
        purpose=data.purpose,
        output_formats=output_formats_json,
        additional_notes=data.additional_notes,
        status="pending"
    )
    db.add(pr)
    db.flush()
    _log_member_action(request, db, member, "create_poster_request", "poster_request", pr.id,
        f"Poster request: {data.event_name}")
    db.commit()
    db.refresh(pr)

    # Get configured department for poster requests
    dept_setting = db.query(Settings).filter(Settings.key == "poster_request_department_id").first()
    if dept_setting and dept_setting.value:
        try:
            dept_id = int(dept_setting.value)
            # Get all members of that department
            dept_members = db.query(Member).join(MemberDepartment).filter(
                MemberDepartment.department_id == dept_id,
                MemberDepartment.status == "approved"
            ).all()

            # Also include HOD
            dept = db.query(Department).filter(Department.id == dept_id).first()
            if dept and dept.hod_member_id:
                hod = db.query(Member).filter(Member.id == dept.hod_member_id).first()
                if hod and hod not in dept_members:
                    dept_members.append(hod)

            # Send email notifications
            try:
                from notifications.dispatcher import dispatch_event
                from notifications.events import EventType

                recipients = [{"id": m.id, "name": m.full_name, "email": m.email, "phone": m.phone}
                              for m in dept_members if m.email]

                # Format speakers for display
                speakers_display = None
                if speakers_json:
                    try:
                        speakers_list = json.loads(speakers_json)
                        speakers_display = ", ".join([
                            f"{s['name']} ({s['role']})" for s in speakers_list
                        ])
                    except:
                        pass

                # Format output formats for display
                output_formats_display = None
                if output_formats_json:
                    try:
                        formats_list = json.loads(output_formats_json)
                        format_labels = {'projector': 'Projector', 'social_media': 'Social Media', 'print': 'Print'}
                        output_formats_display = ", ".join([format_labels.get(f, f) for f in formats_list])
                    except:
                        pass

                if recipients:
                    dispatch_event(db, EventType.POSTER_REQUEST_SUBMITTED, {
                        "request_id": pr.id,
                        "event_name": pr.event_name,
                        "ministry_department": pr.ministry_department or "Not specified",
                        "event_date": pr.event_date.isoformat() if pr.event_date else None,
                        "event_time": pr.event_time,
                        "venue_platform": pr.venue_platform,
                        "requester_name": member.full_name,
                        "requester_email": member.email,
                        "purpose": pr.purpose,
                        "speakers_display": speakers_display,
                        "output_formats_display": output_formats_display,
                        "theme_tagline": pr.theme_tagline,
                        "scripture": pr.scripture,
                        "target_audience": pr.target_audience,
                        "additional_notes": pr.additional_notes,
                        "recipients": recipients
                    })
            except Exception as e:
                print(f"Failed to dispatch poster request notification: {e}")

        except (ValueError, TypeError):
            pass

    return {"success": True, "request": format_poster_request(pr)}


@router.get("/poster-requests")
def get_poster_requests(
    phone: str = Query(...),
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get poster requests - for design team members shows all pending, for others shows their own"""
    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Check if member is in the poster request department
    dept_setting = db.query(Settings).filter(Settings.key == "poster_request_department_id").first()
    is_design_team = False

    if dept_setting and dept_setting.value:
        try:
            dept_id = int(dept_setting.value)
            # Check membership
            membership = db.query(MemberDepartment).filter(
                MemberDepartment.member_id == member.id,
                MemberDepartment.department_id == dept_id,
                MemberDepartment.status == "approved"
            ).first()
            if membership:
                is_design_team = True

            # Check if HOD
            dept = db.query(Department).filter(Department.id == dept_id).first()
            if dept and dept.hod_member_id == member.id:
                is_design_team = True
        except (ValueError, TypeError):
            pass

    # Build query
    query = db.query(PosterRequest).options(
        joinedload(PosterRequest.requester),
        joinedload(PosterRequest.acknowledged_by)
    )

    if is_design_team:
        # Design team sees all requests
        if status:
            query = query.filter(PosterRequest.status == status)
    else:
        # Others see only their own requests
        query = query.filter(PosterRequest.requester_id == member.id)
        if status:
            query = query.filter(PosterRequest.status == status)

    requests = query.order_by(PosterRequest.created_at.desc()).all()

    return {
        "requests": [format_poster_request(pr) for pr in requests],
        "is_design_team": is_design_team,
        "total": len(requests)
    }


@router.put("/poster-requests/{request_id}/acknowledge")
def acknowledge_poster_request(
    request_id: int,
    request: Request,
    phone: str = Query(...),
    db: Session = Depends(get_db)
):
    """Acknowledge a poster request (design team only)"""
    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Check if member is in the poster request department
    dept_setting = db.query(Settings).filter(Settings.key == "poster_request_department_id").first()
    is_design_team = False

    if dept_setting and dept_setting.value:
        try:
            dept_id = int(dept_setting.value)
            membership = db.query(MemberDepartment).filter(
                MemberDepartment.member_id == member.id,
                MemberDepartment.department_id == dept_id,
                MemberDepartment.status == "approved"
            ).first()
            if membership:
                is_design_team = True

            dept = db.query(Department).filter(Department.id == dept_id).first()
            if dept and dept.hod_member_id == member.id:
                is_design_team = True
        except (ValueError, TypeError):
            pass

    if not is_design_team:
        raise HTTPException(status_code=403, detail="Only design team members can acknowledge requests")

    # Find the request
    pr = db.query(PosterRequest).options(
        joinedload(PosterRequest.requester)
    ).filter(PosterRequest.id == request_id).first()

    if not pr:
        raise HTTPException(status_code=404, detail="Request not found")

    if pr.status != "pending":
        raise HTTPException(status_code=400, detail="Request has already been acknowledged")

    # Update the request
    pr.status = "acknowledged"
    pr.acknowledged_by_id = member.id
    pr.acknowledged_at = datetime.now()
    db.commit()
    db.refresh(pr)

    # Notify the requester
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        if pr.requester and pr.requester.email:
            dispatch_event(db, EventType.POSTER_REQUEST_ACKNOWLEDGED, {
                "request_id": pr.id,
                "event_name": pr.event_name,
                "acknowledged_by_name": member.full_name,
                "recipients": [{
                    "id": pr.requester.id,
                    "name": pr.requester.full_name,
                    "email": pr.requester.email,
                    "phone": pr.requester.phone
                }]
            })
    except Exception as e:
        print(f"Failed to dispatch acknowledgment notification: {e}")

    _log_member_action(request, db, member, "acknowledge_poster_request", "poster_request", request_id,
                       f"Acknowledged poster request '{pr.event_name}'")
    db.commit()

    return {"success": True, "request": format_poster_request(pr)}


@router.put("/poster-requests/{request_id}/complete")
def complete_poster_request(
    request_id: int,
    request: Request,
    phone: str = Query(...),
    db: Session = Depends(get_db)
):
    """Mark a poster request as complete (design team only)"""
    # Find member by phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member = None
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member = m
            break

    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    # Check if member is in the poster request department
    dept_setting = db.query(Settings).filter(Settings.key == "poster_request_department_id").first()
    is_design_team = False

    if dept_setting and dept_setting.value:
        try:
            dept_id = int(dept_setting.value)
            membership = db.query(MemberDepartment).filter(
                MemberDepartment.member_id == member.id,
                MemberDepartment.department_id == dept_id,
                MemberDepartment.status == "approved"
            ).first()
            if membership:
                is_design_team = True

            dept = db.query(Department).filter(Department.id == dept_id).first()
            if dept and dept.hod_member_id == member.id:
                is_design_team = True
        except (ValueError, TypeError):
            pass

    if not is_design_team:
        raise HTTPException(status_code=403, detail="Only design team members can complete requests")

    # Find the request
    pr = db.query(PosterRequest).options(
        joinedload(PosterRequest.requester),
        joinedload(PosterRequest.acknowledged_by)
    ).filter(PosterRequest.id == request_id).first()

    if not pr:
        raise HTTPException(status_code=404, detail="Request not found")

    if pr.status != "acknowledged":
        raise HTTPException(status_code=400, detail="Request must be in 'processing' status to mark as done")

    # Update the request
    pr.status = "completed"
    pr.completed_at = datetime.now()
    db.commit()
    db.refresh(pr)

    # Notify the requester that poster is ready
    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        if pr.requester and pr.requester.email:
            dispatch_event(db, EventType.POSTER_REQUEST_COMPLETED, {
                "request_id": pr.id,
                "event_name": pr.event_name,
                "event_date": pr.event_date.isoformat() if pr.event_date else None,
                "completed_by_name": member.full_name,
                "recipients": [{
                    "id": pr.requester.id,
                    "name": pr.requester.full_name,
                    "email": pr.requester.email,
                    "phone": pr.requester.phone
                }]
            })
    except Exception as e:
        print(f"Failed to dispatch completion notification: {e}")

    _log_member_action(request, db, member, "complete_poster_request", "poster_request", request_id,
                       f"Completed poster request '{pr.event_name}'")
    db.commit()

    return {"success": True, "request": format_poster_request(pr)}


@router.get("/admin/poster-requests")
def get_all_poster_requests(
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Admin: Get all poster requests"""
    query = db.query(PosterRequest).options(
        joinedload(PosterRequest.requester),
        joinedload(PosterRequest.acknowledged_by)
    )

    if status:
        query = query.filter(PosterRequest.status == status)

    requests = query.order_by(PosterRequest.created_at.desc()).all()

    return {
        "requests": [format_poster_request(pr) for pr in requests],
        "total": len(requests)
    }


@router.get("/admin/settings/poster-request-department")
def get_poster_request_department(db: Session = Depends(get_db)):
    """Get the configured department for handling poster requests"""
    setting = db.query(Settings).filter(Settings.key == "poster_request_department_id").first()
    dept_id = setting.value if setting else None

    dept = None
    if dept_id:
        try:
            dept = db.query(Department).filter(Department.id == int(dept_id)).first()
        except (ValueError, TypeError):
            pass

    return {
        "department_id": int(dept_id) if dept_id else None,
        "department_name": dept.name if dept else None
    }


@router.put("/admin/settings/poster-request-department")
def set_poster_request_department(
    department_id: int = Body(..., embed=True),
    request: Request = None,
    db: Session = Depends(get_db)
):
    """Set the department that handles poster requests"""
    # Validate department exists
    dept = db.query(Department).filter(Department.id == department_id).first()
    if not dept:
        raise HTTPException(status_code=404, detail="Department not found")

    # Update or create setting
    setting = db.query(Settings).filter(Settings.key == "poster_request_department_id").first()
    if setting:
        setting.value = str(department_id)
    else:
        setting = Settings(key="poster_request_department_id", value=str(department_id))
        db.add(setting)

    db.commit()

    if request:
        _log_admin_action(request, db, "update_settings", "settings", None, f"Set poster request department to '{dept.name}'")
        db.commit()

    return {"success": True, "department_id": department_id, "department_name": dept.name}


# ============ NOTIFICATION ENDPOINTS ============

@router.get("/admin/notifications/email-settings")
def get_email_settings(db: Session = Depends(get_db)):
    """Get all email settings (SMTP and Resend), passwords masked"""
    all_keys = [
        'smtp_enabled', 'smtp_host', 'smtp_port',
        'smtp_username', 'smtp_password',
        'smtp_from_name', 'smtp_from_email', 'smtp_reply_to',
        'resend_enabled', 'resend_api_key',
        'resend_from_name', 'resend_from_email', 'resend_reply_to',
        'reply_to',
    ]
    settings = db.query(Settings).filter(Settings.key.in_(all_keys)).all()
    result = {s.key: s.value for s in settings}

    # Mask secrets
    if result.get('smtp_password'):
        result['smtp_password'] = '********'
    if result.get('resend_api_key'):
        result['resend_api_key'] = '********'

    return result


@router.get("/admin/notifications/smtp-settings")
def get_smtp_settings(db: Session = Depends(get_db)):
    """Get SMTP settings (password is masked) - legacy endpoint"""
    smtp_keys = [
        'smtp_enabled', 'smtp_host', 'smtp_port',
        'smtp_username', 'smtp_password',
        'smtp_from_name', 'smtp_from_email'
    ]
    settings = db.query(Settings).filter(Settings.key.in_(smtp_keys)).all()
    result = {s.key: s.value for s in settings}

    # Mask the password
    if result.get('smtp_password'):
        result['smtp_password'] = '********' if result['smtp_password'] else ''

    return result


@router.put("/admin/notifications/smtp-settings")
def update_smtp_settings(data: SMTPSettingsUpdate, db: Session = Depends(get_db)):
    """Update SMTP / Resend settings (the form sends both through this endpoint)."""
    updates = data.model_dump(exclude_none=True)

    for key, value in updates.items():
        # Skip secret update if it's the masked value
        if key in ('smtp_password', 'resend_api_key') and value == '********':
            continue

        setting = db.query(Settings).filter(Settings.key == key).first()
        if setting:
            setting.value = value
        else:
            db.add(Settings(key=key, value=value))

    db.commit()
    return {"success": True}


@router.post("/admin/notifications/test-email")
def send_test_email(data: TestEmailRequest, db: Session = Depends(get_db)):
    """Send a test email through rfm-notify."""
    from notifications.channels.rfm_notify import RfmNotifyChannel

    channel = RfmNotifyChannel()
    if not channel.is_configured():
        raise HTTPException(
            status_code=400,
            detail="rfm-notify not configured. Set RFM_NOTIFY_URL and RFM_NOTIFY_API_KEY env vars.",
        )

    # Reachability check (hits /health, no auth)
    conn_success, conn_error = channel.test_connection()
    if not conn_success:
        raise HTTPException(status_code=400, detail=f"Connection failed: {conn_error}")

    html = (
        "<div style='font-family:system-ui,sans-serif;padding:24px'>"
        "<h2 style='color:#5b21b6'>Test email from RFM Stellenbosch Portal</h2>"
        "<p>If you received this, the portal is correctly routing email through rfm-notify.</p>"
        "</div>"
    )
    success, error = channel.send(
        data.to_email,
        "Test email from RFM Stellenbosch Portal",
        html,
        event_code="portal.test_email",
        idempotency_key=f"test_email:{data.to_email}:{int(__import__('time').time())}",
    )
    if not success:
        raise HTTPException(status_code=400, detail=f"Failed to send: {error}")

    return {"success": True, "message": f"Test email sent via rfm-notify to {data.to_email}"}


@router.get("/admin/notifications/config")
def get_notification_configs(db: Session = Depends(get_db)):
    """Get all notification event configurations"""
    from notifications.events import EventType, EVENT_LABELS, EVENT_DESCRIPTIONS

    configs = db.query(NotificationConfig).all()
    config_map = {c.event_type: c for c in configs}

    result = []
    for event_type in EventType:
        config = config_map.get(event_type.value)
        result.append({
            "event_type": event_type.value,
            "label": EVENT_LABELS.get(event_type, event_type.value),
            "description": EVENT_DESCRIPTIONS.get(event_type, ""),
            "email_enabled": bool(config.email_enabled) if config else True,
            "sms_enabled": bool(config.sms_enabled) if config else False,
            "push_enabled": bool(config.push_enabled) if config else False
        })

    return result


@router.put("/admin/notifications/config/{event_type}")
def update_notification_config(
    event_type: str,
    data: NotificationConfigUpdate,
    db: Session = Depends(get_db)
):
    """Update notification config for a specific event type"""
    from notifications.events import EventType

    # Validate event type
    valid_types = [e.value for e in EventType]
    if event_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid event type: {event_type}")

    config = db.query(NotificationConfig).filter(
        NotificationConfig.event_type == event_type
    ).first()

    if not config:
        config = NotificationConfig(event_type=event_type)
        db.add(config)

    if data.email_enabled is not None:
        config.email_enabled = data.email_enabled
    if data.sms_enabled is not None:
        config.sms_enabled = data.sms_enabled
    if data.push_enabled is not None:
        config.push_enabled = data.push_enabled

    db.commit()
    return {"success": True, "event_type": event_type}


@router.get("/admin/notifications/logs")
def get_notification_logs(
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    status: Optional[str] = Query(None),
    event_type: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Get notification logs with pagination and filtering"""
    query = db.query(NotificationLog)

    if status:
        query = query.filter(NotificationLog.status == status)
    if event_type:
        query = query.filter(NotificationLog.event_type == event_type)

    total = query.count()
    logs = query.order_by(NotificationLog.created_at.desc()).offset(offset).limit(limit).all()

    return {
        "logs": [
            {
                "id": log.id,
                "event_type": log.event_type,
                "channel": log.channel,
                "recipient_email": log.recipient_email,
                "recipient_phone": log.recipient_phone,
                "subject": log.subject,
                "status": log.status,
                "error_message": log.error_message,
                "sent_at": log.sent_at.isoformat() if log.sent_at else None,
                "created_at": log.created_at.isoformat() if log.created_at else None
            }
            for log in logs
        ],
        "total": total,
        "limit": limit,
        "offset": offset
    }


# ============ SCHEDULER / REMINDERS ============

@router.get("/admin/scheduler/status")
def get_scheduler_status():
    """Get the current status of the background scheduler"""
    from scheduler import get_scheduler_status
    return get_scheduler_status()


@router.post("/admin/scheduler/reminders/trigger")
def trigger_meeting_reminders(meeting_ids: Optional[List[int]] = Query(None)):
    """
    Manually trigger meeting reminders.

    Args:
        meeting_ids: Optional list of specific meeting IDs to include.
                    If not provided, sends for all upcoming meetings (today + tomorrow).
    """
    from scheduler import send_meeting_reminders
    result = send_meeting_reminders(meeting_ids=meeting_ids)
    return result


@router.post("/admin/scheduler/reminders/preview")
def preview_meeting_reminders(db: Session = Depends(get_db)):
    """
    Preview which meetings would get reminders.
    Shows today's meetings.
    Does not send any emails, just returns the list.
    """
    from scheduler import get_meeting_recipients, slot_to_time

    now = datetime.now()
    today = now.date()
    today_str = today.isoformat()

    # Find today's meetings
    meetings = db.query(Meeting).options(
        joinedload(Meeting.department)
    ).filter(
        Meeting.meeting_date == today_str
    ).order_by(Meeting.start_slot).all()

    preview = []
    total_recipients = 0

    for meeting in meetings:
        recipients = get_meeting_recipients(db, meeting)
        total_recipients += len(recipients)

        preview.append({
            "id": meeting.id,
            "title": meeting.title,
            "date": meeting.meeting_date,
            "time": f"{slot_to_time(meeting.start_slot)} - {slot_to_time(meeting.end_slot)}",
            "location": meeting.location,
            "department": meeting.department.name if meeting.department else "All Leaders",
            "is_general": meeting.is_general,
            "recipient_count": len(recipients),
            "recipients": [
                {"name": r["name"], "email": r["email"]}
                for r in recipients[:10]  # Limit preview to first 10
            ],
            "more_recipients": max(0, len(recipients) - 10)
        })

    return {
        "today": today_str,
        "meetings_count": len(meetings),
        "total_recipients": total_recipients,
        "meetings": preview
    }


@router.post("/admin/meetings/{meeting_id}/send-invite")
def send_meeting_invite(meeting_id: int, request: Request, db: Session = Depends(get_db)):
    """
    Manually send meeting invite (RSVP request) to all members of the meeting's department.
    """
    from scheduler import get_meeting_recipients, slot_to_time
    from notifications.dispatcher import dispatch_event
    from notifications.events import EventType

    meeting = db.query(Meeting).options(
        joinedload(Meeting.department)
    ).filter(Meeting.id == meeting_id).first()

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    recipients = get_meeting_recipients(db, meeting)

    if not recipients:
        return {
            "success": False,
            "message": "No recipients found for this meeting",
            "emails_sent": 0
        }

    # Prepare meeting data
    meeting_data = {
        "meeting_id": meeting.id,
        "title": meeting.title,
        "description": meeting.description,
        "meeting_date": meeting.meeting_date,
        "start_time": slot_to_time(meeting.start_slot),
        "end_time": slot_to_time(meeting.end_slot),
        "location": meeting.location,
        "meeting_link": meeting.meeting_link,
        "department_name": meeting.department.name if meeting.department else "All Leaders",
        "is_general": meeting.is_general
    }

    # Dispatch invite notification
    dispatch_event(
        db=db,
        event_type=EventType.MEETING_CREATED,
        data=meeting_data,
        recipients=recipients
    )

    _log_admin_action(request, db, "send_meeting_invite", "meeting", meeting_id,
                      f"Sent invite for '{meeting.title}' to {len(recipients)} recipient(s)")
    db.commit()

    return {
        "success": True,
        "message": f"Meeting invite sent to {len(recipients)} recipient(s)",
        "emails_sent": len(recipients)
    }


@router.post("/admin/meetings/{meeting_id}/send-reminder")
def send_meeting_reminder(meeting_id: int, request: Request, db: Session = Depends(get_db)):
    """
    Manually send meeting reminder to all members of the meeting's department.
    """
    from scheduler import get_meeting_recipients, slot_to_time
    from notifications.dispatcher import dispatch_event
    from notifications.events import EventType

    meeting = db.query(Meeting).options(
        joinedload(Meeting.department)
    ).filter(Meeting.id == meeting_id).first()

    if not meeting:
        raise HTTPException(status_code=404, detail="Meeting not found")

    recipients = get_meeting_recipients(db, meeting)

    if not recipients:
        return {
            "success": False,
            "message": "No recipients found for this meeting",
            "emails_sent": 0
        }

    # Prepare meeting data
    meeting_data = {
        "title": meeting.title,
        "meeting_date": meeting.meeting_date,
        "start_time": slot_to_time(meeting.start_slot),
        "end_time": slot_to_time(meeting.end_slot),
        "location": meeting.location,
        "department_name": meeting.department.name if meeting.department else "All Leaders",
        "meeting_link": meeting.meeting_link,
        "description": meeting.description
    }

    # Dispatch reminder notification
    dispatch_event(
        db=db,
        event_type=EventType.MEETING_REMINDER,
        data=meeting_data,
        recipients=recipients
    )

    _log_admin_action(request, db, "send_meeting_reminder", "meeting", meeting_id,
                      f"Sent reminder for '{meeting.title}' to {len(recipients)} recipient(s)")
    db.commit()

    return {
        "success": True,
        "message": f"Meeting reminder sent to {len(recipients)} recipient(s)",
        "emails_sent": len(recipients)
    }


# ============ SERVICE PROGRAM ENDPOINTS ============

def _title_case_name(name: str) -> str:
    """Normalize a name to title case (first letter uppercase, rest lowercase per word)."""
    if not name:
        return name
    return " ".join(w.capitalize() for w in name.split())


def _get_titled_name(member: "Member") -> str:
    """Get member's full name with leadership title prefix (e.g., 'Pastor John Smith')"""
    name = _title_case_name(member.full_name)
    roles = member.leadership_roles or []
    if "pastor" in roles:
        return f"Pastor {name}"
    elif "elder" in roles:
        return f"Elder {name}"
    elif "deacon" in roles:
        return f"Deacon {name}"
    elif "dr" in roles:
        return f"Dr {name}"
    elif "mr" in roles:
        return f"Mr {name}"
    elif "mrs" in roles:
        return f"Mrs {name}"
    return name


def _program_to_dict(program: ServiceProgram, public: bool = False, db: Session = None) -> dict:
    """Convert a ServiceProgram model to response dict.
    If public=True, the Close marker item is stripped and its time is used
    to calculate the duration of the last real program item.
    If db is provided, participant names are enriched with leadership titles."""
    # Hash is the updated_at unix timestamp - changes on every edit
    ts = program.updated_at or program.created_at
    hashcode = str(int(ts.timestamp())) if ts else "0"

    def _parse_json(val):
        if isinstance(val, str):
            try:
                return json.loads(val)
            except (json.JSONDecodeError, TypeError):
                return []
        return val or []

    # Get creator info with title
    created_by_name = None
    created_by_id = getattr(program, 'created_by_member_id', None)
    if created_by_id and hasattr(program, 'created_by') and program.created_by:
        created_by_name = _get_titled_name(program.created_by)

    items = _parse_json(program.program_items)

    if public and items:
        # Find and remove the Close marker, use its time for last item duration
        close_item = None
        real_items = []
        for it in items:
            if isinstance(it, dict) and (it.get("item") or "").lower() == "close":
                close_item = it
            else:
                real_items.append(it)
        # Calculate duration for the last real item using close time
        if close_item and real_items and close_item.get("time"):
            last = real_items[-1]
            if last.get("time"):
                try:
                    from datetime import datetime as _dt
                    start = _dt.strptime(last["time"], "%H:%M")
                    end = _dt.strptime(close_item["time"], "%H:%M")
                    diff_min = int((end - start).total_seconds() // 60)
                    if diff_min > 0:
                        last["duration_minutes"] = diff_min
                except (ValueError, TypeError):
                    pass
        items = real_items

    participants = _parse_json(program.participants)

    # Enrich participant names with leadership titles
    if db and participants:
        # Collect unique plain names (strip existing titles for lookup)
        title_prefixes = ("pastor ", "elder ", "deacon ", "dr ", "mr ", "mrs ")
        plain_names = set()
        for pt in participants:
            name = (pt.get("name") or "").strip()
            name_lower = name.lower()
            # Strip existing title prefix for DB lookup
            for prefix in title_prefixes:
                if name_lower.startswith(prefix):
                    name = name[len(prefix):]
                    break
            if name:
                plain_names.add(name.lower())
        if plain_names:
            members = db.query(Member).filter(
                func.lower(Member.full_name).in_(list(plain_names))
            ).all()
            title_map = {m.full_name.lower(): _get_titled_name(m) for m in members}
            for pt in participants:
                name = (pt.get("name") or "").strip()
                name_lower = name.lower()
                # Strip existing title for lookup
                lookup_name = name_lower
                for prefix in title_prefixes:
                    if lookup_name.startswith(prefix):
                        lookup_name = lookup_name[len(prefix):]
                        break
                if lookup_name in title_map:
                    pt["name"] = title_map[lookup_name]
                else:
                    # Apply title case even for non-member names
                    pt["name"] = _title_case_name(name)

    return {
        "id": program.id,
        "hash": hashcode,
        "title": program.title,
        "service_date": program.service_date.isoformat(),
        "location_type": program.location_type or "onsite",
        "template_id": program.template_id,
        "program_items": items,
        "participants": participants,
        "admin_announcements": _parse_json(program.admin_announcements),
        "pastors_announcements": _parse_json(program.pastors_announcements),
        "prayer_points": _parse_json(program.prayer_points),
        "status": getattr(program, 'status', 'draft') or 'draft',
        "created_by_member_id": created_by_id,
        "created_by_name": created_by_name,
        "created_at": program.created_at.isoformat() if program.created_at else None,
        "updated_at": program.updated_at.isoformat() if program.updated_at else None
    }


def _cleanup_past_programs(db: Session):
    """Delete programs with service_date before today"""
    today = date.today()
    deleted = db.query(ServiceProgram).filter(ServiceProgram.service_date < today).delete()
    if deleted:
        db.commit()
        print(f"[Cleanup] Deleted {deleted} past service program(s)")
    return deleted


@router.get("/programs/mine")
def get_my_programs(phone: str, db: Session = Depends(get_db)):
    """Get programs where the member (by phone) is listed as a participant"""
    _cleanup_past_programs(db)

    if not phone:
        return []

    # Find all members with this phone
    normalized = phone.strip().replace(" ", "").replace("-", "")
    member_names = set()
    for m in db.query(Member).all():
        m_normalized = m.phone.strip().replace(" ", "").replace("-", "")
        if m_normalized == normalized or m.phone == phone:
            member_names.add(m.full_name.lower())

    if not member_names:
        return []

    # Get all upcoming programs and filter by participant name match
    from sqlalchemy.orm import joinedload
    all_programs = db.query(ServiceProgram).options(
        joinedload(ServiceProgram.created_by)
    ).order_by(ServiceProgram.service_date).all()

    result = []
    for p in all_programs:
        participants = json.loads(p.participants) if isinstance(p.participants, str) else (p.participants or [])
        for pt in participants:
            if pt.get("name", "").lower() in member_names:
                result.append(_program_to_dict(p, db=db))
                break

    return result


@router.get("/programs/today")
def get_todays_programs(db: Session = Depends(get_db)):
    """Public endpoint: get today's service program(s). Auto-cleans past programs."""
    _cleanup_past_programs(db)

    today = date.today()
    # Only return published onsite programs via public API
    programs = db.query(ServiceProgram).filter(
        ServiceProgram.service_date == today,
        ServiceProgram.location_type.in_(["onsite", None]),
        ServiceProgram.status == "published"
    ).order_by(ServiceProgram.id).all()

    return {
        "date": today.isoformat(),
        "programs": [_program_to_dict(p, public=True, db=db) for p in programs]
    }


@router.get("/admin/programs")
def get_all_programs(db: Session = Depends(get_db)):
    """Admin: list all programs (upcoming and today)"""
    _cleanup_past_programs(db)

    from sqlalchemy.orm import joinedload
    programs = db.query(ServiceProgram).options(joinedload(ServiceProgram.created_by)).order_by(ServiceProgram.service_date).all()
    return [_program_to_dict(p, db=db) for p in programs]


@router.get("/admin/programs/{program_id}")
def get_program(program_id: int, db: Session = Depends(get_db)):
    """Admin: get a single program by ID"""
    from sqlalchemy.orm import joinedload
    program = db.query(ServiceProgram).options(joinedload(ServiceProgram.created_by)).filter(ServiceProgram.id == program_id).first()
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")
    return _program_to_dict(program, db=db)


def _get_prayer_points_for_role(prayer_points: list, role: str) -> list:
    """Get prayer points linked to a specific activity/role.
    Returns list of dicts {text, linked_activity} to preserve the link info for display.
    Prayer points can be plain strings (no link) or dicts with {text, linked_activity}."""
    if not prayer_points or not role:
        return []
    result = []
    role_lower = role.lower()
    for pp in prayer_points:
        if isinstance(pp, dict):
            linked = (pp.get("linked_activity") or "").lower()
            if linked and linked == role_lower:
                text = pp.get("text", "")
                if text:
                    result.append({"text": text, "linked_activity": pp.get("linked_activity", "")})
    return result


def _notify_program_participants(db: Session, program_title: str, service_date, participant_names: list, participant_roles: dict, created_by_name: str = None, program_id: int = None, prayer_points: list = None, admin_announcements: list = None, pastors_announcements: list = None):
    """Notify participants who are members in the database about their program roles.
    participant_roles is a dict mapping lowercase name -> list of role strings.
    A person with multiple roles gets one email summarising all their roles.
    Admin-role participants get admin_announcements; Preacher-role participants get pastors_announcements."""
    from notifications.dispatcher import dispatch_event
    from notifications.events import EventType

    if not participant_names:
        return

    # Strip title prefixes for DB lookup
    title_prefixes = ("pastor ", "elder ", "deacon ", "dr ", "mr ", "mrs ")
    def _strip_title(name):
        lower = name.lower()
        for prefix in title_prefixes:
            if lower.startswith(prefix):
                return name[len(prefix):]
        return name

    # Build lookup names (both titled and plain) for matching
    plain_names = [_strip_title(n).lower() for n in participant_names]

    # Find members whose names match participants (case-insensitive, using plain names)
    members = db.query(Member).filter(
        func.lower(Member.full_name).in_(plain_names)
    ).all()

    if not members:
        return

    date_str = service_date.strftime("%A, %d %B %Y") if hasattr(service_date, 'strftime') else str(service_date)

    # Build a role lookup that works with both titled and plain name keys
    def _get_roles_for_member(member):
        name_lower = member.full_name.lower()
        titled_lower = _get_titled_name(member).lower()
        return (participant_roles.get(name_lower)
                or participant_roles.get(titled_lower)
                or ["Participant"])

    # Role keywords that qualify for announcements
    admin_keywords = {"admin", "administrator", "mc", "emcee"}
    preacher_keywords = {"preach", "preacher", "preaching", "sermon", "pastor", "word", "minister", "ministering"}

    for member in members:
        if not member.email:
            continue
        roles = _get_roles_for_member(member)
        if isinstance(roles, str):
            roles = [roles]

        # Collect prayer points linked to any of this participant's roles
        linked_prayer_points = []
        for role in roles:
            linked_prayer_points.extend(_get_prayer_points_for_role(prayer_points or [], role))
        # Deduplicate by text (prayer points are now dicts with text + linked_activity)
        seen_texts = set()
        unique_prayer_points = []
        for pp in linked_prayer_points:
            text = pp.get("text", "") if isinstance(pp, dict) else pp
            if text and text not in seen_texts:
                seen_texts.add(text)
                unique_prayer_points.append(pp)

        # Determine which announcements this participant should receive
        roles_lower = {r.lower() for r in roles}
        member_admin_ann = []
        member_pastor_ann = []
        if admin_announcements and any(kw in rl for rl in roles_lower for kw in admin_keywords):
            member_admin_ann = admin_announcements
        if pastors_announcements and any(kw in rl for rl in roles_lower for kw in preacher_keywords):
            member_pastor_ann = pastors_announcements

        titled_name = _get_titled_name(member)
        try:
            dispatch_event(
                db=db,
                event_type=EventType.PROGRAM_PARTICIPANT_ADDED,
                data={
                    "title": program_title,
                    "service_date": date_str,
                    "roles": roles,
                    "role": ", ".join(roles),
                    "member_name": titled_name,
                    "member_email": member.email,
                    "member_id": member.id,
                    "member_phone": member.phone,
                    "created_by": created_by_name or "Admin",
                    "program_id": program_id,
                    "prayer_points": unique_prayer_points,
                    "admin_announcements": member_admin_ann,
                    "pastors_announcements": member_pastor_ann,
                },
                recipients=[{
                    "id": member.id,
                    "name": titled_name,
                    "email": member.email,
                    "phone": member.phone,
                }]
            )
        except Exception as e:
            print(f"Failed to notify participant {member.full_name}: {e}")


@router.post("/admin/programs")
def create_program(data: ServiceProgramCreate, request: Request, db: Session = Depends(get_db)):
    """Admin: create a new service program"""
    if not data.title:
        raise HTTPException(status_code=400, detail="Title is required")
    if not data.program_items:
        raise HTTPException(status_code=400, detail="At least one program item is required")

    # Resolve who's creating this program. Priority order:
    #   1. Explicit created_by_member_id in the request payload (portal sends this)
    #   2. The admin identity cookie (set when an admin logs in via /admin/login)
    # Without this fallback, programs created from the admin form had no
    # `created_by_member_id` and the list view couldn't show a Service Manager.
    created_by_id = data.created_by_member_id
    if not created_by_id:
        try:
            from routers.pages import get_admin_identity
            identity = get_admin_identity(request)
            if identity and identity.get("member_id"):
                created_by_id = identity["member_id"]
        except Exception:
            pass

    program = ServiceProgram(
        title=data.title,
        service_date=data.service_date,
        location_type=data.location_type or "onsite",
        program_items=json.dumps([item.model_dump() for item in data.program_items]),
        participants=json.dumps([p.model_dump() for p in (data.participants or [])]),
        admin_announcements=json.dumps(data.admin_announcements or []),
        pastors_announcements=json.dumps(data.pastors_announcements or []),
        prayer_points=json.dumps(data.prayer_points or []),
        template_id=data.template_id,
        created_by_member_id=created_by_id
    )
    db.add(program)
    db.commit()
    db.refresh(program)
    # Eager-load creator for dict conversion
    if program.created_by_member_id:
        from sqlalchemy.orm import joinedload
        program = db.query(ServiceProgram).options(joinedload(ServiceProgram.created_by)).filter(ServiceProgram.id == program.id).first()

    _log_admin_action(request, db, "create_program", "program", program.id, f"Created program '{program.title}' for {program.service_date}")
    db.commit()

    # Emails are NOT sent on create — only when the program is published
    return _program_to_dict(program, db=db)


def _check_program_edit_permission(db: Session, program: ServiceProgram, editor_member_id: int = None):
    """Check if a member has permission to edit a program.
    Returns None if allowed, raises HTTPException if not.
    If editor_member_id is None (admin panel), always allows."""
    if editor_member_id is None:
        return  # Admin panel - no restrictions

    member = db.query(Member).filter(Member.id == editor_member_id).first()
    if not member:
        raise HTTPException(status_code=403, detail="Editor not found")

    roles = []
    if member.leadership_roles:
        try:
            roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else member.leadership_roles
        except (ValueError, TypeError):
            roles = []

    # Admin role has full access
    if "admin" in roles:
        return

    # Creator can edit their own program
    if program.created_by_member_id == editor_member_id:
        return

    raise HTTPException(status_code=403, detail="You can only edit programs you created")


@router.put("/admin/programs/{program_id}")
def update_program(program_id: int, data: ServiceProgramUpdate, request: Request, editor_member_id: int = None, db: Session = Depends(get_db)):
    """Admin: update an existing program"""
    program = db.query(ServiceProgram).filter(ServiceProgram.id == program_id).first()
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")

    _check_program_edit_permission(db, program, editor_member_id)

    if data.title is not None:
        program.title = data.title
    if data.service_date is not None:
        program.service_date = data.service_date
    if data.location_type is not None:
        program.location_type = data.location_type
    if data.program_items is not None:
        program.program_items = json.dumps([item.model_dump() for item in data.program_items])
    if data.participants is not None:
        program.participants = json.dumps([p.model_dump() for p in data.participants])
    if data.admin_announcements is not None:
        program.admin_announcements = json.dumps(data.admin_announcements)
    if data.pastors_announcements is not None:
        program.pastors_announcements = json.dumps(data.pastors_announcements)
    if data.prayer_points is not None:
        program.prayer_points = json.dumps(data.prayer_points)
    if data.template_id is not None:
        program.template_id = data.template_id

    db.commit()
    db.refresh(program)

    # Eager-load creator for dict conversion
    from sqlalchemy.orm import joinedload
    program = db.query(ServiceProgram).options(joinedload(ServiceProgram.created_by)).filter(ServiceProgram.id == program.id).first()

    _log_admin_action(request, db, "update_program", "program", program_id, f"Updated program '{program.title}'")
    db.commit()

    # Emails are NOT sent on update — use publish or re-notify instead
    return _program_to_dict(program, db=db)


@router.delete("/admin/programs/{program_id}")
def delete_program(program_id: int, request: Request, editor_member_id: int = None, db: Session = Depends(get_db)):
    """Admin: delete a program"""
    program = db.query(ServiceProgram).filter(ServiceProgram.id == program_id).first()
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")

    _check_program_edit_permission(db, program, editor_member_id)

    prog_title = program.title
    db.delete(program)
    db.commit()

    _log_admin_action(request, db, "delete_program", "program", program_id, f"Deleted program '{prog_title}'")
    db.commit()

    return {"success": True}


def _send_program_notifications(db: Session, program: ServiceProgram):
    """Send email notifications to all participants of a program."""
    participants = json.loads(program.participants) if isinstance(program.participants, str) else (program.participants or [])
    if not participants:
        return

    names = list({p.get("name", "") for p in participants})
    roles = {}
    for p in participants:
        roles.setdefault((p.get("name") or "").lower(), []).append(p.get("role", ""))

    from sqlalchemy.orm import joinedload
    prog = db.query(ServiceProgram).options(joinedload(ServiceProgram.created_by)).filter(ServiceProgram.id == program.id).first()
    creator_name = _get_titled_name(prog.created_by) if prog and prog.created_by_member_id and prog.created_by else None

    pp = json.loads(program.prayer_points or "[]") if isinstance(program.prayer_points, str) else (program.prayer_points or [])
    admin_ann = json.loads(program.admin_announcements or "[]") if isinstance(program.admin_announcements, str) else (program.admin_announcements or [])
    pastor_ann = json.loads(program.pastors_announcements or "[]") if isinstance(program.pastors_announcements, str) else (program.pastors_announcements or [])

    _notify_program_participants(
        db, program.title, program.service_date, names, roles,
        created_by_name=creator_name, program_id=program.id,
        prayer_points=pp, admin_announcements=admin_ann, pastors_announcements=pastor_ann
    )


def _send_published_copy_to_manager(db: Session, program: ServiceProgram):
    """Send the service manager (creator) a copy of the finalised program."""
    if not program.created_by_member_id:
        return
    from sqlalchemy.orm import joinedload
    prog = db.query(ServiceProgram).options(joinedload(ServiceProgram.created_by)).filter(ServiceProgram.id == program.id).first()
    if not prog or not prog.created_by or not prog.created_by.email:
        return

    manager = prog.created_by
    manager_name = _get_titled_name(manager)
    manager_email = manager.email

    # Build program summary
    items = json.loads(program.program_items or "[]") if isinstance(program.program_items, str) else (program.program_items or [])
    participants = json.loads(program.participants or "[]") if isinstance(program.participants, str) else (program.participants or [])
    admin_ann = json.loads(program.admin_announcements or "[]") if isinstance(program.admin_announcements, str) else (program.admin_announcements or [])
    pastor_ann = json.loads(program.pastors_announcements or "[]") if isinstance(program.pastors_announcements, str) else (program.pastors_announcements or [])
    prayer_pts = json.loads(program.prayer_points or "[]") if isinstance(program.prayer_points, str) else (program.prayer_points or [])

    svc_date = program.service_date
    if hasattr(svc_date, 'strftime'):
        date_str = svc_date.strftime('%A, %d %B %Y')
    else:
        date_str = str(svc_date)

    FONT = "-apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif"
    TEXT = "#111827"
    MUTED = "#9ca3af"
    BORDER = "#f3f4f6"

    # Build order of service rows
    items_html = ""
    for it in items:
        t = it.get("time", "")
        label = it.get("item", "")
        if label.lower() == "close":
            continue
        # Format time
        if t:
            try:
                from datetime import datetime as _dt
                items_html += f'<tr><td style="padding:4px 12px 4px 0;color:{MUTED};font-family:monospace;font-size:13px;vertical-align:top;white-space:nowrap;">{_dt.strptime(t, "%H:%M").strftime("%I:%M %p").lstrip("0")}</td><td style="padding:4px 0;color:{TEXT};font-size:14px;">{label}</td></tr>'
            except ValueError:
                items_html += f'<tr><td style="padding:4px 12px 4px 0;color:{MUTED};font-family:monospace;font-size:13px;vertical-align:top;">{t}</td><td style="padding:4px 0;color:{TEXT};font-size:14px;">{label}</td></tr>'
        else:
            items_html += f'<tr><td></td><td style="padding:4px 0;color:{TEXT};font-size:14px;">{label}</td></tr>'

    # Build participants rows
    parts_html = ""
    for pt in participants:
        parts_html += f'<tr><td style="padding:3px 12px 3px 0;color:{MUTED};font-size:13px;vertical-align:top;white-space:nowrap;">{pt.get("role", "")}</td><td style="padding:3px 0;color:{TEXT};font-size:14px;">{pt.get("name", "")}</td></tr>'

    # Build announcements
    ann_html = ""
    if admin_ann:
        ann_html += f'<p style="margin:16px 0 6px 0;font-size:11px;font-weight:600;color:{MUTED};text-transform:uppercase;letter-spacing:0.5px;">Admin Announcements</p>'
        for a in admin_ann:
            ann_html += f'<p style="margin:0 0 4px 0;color:{TEXT};font-size:14px;padding-left:12px;border-left:2px solid {BORDER};">{a}</p>'
    if pastor_ann:
        ann_html += f'<p style="margin:16px 0 6px 0;font-size:11px;font-weight:600;color:{MUTED};text-transform:uppercase;letter-spacing:0.5px;">Pastor\'s Announcements</p>'
        for a in pastor_ann:
            ann_html += f'<p style="margin:0 0 4px 0;color:{TEXT};font-size:14px;padding-left:12px;border-left:2px solid {BORDER};">{a}</p>'

    # Build prayer points
    prayer_html = ""
    if prayer_pts:
        prayer_html += f'<p style="margin:16px 0 6px 0;font-size:11px;font-weight:600;color:{MUTED};text-transform:uppercase;letter-spacing:0.5px;">Prayer Points</p>'
        for pp in prayer_pts:
            if isinstance(pp, dict):
                text = pp.get("text", "")
                linked = pp.get("linked_activity", "")
                linked_tag = f' <span style="font-size:12px;color:{MUTED};font-style:italic;">({linked})</span>' if linked else ""
                prayer_html += f'<p style="margin:0 0 4px 0;color:{TEXT};font-size:14px;padding-left:12px;border-left:2px solid {BORDER};">{text}{linked_tag}</p>'
            elif pp:
                prayer_html += f'<p style="margin:0 0 4px 0;color:{TEXT};font-size:14px;padding-left:12px;border-left:2px solid {BORDER};">{pp}</p>'

    html = f'''<!DOCTYPE html>
<html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#ffffff;font-family:{FONT};">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr><td align="center" style="padding:32px 16px;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="max-width:520px;">
  <tr><td>
    <h2 style="margin:0 0 4px 0;color:{TEXT};font-size:18px;font-weight:700;">{program.title}</h2>
    <p style="margin:0 0 6px 0;color:{MUTED};font-size:14px;">{date_str}</p>
    <p style="margin:0 0 20px 0;color:#059669;font-size:13px;font-weight:600;">Published — notifications sent to all participants</p>

    <p style="margin:0 0 6px 0;font-size:11px;font-weight:600;color:{MUTED};text-transform:uppercase;letter-spacing:0.5px;">Order of Service</p>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-bottom:8px;">{items_html}</table>

    <hr style="border:none;border-top:1px solid {BORDER};margin:16px 0;">
    <p style="margin:0 0 6px 0;font-size:11px;font-weight:600;color:{MUTED};text-transform:uppercase;letter-spacing:0.5px;">Participants</p>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0">{parts_html}</table>

    {ann_html}
    {prayer_html}

    <hr style="border:none;border-top:1px solid {BORDER};margin:20px 0 16px 0;">
    <p style="margin:0;color:{MUTED};font-size:12px;">Revival Fire Ministries &middot; Stellenbosch</p>
  </td></tr>
</table>
</td></tr></table>
</body></html>'''

    subject = f"Program Published: {program.title} — {date_str}"

    # Send via rfm-notify
    try:
        from notifications.channels.rfm_notify import RfmNotifyChannel
        channel = RfmNotifyChannel()
        if channel.is_configured():
            success, error = channel.send(
                manager_email, subject, html,
                event_code="program.published_manager_copy",
                recipient_name=manager_name,
                idempotency_key=f"program_published:{program.id}:{manager_email}",
            )
            if success:
                print(f"Sent published program copy to service manager {manager_name} ({manager_email})")
            else:
                print(f"Failed to send published copy to {manager_email}: {error}")
        else:
            print("rfm-notify not configured — skipping program-published manager copy")
    except Exception as e:
        print(f"Failed to email program copy to service manager: {e}")


@router.post("/admin/programs/{program_id}/publish")
def publish_program(program_id: int, request: Request, editor_member_id: int = None, db: Session = Depends(get_db)):
    """Admin: publish a program and send notifications to all participants"""
    program = db.query(ServiceProgram).filter(ServiceProgram.id == program_id).first()
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")

    _check_program_edit_permission(db, program, editor_member_id)

    program.status = "published"
    db.commit()
    db.refresh(program)

    # Send notifications to all participants
    _send_program_notifications(db, program)

    # Send service manager a copy of the final program
    _send_published_copy_to_manager(db, program)

    _log_admin_action(request, db, "publish_program", "program", program_id, f"Published program '{program.title}'")
    db.commit()

    return _program_to_dict(program, db=db)


@router.post("/admin/programs/{program_id}/unpublish")
def unpublish_program(program_id: int, request: Request, editor_member_id: int = None, db: Session = Depends(get_db)):
    """Admin: unpublish a program (back to draft)"""
    program = db.query(ServiceProgram).filter(ServiceProgram.id == program_id).first()
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")

    _check_program_edit_permission(db, program, editor_member_id)

    program.status = "draft"
    db.commit()
    db.refresh(program)

    _log_admin_action(request, db, "unpublish_program", "program", program_id, f"Unpublished program '{program.title}'")
    db.commit()

    return _program_to_dict(program, db=db)


@router.post("/admin/programs/{program_id}/notify")
def renotify_program(program_id: int, request: Request, editor_member_id: int = None, db: Session = Depends(get_db)):
    """Admin: re-send notifications to all participants of a published program"""
    program = db.query(ServiceProgram).filter(ServiceProgram.id == program_id).first()
    if not program:
        raise HTTPException(status_code=404, detail="Program not found")

    _check_program_edit_permission(db, program, editor_member_id)

    _send_program_notifications(db, program)

    _log_admin_action(request, db, "notify_program", "program", program_id, f"Re-sent notifications for program '{program.title}'")
    db.commit()

    return {"success": True, "message": "Notifications sent to all participants"}


# ============ PROGRAM TEMPLATE ENDPOINTS ============

DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _template_to_dict(template: ProgramTemplate) -> dict:
    """Convert a ProgramTemplate model to response dict"""
    ts = template.updated_at or template.created_at
    hashcode = str(int(ts.timestamp())) if ts else "0"

    def _parse_json(val):
        if isinstance(val, str):
            try:
                return json.loads(val)
            except (json.JSONDecodeError, TypeError):
                return []
        return val or []

    return {
        "id": template.id,
        "hash": hashcode,
        "title": template.title,
        "day_of_week": template.day_of_week,
        "day_name": DAY_NAMES[template.day_of_week] if 0 <= template.day_of_week <= 6 else "Unknown",
        "program_items": _parse_json(template.program_items),
        "participants": _parse_json(template.participants),
        "admin_announcements": _parse_json(template.admin_announcements),
        "pastors_announcements": _parse_json(template.pastors_announcements),
        "prayer_points": _parse_json(template.prayer_points),
        "support_roles": _parse_json(template.support_roles),
        "location_type": template.location_type or "onsite",
        "created_at": template.created_at.isoformat() if template.created_at else None,
        "updated_at": template.updated_at.isoformat() if template.updated_at else None
    }


@router.get("/programs/templates")
def get_public_templates(day: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """Public endpoint: get program templates, optionally filtered by day name or number.
    Examples: /api/programs/templates?day=sunday, /api/programs/templates?day=6
    If no day param, returns all templates."""
    query = db.query(ProgramTemplate)

    if day is not None:
        # Try as day name first
        day_lower = day.strip().lower()
        day_num = None
        for i, name in enumerate(DAY_NAMES):
            if name.lower() == day_lower:
                day_num = i
                break
        # Try as number
        if day_num is None:
            try:
                day_num = int(day)
            except ValueError:
                pass

        if day_num is not None and 0 <= day_num <= 6:
            query = query.filter(ProgramTemplate.day_of_week == day_num)
        else:
            return {"templates": []}

    templates = query.order_by(ProgramTemplate.day_of_week, ProgramTemplate.title).all()
    return {"templates": [_template_to_dict(t) for t in templates]}


@router.get("/programs/templates/today")
def get_todays_templates(db: Session = Depends(get_db)):
    """Public endpoint: get templates for today's day of week."""
    today_dow = date.today().weekday()  # 0=Monday ... 6=Sunday
    templates = db.query(ProgramTemplate).filter(
        ProgramTemplate.day_of_week == today_dow
    ).order_by(ProgramTemplate.title).all()

    return {
        "date": date.today().isoformat(),
        "day_of_week": today_dow,
        "day_name": DAY_NAMES[today_dow],
        "templates": [_template_to_dict(t) for t in templates]
    }


@router.get("/admin/templates")
def get_all_templates(db: Session = Depends(get_db)):
    """Admin: list all program templates"""
    templates = db.query(ProgramTemplate).order_by(
        ProgramTemplate.day_of_week, ProgramTemplate.title
    ).all()
    return [_template_to_dict(t) for t in templates]


@router.get("/admin/templates/{template_id}")
def get_template(template_id: int, db: Session = Depends(get_db)):
    """Admin: get a single template"""
    template = db.query(ProgramTemplate).filter(ProgramTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return _template_to_dict(template)


@router.post("/admin/templates")
def create_template(data: ProgramTemplateCreate, request: Request, db: Session = Depends(get_db)):
    """Admin: create a new program template"""
    if not data.title:
        raise HTTPException(status_code=400, detail="Title is required")
    if data.day_of_week < 0 or data.day_of_week > 6:
        raise HTTPException(status_code=400, detail="day_of_week must be 0 (Monday) to 6 (Sunday)")

    template = ProgramTemplate(
        title=data.title,
        day_of_week=data.day_of_week,
        location_type=data.location_type or "onsite",
        program_items=json.dumps([item.model_dump() for item in (data.program_items or [])]),
        participants=json.dumps([p.model_dump() for p in (data.participants or [])]),
        admin_announcements=json.dumps(data.admin_announcements or []),
        pastors_announcements=json.dumps(data.pastors_announcements or []),
        prayer_points=json.dumps(data.prayer_points or []),
        support_roles=json.dumps(data.support_roles or [])
    )
    db.add(template)
    db.commit()
    db.refresh(template)

    _log_admin_action(request, db, "create_template", "template", template.id, f"Created template '{template.title}'")
    db.commit()

    return _template_to_dict(template)


@router.put("/admin/templates/{template_id}")
def update_template(template_id: int, data: ProgramTemplateUpdate, request: Request, db: Session = Depends(get_db)):
    """Admin: update a template"""
    template = db.query(ProgramTemplate).filter(ProgramTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    if data.title is not None:
        template.title = data.title
    if data.day_of_week is not None:
        if data.day_of_week < 0 or data.day_of_week > 6:
            raise HTTPException(status_code=400, detail="day_of_week must be 0 (Monday) to 6 (Sunday)")
        template.day_of_week = data.day_of_week
    if data.location_type is not None:
        template.location_type = data.location_type
    if data.program_items is not None:
        template.program_items = json.dumps([item.model_dump() for item in data.program_items])
    if data.participants is not None:
        template.participants = json.dumps([p.model_dump() for p in data.participants])
    if data.admin_announcements is not None:
        template.admin_announcements = json.dumps(data.admin_announcements)
    if data.pastors_announcements is not None:
        template.pastors_announcements = json.dumps(data.pastors_announcements)
    if data.prayer_points is not None:
        template.prayer_points = json.dumps(data.prayer_points)
    if data.support_roles is not None:
        template.support_roles = json.dumps(data.support_roles)

    db.commit()
    db.refresh(template)

    _log_admin_action(request, db, "update_template", "template", template_id, f"Updated template '{template.title}'")
    db.commit()

    return _template_to_dict(template)


@router.delete("/admin/templates/{template_id}")
def delete_template(template_id: int, request: Request, db: Session = Depends(get_db)):
    """Admin: delete a template"""
    template = db.query(ProgramTemplate).filter(ProgramTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")

    tmpl_title = template.title
    db.delete(template)
    db.commit()

    _log_admin_action(request, db, "delete_template", "template", template_id, f"Deleted template '{tmpl_title}'")
    db.commit()

    return {"success": True}


# ============ SERVICE SCHEDULE ENDPOINTS ============

DAY_LABELS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


def _schedule_to_dict(schedule: ServiceSchedule) -> dict:
    """Convert a ServiceSchedule model to response dict."""
    svc_date = schedule.service_date
    template = schedule.template
    manager = schedule.service_manager
    program = schedule.program

    return {
        "id": schedule.id,
        "service_date": svc_date.isoformat(),
        "day_of_week": DAY_LABELS[svc_date.weekday()] if svc_date else None,
        "template_id": schedule.template_id,
        "template_name": template.title if template else None,
        "service_manager_id": schedule.service_manager_id,
        "service_manager_name": _get_titled_name(manager) if manager else None,
        "notes": schedule.notes,
        "notified_at": schedule.notified_at.isoformat() if schedule.notified_at else None,
        "reminded_at": schedule.reminded_at.isoformat() if schedule.reminded_at else None,
        "program_id": schedule.program_id,
        "program_status": program.status if program else None,
        "has_program": schedule.program_id is not None,
        "created_at": schedule.created_at.isoformat() if schedule.created_at else None,
    }


@router.get("/admin/schedules")
def get_schedules(db: Session = Depends(get_db)):
    """Admin: list all upcoming service schedules"""
    today = date.today()
    schedules = db.query(ServiceSchedule).options(
        joinedload(ServiceSchedule.template),
        joinedload(ServiceSchedule.service_manager),
        joinedload(ServiceSchedule.program)
    ).filter(ServiceSchedule.service_date >= today).order_by(ServiceSchedule.service_date).all()
    return [_schedule_to_dict(s) for s in schedules]


@router.get("/admin/schedules/all")
def get_all_schedules(
    from_date: str = None,
    to_date: str = None,
    template_id: int = None,
    manager_id: int = None,
    status: str = None,
    db: Session = Depends(get_db)
):
    """Admin: list all service schedules with optional filters"""
    query = db.query(ServiceSchedule).options(
        joinedload(ServiceSchedule.template),
        joinedload(ServiceSchedule.service_manager),
        joinedload(ServiceSchedule.program)
    )

    if from_date:
        query = query.filter(ServiceSchedule.service_date >= from_date)
    if to_date:
        query = query.filter(ServiceSchedule.service_date <= to_date)
    if template_id:
        query = query.filter(ServiceSchedule.template_id == template_id)
    if manager_id:
        query = query.filter(ServiceSchedule.service_manager_id == manager_id)
    if status == "has_program":
        query = query.filter(ServiceSchedule.program_id != None)
    elif status == "no_program":
        query = query.filter(ServiceSchedule.program_id == None)
    elif status == "published":
        query = query.join(ServiceProgram, ServiceSchedule.program_id == ServiceProgram.id).filter(ServiceProgram.status == "published")
    elif status == "draft":
        query = query.join(ServiceProgram, ServiceSchedule.program_id == ServiceProgram.id).filter(ServiceProgram.status == "draft")

    schedules = query.order_by(ServiceSchedule.service_date.desc()).all()
    return [_schedule_to_dict(s) for s in schedules]


@router.post("/admin/schedules")
def create_schedule(data: ServiceScheduleCreate, request: Request, db: Session = Depends(get_db)):
    """Admin: schedule a future service"""
    # Validate service manager exists
    manager = db.query(Member).filter(Member.id == data.service_manager_id).first()
    if not manager:
        raise HTTPException(status_code=404, detail="Service manager not found")

    # Validate template if provided
    if data.template_id:
        template = db.query(ProgramTemplate).filter(ProgramTemplate.id == data.template_id).first()
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")

    # Check for duplicate date
    existing = db.query(ServiceSchedule).filter(ServiceSchedule.service_date == data.service_date).first()
    if existing:
        raise HTTPException(status_code=400, detail="A service is already scheduled for this date")

    schedule = ServiceSchedule(
        service_date=data.service_date,
        template_id=data.template_id,
        service_manager_id=data.service_manager_id,
        notes=data.notes
    )
    db.add(schedule)
    db.commit()
    db.refresh(schedule)

    # Eager-load relationships
    schedule = db.query(ServiceSchedule).options(
        joinedload(ServiceSchedule.template),
        joinedload(ServiceSchedule.service_manager),
        joinedload(ServiceSchedule.program)
    ).filter(ServiceSchedule.id == schedule.id).first()

    _log_admin_action(request, db, "create_schedule", "schedule", schedule.id, f"Scheduled service for {data.service_date}")
    db.commit()

    return _schedule_to_dict(schedule)


@router.put("/admin/schedules/{schedule_id}")
def update_schedule(schedule_id: int, data: ServiceScheduleUpdate, request: Request, db: Session = Depends(get_db)):
    """Admin: update a service schedule"""
    schedule = db.query(ServiceSchedule).filter(ServiceSchedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    if data.service_date is not None:
        # Check duplicate
        existing = db.query(ServiceSchedule).filter(
            ServiceSchedule.service_date == data.service_date,
            ServiceSchedule.id != schedule_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="A service is already scheduled for this date")
        schedule.service_date = data.service_date

    if data.template_id is not None:
        schedule.template_id = data.template_id
    if data.service_manager_id is not None:
        manager = db.query(Member).filter(Member.id == data.service_manager_id).first()
        if not manager:
            raise HTTPException(status_code=404, detail="Service manager not found")
        schedule.service_manager_id = data.service_manager_id
        # Reset notification state if manager changed
        schedule.notified_at = None
        schedule.reminded_at = None
    if data.notes is not None:
        schedule.notes = data.notes

    db.commit()
    db.refresh(schedule)

    schedule = db.query(ServiceSchedule).options(
        joinedload(ServiceSchedule.template),
        joinedload(ServiceSchedule.service_manager),
        joinedload(ServiceSchedule.program)
    ).filter(ServiceSchedule.id == schedule.id).first()

    _log_admin_action(request, db, "update_schedule", "schedule", schedule_id, f"Updated schedule for {schedule.service_date}")
    db.commit()

    return _schedule_to_dict(schedule)


@router.delete("/admin/schedules/{schedule_id}")
def delete_schedule(schedule_id: int, request: Request, db: Session = Depends(get_db)):
    """Admin: delete a service schedule"""
    schedule = db.query(ServiceSchedule).filter(ServiceSchedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    svc_date = str(schedule.service_date)
    db.delete(schedule)
    db.commit()

    _log_admin_action(request, db, "delete_schedule", "schedule", schedule_id, f"Deleted schedule for {svc_date}")
    db.commit()

    return {"success": True}


@router.get("/api/my-schedules")
def get_my_schedules(phone: str, db: Session = Depends(get_db)):
    """Get upcoming scheduled services for a service manager by phone"""
    members = db.query(Member).filter(Member.phone == phone).all()
    if not members:
        return []
    member_ids = [m.id for m in members]
    today = date.today()
    schedules = db.query(ServiceSchedule).options(
        joinedload(ServiceSchedule.template),
        joinedload(ServiceSchedule.service_manager),
        joinedload(ServiceSchedule.program)
    ).filter(
        ServiceSchedule.service_manager_id.in_(member_ids),
        ServiceSchedule.service_date >= today
    ).order_by(ServiceSchedule.service_date).all()
    return [_schedule_to_dict(s) for s in schedules]



# ============================================================================
# HOME CHURCH ROSTER
# ============================================================================

# Home Church dept matching — case-insensitive, tolerates variants like
# "Home Church Leadership" vs "Home Church Leaders", "Home Church committee"
# (lowercase c). We match by keyword so church renaming doesn't break things.
HOME_CHURCH_ROLE_KEYWORDS = {
    # role -> list of keywords that must appear (any one) alongside "home church"
    "committee": ["committee"],
    "leaders": ["leader", "leadership"],
    "preachers": ["preacher", "preaching"],
}


def _home_church_dept_ids(db: Session, role: str) -> List[int]:
    """Return department IDs for a given home-church role (committee/leaders/preachers).
    Case-insensitive match against 'home church <keyword>'."""
    keywords = HOME_CHURCH_ROLE_KEYWORDS.get(role, [])
    if not keywords:
        return []
    # Fetch all departments once and filter in Python so we can do case-insensitive
    # substring matching that works across SQLite/Postgres collations.
    ids = []
    for d in db.query(Department).all():
        name = (d.name or "").lower()
        if "home church" in name and any(k in name for k in keywords):
            ids.append(d.id)
    return ids


def _preacher_pool_dept_ids(db: Session) -> List[int]:
    """Preachers may come from any of the three Home Church departments."""
    seen = set()
    out = []
    for role in ("preachers", "committee", "leaders"):
        for d_id in _home_church_dept_ids(db, role):
            if d_id not in seen:
                seen.add(d_id)
                out.append(d_id)
    return out


def _require_hc_access(request: Request, db: Session, hc: HomeChurch) -> Optional[Member]:
    """Admins, committee members, AND home church leaders can manage members.
      - admin (admin_session cookie) -> any home church, no restrictions
      - committee member             -> any home church
      - home church leader           -> only the home church they lead

    Returns the acting Member when one is resolvable (used for audit logs),
    or None for an admin whose identity isn't tied to a local Member row.
    Raises 401/403 when access is denied."""
    from routers.pages import get_admin_identity, is_authenticated, MEMBER_COOKIE_NAME, _verify_member_session

    # Admin path — always allowed, regardless of whether a local Member row exists
    if is_authenticated(request):
        identity = get_admin_identity(request)
        if identity and identity.get("member_id"):
            m = db.query(Member).filter(Member.id == identity["member_id"]).first()
            if m:
                return m
        return None  # admin without a mapped local member — still allowed

    token = request.cookies.get(MEMBER_COOKIE_NAME)
    member_id = _verify_member_session(token)
    if not member_id:
        raise HTTPException(status_code=401, detail="Not authenticated")
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=401, detail="Not authenticated")
    if _is_committee_member(db, member):
        return member
    if hc.leader_member_id == member.id:
        return member
    raise HTTPException(status_code=403, detail="You don't have access to this home church")


def _is_committee_member(db: Session, member: Member) -> bool:
    if not member:
        return False
    dept_ids = _home_church_dept_ids(db, "committee")
    if not dept_ids:
        return False
    count = db.query(MemberDepartment).filter(
        MemberDepartment.member_id == member.id,
        MemberDepartment.department_id.in_(dept_ids),
        MemberDepartment.status == "approved",
    ).count()
    return count > 0


def _require_committee_or_admin(request: Request, db: Session) -> Optional[Member]:
    """Return acting member or raise 401/403. Admins pass through; otherwise requires committee."""
    from routers.pages import get_admin_identity, is_authenticated
    if is_authenticated(request):
        identity = get_admin_identity(request)
        if identity and identity.get("member_id"):
            m = db.query(Member).filter(Member.id == identity["member_id"]).first()
            if m:
                return m
        return None
    from routers.pages import MEMBER_COOKIE_NAME, _verify_member_session
    token = request.cookies.get(MEMBER_COOKIE_NAME)
    member_id = _verify_member_session(token)
    if not member_id:
        raise HTTPException(status_code=401, detail="Not authenticated")
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member or not _is_committee_member(db, member):
        raise HTTPException(status_code=403, detail="Home Church Committee access required")
    return member


def _preacher_pool_member_ids(db: Session) -> List[int]:
    dept_ids = _preacher_pool_dept_ids(db)
    if not dept_ids:
        return []
    rows = db.query(MemberDepartment.member_id).filter(
        MemberDepartment.department_id.in_(dept_ids),
        MemberDepartment.status == "approved",
    ).distinct().all()
    return [r[0] for r in rows]


def _home_church_to_dict(hc: HomeChurch, db: Session) -> dict:
    leader = None
    if hc.leader_member_id:
        lm = db.query(Member).filter(Member.id == hc.leader_member_id).first()
        if lm:
            leader = {"id": lm.id, "full_name": lm.full_name, "phone": lm.phone, "email": lm.email}
    return {
        "id": hc.id,
        "name": hc.name,
        "leader_member_id": hc.leader_member_id,
        "leader": leader,
        "address": hc.address,
        "suburb": hc.suburb,
        "meeting_day": hc.meeting_day,
        "meeting_time": hc.meeting_time,
        "whatsapp_link": hc.whatsapp_link,
        "notes": hc.notes,
        "is_active": hc.is_active,
    }


def _program_type_to_dict(pt: HomeChurchProgramType) -> dict:
    return {
        "id": pt.id,
        "name": pt.name,
        "requires_preacher": pt.requires_preacher,
        "color": pt.color,
        "icon": pt.icon,
        "sort_order": pt.sort_order,
        "is_active": pt.is_active,
    }


def _roster_entry_to_dict(entry: HomeChurchRoster, db: Session) -> dict:
    preacher = None
    if entry.preacher_member_id:
        pm = db.query(Member).filter(Member.id == entry.preacher_member_id).first()
        if pm:
            preacher = {"id": pm.id, "full_name": pm.full_name, "phone": pm.phone, "email": pm.email}
    program_type = None
    if entry.program_type_id:
        pt = db.query(HomeChurchProgramType).filter(HomeChurchProgramType.id == entry.program_type_id).first()
        if pt:
            program_type = _program_type_to_dict(pt)
    return {
        "id": entry.id,
        "home_church_id": entry.home_church_id,
        "roster_date": entry.roster_date.isoformat() if entry.roster_date else None,
        "program_type_id": entry.program_type_id,
        "program_type": program_type,
        "preacher_member_id": entry.preacher_member_id,
        "preacher": preacher,
        "notes": entry.notes,
        "status": entry.status,
        "published_at": entry.published_at.isoformat() if entry.published_at else None,
    }


def _next_weekday(d: Optional[date] = None, weekday: int = 0) -> date:
    if d is None:
        d = date.today()
    delta = (weekday - d.weekday()) % 7
    return d + timedelta(days=delta)


# ---- PROGRAM TYPES ----

@router.get("/home-church/program-types")
def list_program_types_public(db: Session = Depends(get_db)):
    types = db.query(HomeChurchProgramType).filter(
        HomeChurchProgramType.is_active == True
    ).order_by(HomeChurchProgramType.sort_order).all()
    return [_program_type_to_dict(t) for t in types]


@router.get("/admin/home-church/program-types")
def admin_list_program_types(request: Request, db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    types = db.query(HomeChurchProgramType).order_by(HomeChurchProgramType.sort_order).all()
    return [_program_type_to_dict(t) for t in types]


@router.post("/admin/home-church/program-types")
def admin_create_program_type(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Name required")
    pt = HomeChurchProgramType(
        name=name,
        requires_preacher=bool(data.get("requires_preacher", False)),
        color=data.get("color") or "gray",
        icon=data.get("icon") or "\U0001F4CC",
        sort_order=int(data.get("sort_order", 99)),
        is_active=bool(data.get("is_active", True)),
    )
    db.add(pt)
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=409, detail="A program type with that name already exists")
    db.refresh(pt)
    _log_admin_action(request, db, "create_home_church_program_type", "home_church_program_type", pt.id, f"Created program type '{pt.name}'")
    db.commit()
    return _program_type_to_dict(pt)


@router.put("/admin/home-church/program-types/{pt_id}")
def admin_update_program_type(pt_id: int, request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    pt = db.query(HomeChurchProgramType).filter(HomeChurchProgramType.id == pt_id).first()
    if not pt:
        raise HTTPException(status_code=404, detail="Program type not found")
    for field in ("name", "color", "icon"):
        if field in data and data[field] is not None:
            setattr(pt, field, data[field])
    if "requires_preacher" in data:
        pt.requires_preacher = bool(data["requires_preacher"])
    if "sort_order" in data:
        pt.sort_order = int(data["sort_order"])
    if "is_active" in data:
        pt.is_active = bool(data["is_active"])
    db.commit()
    _log_admin_action(request, db, "update_home_church_program_type", "home_church_program_type", pt.id, f"Updated program type '{pt.name}'")
    db.commit()
    return _program_type_to_dict(pt)


@router.delete("/admin/home-church/program-types/{pt_id}")
def admin_delete_program_type(pt_id: int, request: Request, db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    pt = db.query(HomeChurchProgramType).filter(HomeChurchProgramType.id == pt_id).first()
    if not pt:
        raise HTTPException(status_code=404, detail="Program type not found")
    name = pt.name
    used = db.query(HomeChurchRoster).filter(HomeChurchRoster.program_type_id == pt_id).count()
    if used > 0:
        pt.is_active = False
        db.commit()
        _log_admin_action(request, db, "deactivate_home_church_program_type", "home_church_program_type", pt_id, f"Deactivated '{name}' (referenced by {used} roster entries)")
        db.commit()
        return {"success": True, "deactivated": True}
    db.delete(pt)
    db.commit()
    _log_admin_action(request, db, "delete_home_church_program_type", "home_church_program_type", pt_id, f"Deleted program type '{name}'")
    db.commit()
    return {"success": True, "deleted": True}


# ---- HOME CHURCHES ----

@router.get("/admin/home-churches")
def admin_list_home_churches(request: Request, db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    churches = db.query(HomeChurch).order_by(HomeChurch.name).all()
    return [_home_church_to_dict(c, db) for c in churches]


@router.post("/admin/home-churches")
def admin_create_home_church(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Home church name is required")
    hc = HomeChurch(
        name=name,
        leader_member_id=data.get("leader_member_id"),
        address=data.get("address"),
        suburb=data.get("suburb"),
        meeting_day=int(data.get("meeting_day", 0)),
        meeting_time=data.get("meeting_time") or "19:00",
        whatsapp_link=data.get("whatsapp_link"),
        notes=data.get("notes"),
        is_active=bool(data.get("is_active", True)),
    )
    db.add(hc)
    db.commit()
    db.refresh(hc)
    _log_admin_action(request, db, "create_home_church", "home_church", hc.id, f"Created home church '{hc.name}'")
    db.commit()
    _push_home_church_to_central(hc, request, db, action="create")
    return _home_church_to_dict(hc, db)


@router.put("/admin/home-churches/{hc_id}")
def admin_update_home_church(hc_id: int, request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    hc = db.query(HomeChurch).filter(HomeChurch.id == hc_id).first()
    if not hc:
        raise HTTPException(status_code=404, detail="Home church not found")
    for field in ("name", "address", "suburb", "meeting_time", "whatsapp_link", "notes"):
        if field in data:
            setattr(hc, field, data[field])
    if "leader_member_id" in data:
        hc.leader_member_id = data["leader_member_id"]
    if "meeting_day" in data and data["meeting_day"] is not None:
        hc.meeting_day = int(data["meeting_day"])
    if "is_active" in data:
        hc.is_active = bool(data["is_active"])
    db.commit()
    _log_admin_action(request, db, "update_home_church", "home_church", hc.id, f"Updated '{hc.name}'")
    db.commit()
    _push_home_church_to_central(hc, request, db, action="update")
    return _home_church_to_dict(hc, db)


@router.delete("/admin/home-churches/{hc_id}")
def admin_delete_home_church(hc_id: int, request: Request, db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    hc = db.query(HomeChurch).filter(HomeChurch.id == hc_id).first()
    if not hc:
        raise HTTPException(status_code=404, detail="Home church not found")
    name = hc.name
    # Push delete BEFORE removing the local row so we still have the external id
    _push_home_church_to_central(hc, request, db, action="delete")
    db.delete(hc)
    db.commit()
    _log_admin_action(request, db, "delete_home_church", "home_church", hc_id, f"Deleted home church '{name}'")
    db.commit()
    return {"success": True}


# ---- PREACHER POOL ----

@router.get("/admin/home-church/preachers")
def admin_list_preachers(request: Request, db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    member_ids = _preacher_pool_member_ids(db)
    if not member_ids:
        return []
    members = db.query(Member).filter(Member.id.in_(member_ids), Member.is_active == True).order_by(Member.full_name).all()
    return [{"id": m.id, "full_name": m.full_name, "phone": m.phone, "email": m.email} for m in members]


# ---- HOME CHURCH MEMBERSHIP (sourced from rfm-database) ----
#
# Membership lives in the central rfm-database — `church_members.home_church_id`
# points at the home_churches table. The portal here is just a UI: every read
# and write below proxies to the central API. Local data isn't used.

def _hc_member_dict(api_member: dict) -> dict:
    """Trim a central member record to the fields we render on the page."""
    if not isinstance(api_member, dict):
        return {}
    full = " ".join(filter(None, [api_member.get("first_name"), api_member.get("last_name")])).strip()
    if not full:
        full = api_member.get("full_name") or ""
    return {
        "id": api_member.get("id"),
        "full_name": full,
        "phone": (api_member.get("phone") or "").replace(" ", ""),
        "email": api_member.get("email") or "",
        "home_church_id": api_member.get("home_church_id"),
        "gender": api_member.get("gender"),
    }


def _ensure_central_synced(hc: HomeChurch, request: Request, db: Session):
    """Make sure this home church has a central UUID; auto-push if not."""
    if hc.external_home_church_id:
        return
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        raise HTTPException(
            status_code=400,
            detail="Central database integration is disabled or unconfigured."
        )
    result = _push_home_church_to_central(hc, request, db, action="create")
    if not result["ok"] or not hc.external_home_church_id:
        raise HTTPException(
            status_code=502,
            detail=f"Could not sync this home church to the central database: {result.get('error') or 'unknown error'}",
        )


@router.get("/admin/home-church/{hc_id}/members")
def admin_hc_members(hc_id: int, request: Request, db: Session = Depends(get_db)):
    """List members assigned to this home church (sourced from rfm-database)."""
    hc = db.query(HomeChurch).filter(HomeChurch.id == hc_id).first()
    if not hc:
        raise HTTPException(status_code=404, detail="Home church not found")
    _require_hc_access(request, db, hc)
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="Central database integration is disabled or unconfigured")
    _ensure_central_synced(hc, request, db)

    assembly_id = _resolve_default_assembly_id(db)
    members: List[dict] = []
    page = 1
    while True:
        r = _rfm.list_home_church_members(
            hc.external_home_church_id,
            assembly_id=str(assembly_id) if assembly_id else None,
            page=page, size=100, db=db,
        )
        if not r.ok:
            raise HTTPException(status_code=502, detail=f"Central API error: {r.error}")
        items = r.data if isinstance(r.data, list) else (r.data or {}).get("data") or []
        members.extend([_hc_member_dict(m) for m in items])
        if not isinstance(r.data, dict):
            break
        meta = r.data.get("meta") or {}
        if len(items) < 100 or page >= int(meta.get("pages") or 1):
            break
        page += 1
    members.sort(key=lambda m: (m.get("full_name") or "").lower())
    return {
        "home_church": {"id": hc.id, "name": hc.name, "external_id": hc.external_home_church_id},
        "members": members,
        "total": len(members),
    }


@router.post("/admin/home-church/{hc_id}/members")
def admin_hc_add_member(hc_id: int, payload: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    """Add a member to this home church.

    Accepts either a local `member_id` (the portal's Member.id, which we
    resolve to the central UUID) or `external_member_id` directly."""
    hc = db.query(HomeChurch).filter(HomeChurch.id == hc_id).first()
    if not hc:
        raise HTTPException(status_code=404, detail="Home church not found")
    actor = _require_hc_access(request, db, hc)
    _ensure_central_synced(hc, request, db)

    external_id = (payload.get("external_member_id") or "").strip() or None
    if not external_id and payload.get("member_id"):
        local = db.query(Member).filter(Member.id == int(payload["member_id"])).first()
        if not local:
            raise HTTPException(status_code=404, detail="Member not found locally")
        if not local.external_member_id:
            raise HTTPException(
                status_code=400,
                detail=f"{local.full_name} isn't synced to the central database yet. Use Member Sync first.",
            )
        external_id = local.external_member_id
    if not external_id:
        raise HTTPException(status_code=400, detail="Provide member_id or external_member_id")

    r = _rfm.update_member(external_id, {"home_church_id": hc.external_home_church_id}, db=db)
    if not r.ok:
        raise HTTPException(status_code=502, detail=f"Central API error: {r.error}")
    from routers.pages import is_authenticated as _is_admin_auth
    if _is_admin_auth(request):
        _log_admin_action(request, db, "hc_add_member", "home_church", hc.id,
                          f"Added external member {external_id} to '{hc.name}'")
    else:
        _log_member_action(request, db, actor, "hc_add_member", "home_church", hc.id,
                           f"Added external member {external_id} to '{hc.name}'")
    db.commit()
    return {"success": True, "member": _hc_member_dict(r.data if isinstance(r.data, dict) else {})}


@router.delete("/admin/home-church/{hc_id}/members/{external_member_id}")
def admin_hc_remove_member(hc_id: int, external_member_id: str, request: Request, db: Session = Depends(get_db)):
    """Remove a member from this home church (sets home_church_id = null centrally)."""
    hc = db.query(HomeChurch).filter(HomeChurch.id == hc_id).first()
    if not hc:
        raise HTTPException(status_code=404, detail="Home church not found")
    actor = _require_hc_access(request, db, hc)

    r = _rfm.update_member(external_member_id, {"home_church_id": None}, db=db)
    if not r.ok:
        raise HTTPException(status_code=502, detail=f"Central API error: {r.error}")
    from routers.pages import is_authenticated as _is_admin_auth
    if _is_admin_auth(request):
        _log_admin_action(request, db, "hc_remove_member", "home_church", hc.id,
                          f"Removed external member {external_member_id} from '{hc.name}'")
    else:
        _log_member_action(request, db, actor, "hc_remove_member", "home_church", hc.id,
                           f"Removed external member {external_member_id} from '{hc.name}'")
    db.commit()
    return {"success": True}


@router.get("/admin/home-church/membership-stats")
def admin_hc_membership_stats(request: Request, db: Session = Depends(get_db)):
    """Counts: total members, with HC, without HC, plus per-home-church count.
    Committee/admin only — leaders don't need the global view."""
    _require_committee_or_admin(request, db)
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="Central database integration is disabled or unconfigured")

    assembly_id = _resolve_default_assembly_id(db)
    if not assembly_id:
        raise HTTPException(status_code=400, detail="Could not resolve assembly_id")

    # Total members in the assembly
    total_r = _rfm.search_members(assembly_id=str(assembly_id), page=1, size=1, db=db)
    if not total_r.ok:
        raise HTTPException(status_code=502, detail=f"Central API error: {total_r.error}")
    meta = (total_r.data or {}).get("meta") or {} if isinstance(total_r.data, dict) else {}
    total_members = int(meta.get("total") or 0)

    # Per-home-church count + sum
    hcs = db.query(HomeChurch).filter(HomeChurch.external_home_church_id.isnot(None)).all()
    per_hc = []
    members_with_hc = 0
    for hc in hcs:
        r = _rfm.list_home_church_members(
            hc.external_home_church_id,
            assembly_id=str(assembly_id),
            page=1, size=1, db=db,
        )
        count = 0
        if r.ok and isinstance(r.data, dict):
            count = int(((r.data.get("meta") or {}).get("total")) or 0)
        members_with_hc += count
        per_hc.append({"id": hc.id, "name": hc.name, "count": count})
    per_hc.sort(key=lambda x: x["count"], reverse=True)

    return {
        "total_members": total_members,
        "with_home_church": members_with_hc,
        "without_home_church": max(0, total_members - members_with_hc),
        "per_home_church": per_hc,
        "unsynced_home_church_count": db.query(HomeChurch).filter(HomeChurch.external_home_church_id.is_(None)).count(),
    }


# ---- ROSTER ----

@router.get("/admin/home-church/roster")
def admin_get_roster(
    request: Request,
    start_date: Optional[str] = Query(None),
    weeks: int = Query(8, ge=1, le=26),
    db: Session = Depends(get_db),
):
    _require_committee_or_admin(request, db)
    if start_date:
        try:
            start = date.fromisoformat(start_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date")
    else:
        start = _next_weekday()

    dates = [start + timedelta(days=7 * i) for i in range(weeks)]
    end_date = dates[-1]

    churches = db.query(HomeChurch).filter(HomeChurch.is_active == True).order_by(HomeChurch.name).all()
    entries = db.query(HomeChurchRoster).filter(
        HomeChurchRoster.roster_date >= start,
        HomeChurchRoster.roster_date <= end_date,
    ).all()

    entry_map = {}
    for e in entries:
        entry_map[(e.home_church_id, e.roster_date)] = e

    program_types = db.query(HomeChurchProgramType).filter(HomeChurchProgramType.is_active == True).order_by(HomeChurchProgramType.sort_order).all()

    matrix = []
    for c in churches:
        row = {"home_church": _home_church_to_dict(c, db), "cells": []}
        for d in dates:
            entry = entry_map.get((c.id, d))
            row["cells"].append({
                "date": d.isoformat(),
                "entry": _roster_entry_to_dict(entry, db) if entry else None,
            })
        matrix.append(row)

    return {
        "start_date": start.isoformat(),
        "end_date": end_date.isoformat(),
        "dates": [d.isoformat() for d in dates],
        "weeks": weeks,
        "home_churches": [_home_church_to_dict(c, db) for c in churches],
        "program_types": [_program_type_to_dict(t) for t in program_types],
        "matrix": matrix,
    }


@router.put("/admin/home-church/roster/cell")
def admin_upsert_roster_cell(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    try:
        hc_id = int(data["home_church_id"])
        roster_date = date.fromisoformat(data["roster_date"])
    except (KeyError, ValueError, TypeError):
        raise HTTPException(status_code=400, detail="home_church_id and roster_date required")

    program_type_id = data.get("program_type_id")
    preacher_id = data.get("preacher_member_id")
    notes = data.get("notes")

    program_type = None
    if program_type_id:
        program_type = db.query(HomeChurchProgramType).filter(HomeChurchProgramType.id == program_type_id).first()
        if not program_type:
            raise HTTPException(status_code=404, detail="Program type not found")
        if not program_type.requires_preacher:
            preacher_id = None

    if preacher_id:
        preacher = db.query(Member).filter(Member.id == preacher_id).first()
        if not preacher:
            raise HTTPException(status_code=404, detail="Preacher not found")

    entry = db.query(HomeChurchRoster).filter(
        HomeChurchRoster.home_church_id == hc_id,
        HomeChurchRoster.roster_date == roster_date,
    ).first()

    is_new = entry is None
    if is_new:
        entry = HomeChurchRoster(
            home_church_id=hc_id,
            roster_date=roster_date,
            program_type_id=program_type_id,
            preacher_member_id=preacher_id,
            notes=notes,
            status="draft",
        )
        db.add(entry)
    else:
        entry.program_type_id = program_type_id
        entry.preacher_member_id = preacher_id
        if "notes" in data:
            entry.notes = notes

    db.commit()
    db.refresh(entry)

    _log_admin_action(
        request, db,
        "upsert_home_church_roster",
        "home_church_roster",
        entry.id,
        f"{'Created' if is_new else 'Updated'} roster for home church {hc_id} on {roster_date.isoformat()}"
    )
    db.commit()

    return _roster_entry_to_dict(entry, db)


@router.delete("/admin/home-church/roster/cell")
def admin_clear_roster_cell(
    request: Request,
    home_church_id: int = Query(...),
    roster_date: str = Query(...),
    db: Session = Depends(get_db),
):
    _require_committee_or_admin(request, db)
    try:
        d = date.fromisoformat(roster_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date")
    entry = db.query(HomeChurchRoster).filter(
        HomeChurchRoster.home_church_id == home_church_id,
        HomeChurchRoster.roster_date == d,
    ).first()
    if not entry:
        return {"success": True, "removed": False}
    db.delete(entry)
    db.commit()
    _log_admin_action(request, db, "clear_home_church_roster", "home_church_roster", entry.id, f"Cleared roster for home church {home_church_id} on {d.isoformat()}")
    db.commit()
    return {"success": True, "removed": True}


@router.put("/admin/home-church/roster/week")
def admin_set_week_program(
    request: Request,
    data: dict = Body(...),
    db: Session = Depends(get_db),
):
    """Cascade a single program type to every active home church for a given
    date. The common case: all home churches share the same programme each
    week. Creates missing roster entries and updates existing ones. If the
    new type doesn't require a preacher, preacher assignments are cleared.
    Already-published rows are left alone unless force=true."""
    _require_committee_or_admin(request, db)

    try:
        roster_date = date.fromisoformat(data["roster_date"])
    except (KeyError, ValueError, TypeError):
        raise HTTPException(status_code=400, detail="roster_date required")

    program_type_id = data.get("program_type_id")
    force = bool(data.get("force", False))

    program_type = None
    if program_type_id:
        program_type = db.query(HomeChurchProgramType).filter(
            HomeChurchProgramType.id == program_type_id
        ).first()
        if not program_type:
            raise HTTPException(status_code=404, detail="Program type not found")

    churches = db.query(HomeChurch).filter(HomeChurch.is_active == True).all()
    existing = {
        e.home_church_id: e
        for e in db.query(HomeChurchRoster).filter(
            HomeChurchRoster.roster_date == roster_date
        ).all()
    }

    updated_count = 0
    created_count = 0
    skipped_published = 0

    for c in churches:
        entry = existing.get(c.id)
        if entry is None:
            entry = HomeChurchRoster(
                home_church_id=c.id,
                roster_date=roster_date,
                program_type_id=program_type_id,
                status="draft",
            )
            # Non-preaching type: no preacher
            if program_type and not program_type.requires_preacher:
                entry.preacher_member_id = None
            db.add(entry)
            created_count += 1
        else:
            if entry.status == "published" and not force:
                skipped_published += 1
                continue
            entry.program_type_id = program_type_id
            if program_type and not program_type.requires_preacher:
                entry.preacher_member_id = None
            updated_count += 1

    db.commit()

    _log_admin_action(
        request, db,
        "set_week_program",
        "home_church_roster",
        None,
        f"Set week {roster_date.isoformat()} to "
        f"{program_type.name if program_type else 'none'} "
        f"({created_count} created, {updated_count} updated, {skipped_published} published skipped)"
    )
    db.commit()

    return {
        "success": True,
        "created": created_count,
        "updated": updated_count,
        "skipped_published": skipped_published,
    }


@router.post("/admin/home-church/roster/auto-fill")
def admin_auto_fill_roster(request: Request, data: dict = Body(default={}), db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    start_date_str = data.get("start_date")
    weeks = int(data.get("weeks", 4))
    if start_date_str:
        try:
            start = date.fromisoformat(start_date_str)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date")
    else:
        start = _next_weekday()
    end_date = start + timedelta(days=7 * weeks)

    entries = db.query(HomeChurchRoster).join(HomeChurchProgramType).filter(
        HomeChurchRoster.roster_date >= start,
        HomeChurchRoster.roster_date < end_date,
        HomeChurchProgramType.requires_preacher == True,
        HomeChurchRoster.preacher_member_id.is_(None),
    ).order_by(HomeChurchRoster.roster_date, HomeChurchRoster.home_church_id).all()

    if not entries:
        return {"success": True, "filled": 0, "message": "No empty preacher slots. Set program types first."}

    member_ids = _preacher_pool_member_ids(db)
    if not member_ids:
        raise HTTPException(status_code=400, detail="No preachers available in the pool")
    preachers = db.query(Member).filter(Member.id.in_(member_ids), Member.is_active == True).all()
    if not preachers:
        raise HTTPException(status_code=400, detail="No active preachers available")

    recent_assignments = {}
    existing = db.query(HomeChurchRoster).filter(
        HomeChurchRoster.roster_date >= start - timedelta(days=56),
        HomeChurchRoster.preacher_member_id.isnot(None),
    ).all()
    for e in existing:
        recent_assignments.setdefault(e.preacher_member_id, []).append(e.roster_date)

    hc_leader = {hc.id: hc.leader_member_id for hc in db.query(HomeChurch).all()}

    filled = 0
    for entry in entries:
        leader_id = hc_leader.get(entry.home_church_id)

        def score(p):
            count = len(recent_assignments.get(p.id, []))
            self_preach = 1 if p.id == leader_id else 0
            return (self_preach, count)

        candidates = sorted(preachers, key=score)
        chosen = candidates[0]
        entry.preacher_member_id = chosen.id
        recent_assignments.setdefault(chosen.id, []).append(entry.roster_date)
        filled += 1

    db.commit()
    _log_admin_action(request, db, "auto_fill_home_church_roster", "home_church_roster", None, f"Auto-filled {filled} preacher slots from {start.isoformat()} for {weeks} weeks")
    db.commit()
    return {"success": True, "filled": filled}


@router.post("/admin/home-church/roster/publish")
def admin_publish_roster(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    roster_date_str = data.get("roster_date")
    if not roster_date_str:
        raise HTTPException(status_code=400, detail="roster_date required")
    try:
        d = date.fromisoformat(roster_date_str)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date")

    entries = db.query(HomeChurchRoster).options(
        joinedload(HomeChurchRoster.home_church).joinedload(HomeChurch.leader),
        joinedload(HomeChurchRoster.program_type),
        joinedload(HomeChurchRoster.preacher),
    ).filter(HomeChurchRoster.roster_date == d).all()

    if not entries:
        raise HTTPException(status_code=404, detail="No roster entries found for that date")

    now = datetime.utcnow()
    published_count = 0
    for e in entries:
        if e.status != "published":
            published_count += 1
        e.status = "published"
        e.published_at = now
    db.commit()

    try:
        from notifications.dispatcher import dispatch_event
        from notifications.events import EventType

        for e in entries:
            hc = e.home_church
            if not hc or not hc.leader or not hc.leader.email:
                continue
            dispatch_event(db, EventType.HOME_CHURCH_ROSTER_PUBLISHED, {
                "leader_id": hc.leader.id,
                "leader_name": hc.leader.full_name,
                "leader_email": hc.leader.email,
                "home_church_name": hc.name,
                "roster_date": d.isoformat(),
                "meeting_time": hc.meeting_time,
                "program_type_name": e.program_type.name if e.program_type else "Not set",
                "program_type_icon": e.program_type.icon if e.program_type else "\U0001F4CC",
                "requires_preacher": e.program_type.requires_preacher if e.program_type else False,
                "preacher_name": e.preacher.full_name if e.preacher else None,
                "preacher_phone": e.preacher.phone if e.preacher else None,
                "recipients": [{"id": hc.leader.id, "name": hc.leader.full_name, "email": hc.leader.email, "phone": hc.leader.phone}],
            })

        for e in entries:
            if not e.preacher or not e.preacher.email:
                continue
            hc = e.home_church
            dispatch_event(db, EventType.HOME_CHURCH_PREACHER_ASSIGNED, {
                "preacher_id": e.preacher.id,
                "preacher_name": e.preacher.full_name,
                "preacher_email": e.preacher.email,
                "home_church_name": hc.name if hc else "",
                "home_church_address": hc.address if hc else "",
                "leader_name": hc.leader.full_name if (hc and hc.leader) else "",
                "leader_phone": hc.leader.phone if (hc and hc.leader) else "",
                "roster_date": d.isoformat(),
                "meeting_time": hc.meeting_time if hc else "19:00",
                "recipients": [{"id": e.preacher.id, "name": e.preacher.full_name, "email": e.preacher.email, "phone": e.preacher.phone}],
            })
    except Exception as exc:
        print(f"Failed to dispatch home church roster notifications: {exc}")

    _log_admin_action(request, db, "publish_home_church_roster", "home_church_roster", None, f"Published roster for {d.isoformat()} ({len(entries)} entries)")
    db.commit()

    return {"success": True, "date": d.isoformat(), "entries_published": len(entries), "newly_published": published_count}


# ---- PUBLIC (portal) ----

@router.get("/home-church/my-upcoming")
def my_upcoming_home_church(
    phone: str = Query(...),
    weeks: int = Query(8, ge=1, le=26),
    db: Session = Depends(get_db),
):
    member = _find_member_by_phone(db, phone)
    if not member:
        return {"as_leader": [], "as_preacher": []}

    today = date.today()
    horizon = today + timedelta(days=7 * weeks)

    led_churches = db.query(HomeChurch).filter(HomeChurch.leader_member_id == member.id).all()
    led_ids = [c.id for c in led_churches]
    as_leader = []
    if led_ids:
        entries = db.query(HomeChurchRoster).options(
            joinedload(HomeChurchRoster.home_church),
            joinedload(HomeChurchRoster.program_type),
            joinedload(HomeChurchRoster.preacher),
        ).filter(
            HomeChurchRoster.home_church_id.in_(led_ids),
            HomeChurchRoster.roster_date >= today,
            HomeChurchRoster.roster_date <= horizon,
            HomeChurchRoster.status == "published",
        ).order_by(HomeChurchRoster.roster_date).all()
        as_leader = [_roster_entry_to_dict(e, db) for e in entries]
        # attach home church info
        for i, e in enumerate(entries):
            as_leader[i]["home_church"] = _home_church_to_dict(e.home_church, db) if e.home_church else None

    entries = db.query(HomeChurchRoster).options(
        joinedload(HomeChurchRoster.home_church).joinedload(HomeChurch.leader),
        joinedload(HomeChurchRoster.program_type),
    ).filter(
        HomeChurchRoster.preacher_member_id == member.id,
        HomeChurchRoster.roster_date >= today,
        HomeChurchRoster.roster_date <= horizon,
        HomeChurchRoster.status == "published",
    ).order_by(HomeChurchRoster.roster_date).all()
    as_preacher = []
    for e in entries:
        d = _roster_entry_to_dict(e, db)
        hc = e.home_church
        d["home_church"] = _home_church_to_dict(hc, db) if hc else None
        as_preacher.append(d)

    return {"as_leader": as_leader, "as_preacher": as_preacher}



# ============================================================================
# HOME CHURCH ATTENDANCE (committee captures; leaders submit via WhatsApp)
# ============================================================================

def _attendance_to_dict(a: HomeChurchAttendance, db: Session) -> dict:
    submitted_by = None
    if a.submitted_by_member_id:
        m = db.query(Member).filter(Member.id == a.submitted_by_member_id).first()
        if m:
            submitted_by = {"id": m.id, "full_name": m.full_name}
    try:
        offering = float(a.offering_amount) if a.offering_amount else 0.0
    except (TypeError, ValueError):
        offering = 0.0
    return {
        "id": a.id,
        "home_church_id": a.home_church_id,
        "roster_date": a.roster_date.isoformat() if a.roster_date else None,
        "did_not_meet": bool(a.did_not_meet),
        "attendance_count": a.attendance_count or 0,
        "adults_count": a.adults_count,
        "children_count": a.children_count,
        "new_visitors_count": a.new_visitors_count or 0,
        "offering_amount": offering,
        "notes": a.notes,
        "submitted_by": submitted_by,
        "submitted_at": a.submitted_at.isoformat() if a.submitted_at else None,
        "updated_at": a.updated_at.isoformat() if a.updated_at else None,
        "reminder_sent_at": a.reminder_sent_at.isoformat() if a.reminder_sent_at else None,
    }


@router.get("/admin/home-church/attendance")
def admin_list_attendance(
    request: Request,
    start_date: Optional[str] = Query(None),
    weeks: int = Query(8, ge=1, le=52),
    db: Session = Depends(get_db),
):
    """Return the attendance matrix for the given range: each active home church
    x each Monday, with its report (if captured) or a null placeholder."""
    _require_committee_or_admin(request, db)

    if start_date:
        try:
            start = date.fromisoformat(start_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date")
    else:
        # Default: start one week in the PAST so the committee can fill
        # in last Monday's numbers, working forward.
        start = _next_weekday() - timedelta(days=7)

    dates = [start + timedelta(days=7 * i) for i in range(weeks)]
    end_date = dates[-1]

    churches = db.query(HomeChurch).filter(HomeChurch.is_active == True).order_by(HomeChurch.name).all()
    reports = db.query(HomeChurchAttendance).filter(
        HomeChurchAttendance.roster_date >= start,
        HomeChurchAttendance.roster_date <= end_date,
    ).all()

    index = {(r.home_church_id, r.roster_date): r for r in reports}

    matrix = []
    for c in churches:
        row = {"home_church": _home_church_to_dict(c, db), "cells": []}
        for d in dates:
            r = index.get((c.id, d))
            row["cells"].append({
                "date": d.isoformat(),
                "report": _attendance_to_dict(r, db) if r else None,
            })
        matrix.append(row)

    return {
        "start_date": start.isoformat(),
        "end_date": end_date.isoformat(),
        "dates": [d.isoformat() for d in dates],
        "weeks": weeks,
        "home_churches": [_home_church_to_dict(c, db) for c in churches],
        "matrix": matrix,
    }


@router.put("/admin/home-church/attendance")
def admin_upsert_attendance(
    request: Request,
    data: dict = Body(...),
    db: Session = Depends(get_db),
):
    """Committee captures (or updates) attendance for a given home church / date."""
    actor = _require_committee_or_admin(request, db)

    try:
        hc_id = int(data["home_church_id"])
        d = date.fromisoformat(data["roster_date"])
    except (KeyError, ValueError, TypeError):
        raise HTTPException(status_code=400, detail="home_church_id and roster_date required")

    hc = db.query(HomeChurch).filter(HomeChurch.id == hc_id).first()
    if not hc:
        raise HTTPException(status_code=404, detail="Home church not found")

    did_not_meet = bool(data.get("did_not_meet", False))
    attendance_count = 0
    adults = None
    children = None
    new_visitors = 0
    offering_str = "0"
    notes = data.get("notes")

    if not did_not_meet:
        try:
            attendance_count = int(data.get("attendance_count") or 0)
            if data.get("adults_count") not in (None, ""):
                adults = int(data["adults_count"])
            if data.get("children_count") not in (None, ""):
                children = int(data["children_count"])
            new_visitors = int(data.get("new_visitors_count") or 0)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Counts must be whole numbers")
        # Offering as plain text (we store and display; avoids float rounding)
        offering_raw = data.get("offering_amount", "0")
        if offering_raw is None:
            offering_raw = "0"
        try:
            # Validate parseable
            float(str(offering_raw).replace(",", "").strip() or "0")
        except ValueError:
            raise HTTPException(status_code=400, detail="Offering must be a number")
        offering_str = str(offering_raw)

    entry = db.query(HomeChurchAttendance).filter(
        HomeChurchAttendance.home_church_id == hc_id,
        HomeChurchAttendance.roster_date == d,
    ).first()

    is_new = entry is None
    if is_new:
        entry = HomeChurchAttendance(
            home_church_id=hc_id,
            roster_date=d,
            did_not_meet=did_not_meet,
            attendance_count=attendance_count,
            adults_count=adults,
            children_count=children,
            new_visitors_count=new_visitors,
            offering_amount=offering_str,
            notes=notes,
            submitted_by_member_id=actor.id if actor else None,
        )
        db.add(entry)
    else:
        entry.did_not_meet = did_not_meet
        entry.attendance_count = attendance_count
        entry.adults_count = adults
        entry.children_count = children
        entry.new_visitors_count = new_visitors
        entry.offering_amount = offering_str
        entry.notes = notes
        if actor:
            entry.submitted_by_member_id = actor.id

    db.commit()
    db.refresh(entry)

    _log_admin_action(
        request, db,
        "upsert_home_church_attendance",
        "home_church_attendance",
        entry.id,
        f"{'Captured' if is_new else 'Updated'} {hc.name} attendance for {d.isoformat()} (attendance={attendance_count}, offering={offering_str})"
    )
    db.commit()

    return _attendance_to_dict(entry, db)


@router.delete("/admin/home-church/attendance/{attendance_id}")
def admin_delete_attendance(attendance_id: int, request: Request, db: Session = Depends(get_db)):
    _require_committee_or_admin(request, db)
    entry = db.query(HomeChurchAttendance).filter(HomeChurchAttendance.id == attendance_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Not found")
    hc = entry.home_church
    d = entry.roster_date
    db.delete(entry)
    db.commit()
    _log_admin_action(request, db, "delete_home_church_attendance", "home_church_attendance", attendance_id,
                     f"Deleted attendance for {hc.name if hc else '?'} on {d.isoformat() if d else '?'}")
    db.commit()
    return {"success": True}


@router.get("/admin/home-church/attendance/stats")
def admin_attendance_stats(
    request: Request,
    end_date: Optional[str] = Query(None, description="ISO date; defaults to today. Analyses the 8 weeks ending on the Monday <= this date."),
    db: Session = Depends(get_db),
):
    """Weekly totals, monthly averages, and improving / attention flags.

    Compares the most recent 4 Mondays against the prior 4 for each home church.
    """
    _require_committee_or_admin(request, db)

    if end_date:
        try:
            end = date.fromisoformat(end_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date")
    else:
        end = date.today()

    # Anchor on the most recent Monday <= end
    # Python: weekday() returns 0 for Monday
    anchor_monday = end - timedelta(days=end.weekday())

    # Get the 8 recent Mondays (anchor going backwards)
    recent = [anchor_monday - timedelta(days=7 * i) for i in range(8)]  # [0]=latest
    recent.reverse()  # oldest first
    first_day = recent[0]
    last_day = recent[-1]

    churches = db.query(HomeChurch).filter(HomeChurch.is_active == True).order_by(HomeChurch.name).all()
    reports = db.query(HomeChurchAttendance).filter(
        HomeChurchAttendance.roster_date >= first_day,
        HomeChurchAttendance.roster_date <= last_day,
    ).all()
    by_hc = {}
    for r in reports:
        by_hc.setdefault(r.home_church_id, {})[r.roster_date] = r

    # Weekly totals across all churches
    weekly_totals = []
    for d in recent:
        week = {"date": d.isoformat(), "attendance": 0, "offering": 0.0, "reporting": 0, "total": len(churches)}
        for c in churches:
            r = by_hc.get(c.id, {}).get(d)
            if r and not r.did_not_meet:
                week["attendance"] += r.attendance_count or 0
                try:
                    week["offering"] += float(r.offering_amount or 0)
                except (TypeError, ValueError):
                    pass
            if r:
                week["reporting"] += 1
        weekly_totals.append(week)

    # Per-church stats: 4-week recent vs 4-week prior, monthly average, flags
    per_church = []
    def _safe_avg(vals):
        return (sum(vals) / len(vals)) if vals else 0.0

    last4 = recent[-4:]   # most recent 4
    prior4 = recent[:4]   # the 4 before

    for c in churches:
        hc_reports = by_hc.get(c.id, {})
        # Attendance figures for last 4 (ignoring did_not_meet rows and missing rows)
        last4_att = [hc_reports[d].attendance_count for d in last4 if d in hc_reports and not hc_reports[d].did_not_meet]
        prior4_att = [hc_reports[d].attendance_count for d in prior4 if d in hc_reports and not hc_reports[d].did_not_meet]
        last4_off = []
        for d in last4:
            if d in hc_reports and not hc_reports[d].did_not_meet:
                try:
                    last4_off.append(float(hc_reports[d].offering_amount or 0))
                except (TypeError, ValueError):
                    pass

        missing_last4 = sum(1 for d in last4 if d not in hc_reports)
        last4_avg = _safe_avg(last4_att)
        prior4_avg = _safe_avg(prior4_att)

        # Change % (only meaningful if we have a prior baseline)
        change_pct = None
        if prior4_avg > 0:
            change_pct = ((last4_avg - prior4_avg) / prior4_avg) * 100

        # Flag
        flag = "steady"
        reason = None
        if missing_last4 >= 2:
            flag = "attention"
            reason = f"{missing_last4} reports missing in last 4 weeks"
        elif change_pct is not None and change_pct >= 15:
            flag = "improving"
            reason = f"+{change_pct:.0f}% vs previous 4 weeks"
        elif change_pct is not None and change_pct <= -15:
            flag = "attention"
            reason = f"{change_pct:.0f}% vs previous 4 weeks"
        elif change_pct is not None:
            reason = f"{change_pct:+.0f}% vs previous 4 weeks"

        per_church.append({
            "home_church": _home_church_to_dict(c, db),
            "last4_avg_attendance": round(last4_avg, 1),
            "prior4_avg_attendance": round(prior4_avg, 1),
            "last4_total_offering": round(sum(last4_off), 2),
            "last4_avg_offering": round(_safe_avg(last4_off), 2),
            "missing_last4": missing_last4,
            "change_pct": round(change_pct, 1) if change_pct is not None else None,
            "flag": flag,
            "reason": reason,
            # Sparkline-friendly: last 8 weeks of attendance (None when no report)
            "trend_attendance": [
                (hc_reports[d].attendance_count if d in hc_reports and not hc_reports[d].did_not_meet else None)
                for d in recent
            ],
        })

    # Grand totals for the most recent week (anchor) and last 4 weeks combined
    latest_week = weekly_totals[-1] if weekly_totals else None
    last4_total_attendance = sum(w["attendance"] for w in weekly_totals[-4:])
    last4_total_offering = sum(w["offering"] for w in weekly_totals[-4:])
    last4_avg_attendance = round(last4_total_attendance / 4, 1)
    last4_avg_offering = round(last4_total_offering / 4, 2)

    return {
        "anchor_monday": anchor_monday.isoformat(),
        "range": {"from": first_day.isoformat(), "to": last_day.isoformat()},
        "weekly_totals": weekly_totals,
        "latest_week": latest_week,
        "month_summary": {
            "total_attendance": last4_total_attendance,
            "total_offering": round(last4_total_offering, 2),
            "avg_attendance_per_week": last4_avg_attendance,
            "avg_offering_per_week": last4_avg_offering,
            "reporting_rate": round(
                100 * sum(w["reporting"] for w in weekly_totals[-4:]) / max(1, 4 * len(churches)),
                1,
            ),
        },
        "per_church": per_church,
    }


@router.get("/admin/home-church/{home_church_id}/history")
def admin_home_church_history(
    home_church_id: int,
    request: Request,
    weeks: int = Query(12, ge=4, le=52),
    db: Session = Depends(get_db),
):
    """Per-home-church history view: N weeks of attendance + offering with
    headline stats (current vs prior month, group average comparison)."""
    _require_committee_or_admin(request, db)

    hc = db.query(HomeChurch).filter(HomeChurch.id == home_church_id).first()
    if not hc:
        raise HTTPException(status_code=404, detail="Home church not found")

    today = date.today()
    # Anchor on most recent Monday <= today
    anchor = today - timedelta(days=today.weekday())
    mondays = [anchor - timedelta(days=7 * i) for i in range(weeks - 1, -1, -1)]  # oldest first
    first = mondays[0]
    last = mondays[-1]

    # This home church's reports
    own_reports = {
        r.roster_date: r
        for r in db.query(HomeChurchAttendance).filter(
            HomeChurchAttendance.home_church_id == home_church_id,
            HomeChurchAttendance.roster_date >= first,
            HomeChurchAttendance.roster_date <= last,
        ).all()
    }

    # Group reports (for average comparison)
    all_reports = db.query(HomeChurchAttendance).filter(
        HomeChurchAttendance.roster_date >= first,
        HomeChurchAttendance.roster_date <= last,
    ).all()
    group_by_date = {}
    for r in all_reports:
        if not r.did_not_meet and (r.attendance_count or 0) > 0:
            group_by_date.setdefault(r.roster_date, []).append(r)

    series = []
    for d in mondays:
        r = own_reports.get(d)
        row = {
            "date": d.isoformat(),
            "attendance": None,
            "offering": None,
            "did_not_meet": False,
            "group_avg_attendance": None,
        }
        if r:
            row["did_not_meet"] = bool(r.did_not_meet)
            if not r.did_not_meet:
                row["attendance"] = r.attendance_count or 0
                try:
                    row["offering"] = float(r.offering_amount or 0)
                except (TypeError, ValueError):
                    row["offering"] = 0.0
        # Group average for that week (across all reporting churches)
        grp = group_by_date.get(d, [])
        if grp:
            row["group_avg_attendance"] = round(sum(x.attendance_count or 0 for x in grp) / len(grp), 1)
        series.append(row)

    # Recent 4 weeks vs prior 4 weeks (same logic as the main stats)
    last4 = [s for s in series[-4:] if s["attendance"] is not None]
    prior4 = [s for s in series[-8:-4] if s["attendance"] is not None] if len(series) >= 8 else []

    def _avg(xs, field):
        vals = [x[field] for x in xs if x[field] is not None]
        return (sum(vals) / len(vals)) if vals else 0.0

    last4_avg = _avg(last4, "attendance")
    prior4_avg = _avg(prior4, "attendance")
    change_pct = None
    if prior4_avg > 0:
        change_pct = ((last4_avg - prior4_avg) / prior4_avg) * 100

    last4_total_offering = sum((s["offering"] or 0) for s in series[-4:])
    prior4_total_offering = sum((s["offering"] or 0) for s in series[-8:-4]) if len(series) >= 8 else 0

    missing_last4 = sum(1 for s in series[-4:] if s["attendance"] is None and not s["did_not_meet"])

    # Recent notes (last 4 weeks of captures with notes)
    notes = []
    for d in reversed(mondays[-4:]):
        r = own_reports.get(d)
        if r and r.notes:
            notes.append({"date": d.isoformat(), "notes": r.notes})

    return {
        "home_church": _home_church_to_dict(hc, db),
        "weeks": weeks,
        "series": series,
        "summary": {
            "last4_avg_attendance": round(last4_avg, 1),
            "prior4_avg_attendance": round(prior4_avg, 1),
            "change_pct": round(change_pct, 1) if change_pct is not None else None,
            "last4_total_offering": round(last4_total_offering, 2),
            "prior4_total_offering": round(prior4_total_offering, 2),
            "missing_last4": missing_last4,
        },
        "recent_notes": notes,
    }



# ============================================================================
# RFM-DATABASE SYNC (admin tools — Phase 2)
# ----------------------------------------------------------------------------
# Lets the admin see which local members link to the central rfm-database
# (matched), which look ambiguous, and which are orphans (no record there).
# Everything is gated by RFM_API_INTEGRATION_ENABLED.
# ============================================================================

import rfm_api_client as _rfm


@router.get("/admin/rfm-sync/status")
def admin_rfm_sync_status(request: Request, db: Session = Depends(get_db)):
    """Summary counts + connectivity check. Used by the admin UI banner."""
    _require_committee_or_admin(request, db)

    enabled = _rfm.is_enabled(db)
    configured = _rfm.is_configured(db)

    counts = {"matched": 0, "ambiguous": 0, "unmatched": 0, "manual": 0, "pending": 0}
    for status, count in db.query(
        Member.external_match_status, func.count(Member.id)
    ).group_by(Member.external_match_status).all():
        key = status if status in counts else "pending"
        counts[key] = (counts.get(key) or 0) + count

    null_count = db.query(Member).filter(Member.external_match_status.is_(None)).count()
    counts["pending"] = null_count

    total = db.query(Member).count()

    health = None
    if enabled and configured:
        h = _rfm.health_check(db)
        health = {"ok": h.ok, "error": h.error, "status": h.status}

    return {
        "enabled": enabled,
        "configured": configured,
        "api_url": _rfm.get_api_url(db) or None,
        "ttl_seconds": _rfm.get_ttl_seconds(db),
        "counts": counts,
        "total_members": total,
        "health": health,
    }


@router.get("/admin/rfm-sync/members")
def admin_rfm_sync_list(
    request: Request,
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    _require_committee_or_admin(request, db)
    q = db.query(Member)
    if status == "pending":
        q = q.filter(Member.external_match_status.is_(None))
    elif status:
        q = q.filter(Member.external_match_status == status)
    members = q.order_by(Member.full_name).all()
    return [
        {
            "id": m.id,
            "full_name": m.full_name,
            "phone": m.phone,
            "email": m.email,
            "external_member_id": m.external_member_id,
            "external_match_status": m.external_match_status,
            "external_synced_at": m.external_synced_at.isoformat() if m.external_synced_at else None,
        }
        for m in members
    ]


@router.post("/admin/rfm-sync/match-one")
def admin_rfm_sync_match_one(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    """Run matching for one local member; returns candidates for the UI."""
    _require_committee_or_admin(request, db)
    member_id = data.get("member_id")
    if not member_id:
        raise HTTPException(status_code=400, detail="member_id required")
    member = db.query(Member).filter(Member.id == int(member_id)).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    result = _rfm.match_local_member(member, db=db)
    return {
        "member": {"id": member.id, "full_name": member.full_name, "phone": member.phone},
        "result": result,
    }


@router.post("/admin/rfm-sync/api-search")
def admin_rfm_sync_api_search(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    """Free-form search against the central database — used by the admin
    when the auto-matcher can't find someone they know is there. Takes
    `query` (any string: name, phone, email) and `member_id` (optional, the
    local member being reviewed — used to compute name match scores in the
    response so the UI can show how close each result is)."""
    _require_committee_or_admin(request, db)
    query = (data.get("query") or "").strip()
    if not query:
        raise HTTPException(status_code=400, detail="query required")
    if len(query) < 2:
        raise HTTPException(status_code=400, detail="query must be at least 2 characters")

    if not _rfm.is_enabled(db):
        raise HTTPException(status_code=400, detail="rfm-db integration is disabled")
    if not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="rfm-db API not configured")

    # If the query looks like a phone, normalise it to last-9-digits before
    # searching (so "+27 61 919 7741" still matches a stored "0619197741").
    phone_norm = _rfm.normalise_phone(query)
    search_term = phone_norm if (phone_norm and len(phone_norm) >= 9) else query

    api_result = _rfm.search_members(search=search_term, page=1, size=50, db=db)
    if api_result.disabled:
        raise HTTPException(status_code=400, detail="rfm-db integration disabled")
    if not api_result.ok:
        raise HTTPException(status_code=502, detail=api_result.error or "Central database error")

    raw = api_result.data or []
    if isinstance(raw, dict):
        raw = raw.get("data") or []

    # Score each result against the local member's name (for relative ranking)
    local_name = ""
    member_id = data.get("member_id")
    if member_id:
        m = db.query(Member).filter(Member.id == int(member_id)).first()
        if m:
            local_name = m.full_name or ""

    results = sorted(
        (_rfm._score_and_strip(item, _rfm.name_match_score(local_name, item)) for item in raw),
        key=lambda x: x["_score"], reverse=True,
    )
    return {"query": query, "count": len(results), "results": results}


@router.post("/admin/rfm-sync/match-all")
def admin_rfm_sync_match_all(request: Request, data: dict = Body(default={}), db: Session = Depends(get_db)):
    """Walk every member that hasn't been matched yet (or all, with force=true)
    and persist the matching result. Auto-matched rows store external_id;
    ambiguous and unmatched rows are flagged for review."""
    _require_committee_or_admin(request, db)
    if not _rfm.is_enabled(db):
        raise HTTPException(status_code=400, detail="rfm-db integration is disabled")
    if not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="rfm-db API not configured (URL/key missing)")

    force = bool(data.get("force", False))
    q = db.query(Member)
    if not force:
        q = q.filter(
            (Member.external_match_status.is_(None))
            | (Member.external_match_status == "ambiguous")
            | (Member.external_match_status == "unmatched")
        )
    members = q.all()

    summary = {"matched": 0, "ambiguous": 0, "unmatched": 0, "errors": 0, "skipped": 0, "enriched": 0}
    errors_sample = []
    enrichment_sample = []
    from datetime import datetime as _dt

    for m in members:
        result = _rfm.match_local_member(m, db=db)
        status = result.get("status")
        if status == "matched":
            m.external_member_id = result.get("external_id")
            m.external_assembly_id = result.get("assembly_id")
            m.external_match_status = "matched"
            m.external_synced_at = _dt.utcnow()
            summary["matched"] += 1
            # Enrichment: push fields the API was missing where we have valid
            # values (validated SA mobile / email). Never pushes invalid data.
            hint = result.get("enrich_hint") or {}
            if hint:
                push = _rfm.push_enrichment(result.get("external_id"), hint, db=db)
                if push.ok:
                    summary["enriched"] += 1
                    if len(enrichment_sample) < 5:
                        enrichment_sample.append({
                            "member_id": m.id,
                            "name": m.full_name,
                            "fields": list(hint.keys()),
                        })
        elif status == "ambiguous":
            m.external_match_status = "ambiguous"
            summary["ambiguous"] += 1
        elif status == "unmatched":
            m.external_match_status = "unmatched"
            summary["unmatched"] += 1
        elif status == "disabled":
            summary["skipped"] += 1
        else:
            summary["errors"] += 1
            if len(errors_sample) < 5:
                errors_sample.append({"member_id": m.id, "error": result.get("error")})

    db.commit()
    _log_admin_action(
        request, db,
        "rfm_sync_match_all",
        "members",
        None,
        f"Matched {summary['matched']}, ambiguous {summary['ambiguous']}, "
        f"unmatched {summary['unmatched']}, errors {summary['errors']}, "
        f"enriched {summary['enriched']}"
        + (" (forced)" if force else ""),
    )
    db.commit()
    return {
        "summary": summary,
        "errors_sample": errors_sample,
        "enrichment_sample": enrichment_sample,
        "scanned": len(members),
    }


@router.post("/admin/rfm-sync/confirm")
def admin_rfm_sync_confirm(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    """Manually link a local member to a specific external_member_id and
    optionally push explicit field overrides (phone / first_name / last_name /
    email) up to the API. Useful when:
      * The API record has the wrong phone and the portal has the right one
      * The API has a typo in first/last name we can correct from local
      * The API record is missing phone or email entirely
    Body:
      {
        member_id: int (required),
        external_member_id: uuid (required),
        assembly_id: uuid (optional),
        overrides: {
          phone: str | null,
          first_name: str | null,
          last_name: str | null,
          email: str | null,
        }   ← when omitted, falls back to fill-only-if-empty enrichment
      }"""
    _require_committee_or_admin(request, db)
    try:
        member_id = int(data["member_id"])
        external_id = str(data["external_member_id"])
    except (KeyError, ValueError, TypeError):
        raise HTTPException(status_code=400, detail="member_id and external_member_id required")
    assembly_id = data.get("assembly_id")
    overrides = data.get("overrides") or {}
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    from datetime import datetime as _dt
    member.external_member_id = external_id
    if assembly_id:
        member.external_assembly_id = str(assembly_id)
    member.external_match_status = "manual"
    member.external_synced_at = _dt.utcnow()
    db.commit()

    # Build the API patch payload.
    push_fields = {}
    if overrides:
        # Explicit overrides path — admin chose specific fields.
        # Validate phone (only push valid SA mobiles), light-validate email.
        if overrides.get("phone"):
            canonical = _rfm.to_sa_canonical_mobile(overrides["phone"])
            if canonical:
                push_fields["phone"] = canonical
        if overrides.get("first_name"):
            v = str(overrides["first_name"]).strip()
            if v:
                push_fields["first_name"] = v
        if overrides.get("last_name"):
            v = str(overrides["last_name"]).strip()
            if v:
                push_fields["last_name"] = v
        if overrides.get("email"):
            v = str(overrides["email"]).strip()
            if "@" in v and "." in v.split("@")[-1]:
                push_fields["email"] = v
    else:
        # Fallback: re-fetch the API record and only fill missing fields.
        if _rfm.is_enabled(db) and _rfm.is_configured(db):
            api_get = _rfm.get_member(external_id, db=db)
            if api_get.ok and isinstance(api_get.data, dict):
                push_fields = _rfm._enrich_diff(
                    api_get.data,
                    phone=member.phone or "",
                    email=member.email or "",
                )

    pushed_fields = []
    if push_fields:
        push = _rfm.push_enrichment(external_id, push_fields, db=db)
        if push.ok:
            pushed_fields = list(push_fields.keys())

    _log_admin_action(
        request, db, "rfm_sync_manual_match", "member", member.id,
        f"Manually matched {member.full_name} -> external {external_id}"
        + (f" (pushed to API: {', '.join(pushed_fields)})" if pushed_fields else ""),
    )
    db.commit()
    return {
        "success": True,
        "member_id": member.id,
        "external_member_id": external_id,
        "pushed_fields": pushed_fields,
    }


def _resolve_default_assembly_id(db: Session) -> Optional[str]:
    """Best-effort assembly UUID resolver for 'create new in API' flows:
      1. RFM_ASSEMBLY_ID env var or Settings entry
      2. external_assembly_id of any locally-matched member (we already store
         this when matching, so as long as one member is matched we know it)
      3. First assembly returned by the API (works when the service key is
         scoped to a single assembly — the common production setup)
    Returns the UUID string, or None if all three fall through."""
    import os as _os
    raw = _os.getenv("RFM_ASSEMBLY_ID")
    if raw:
        return raw.strip()

    from models import Settings as _Settings
    s = db.query(_Settings).filter(_Settings.key == "rfm_assembly_id").first()
    if s and s.value:
        return s.value.strip()

    # From any matched local member
    m = (
        db.query(Member)
        .filter(Member.external_assembly_id.isnot(None))
        .order_by(Member.external_synced_at.desc().nullslast() if hasattr(Member.external_synced_at, "desc") else Member.id.desc())
        .first()
    )
    if m and m.external_assembly_id:
        return str(m.external_assembly_id)

    # Last resort — call the API
    if _rfm.is_enabled(db) and _rfm.is_configured(db):
        result = _rfm.list_assemblies(db=db)
        if result.ok and result.data:
            data = result.data
            if isinstance(data, dict):
                data = data.get("data") or []
            if isinstance(data, list) and data:
                return str(data[0].get("id"))
    return None


@router.post("/admin/rfm-sync/create-from-local")
def admin_rfm_sync_create_from_local(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    """Push a local-only member up to the central database, then link the
    local member to the new external record.

    Used when a member exists in the portal but not in rfm-database (orphan
    where no manual search candidate is right either). The admin reviews/
    edits the derived first/last/phone/email in the UI panel and confirms.

    Body:
      {
        member_id:   int (required) — the local member to push,
        first_name:  str (required),
        last_name:   str (required),
        phone:       str (optional),
        email:       str (optional),
        physical_address: str (optional),
      }"""
    _require_committee_or_admin(request, db)
    if not _rfm.is_enabled(db):
        raise HTTPException(status_code=400, detail="rfm-db integration is disabled")
    if not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="rfm-db API not configured (URL/key missing)")

    try:
        member_id = int(data["member_id"])
    except (KeyError, ValueError, TypeError):
        raise HTTPException(status_code=400, detail="member_id required")

    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")

    first_name = str(data.get("first_name") or "").strip()
    last_name = str(data.get("last_name") or "").strip()
    if not first_name or not last_name:
        raise HTTPException(status_code=400, detail="first_name and last_name required")

    assembly_id = data.get("assembly_id") or _resolve_default_assembly_id(db)
    if not assembly_id:
        raise HTTPException(
            status_code=400,
            detail=(
                "Could not determine the assembly UUID to create this member under. "
                "Set the RFM_ASSEMBLY_ID env var, or match at least one existing member first."
            ),
        )

    # Validate optional fields with the same rules we use for enrichment
    payload = {
        "assembly_id": str(assembly_id),
        "first_name": first_name,
        "last_name": last_name,
    }
    raw_phone = (data.get("phone") or "").strip()
    if raw_phone:
        canonical = _rfm.to_sa_canonical_mobile(raw_phone)
        if canonical:
            payload["phone"] = canonical
    raw_email = (data.get("email") or "").strip()
    if raw_email and "@" in raw_email and "." in raw_email.split("@")[-1]:
        payload["email"] = raw_email
    raw_address = (data.get("physical_address") or "").strip()
    if raw_address:
        payload["physical_address"] = raw_address

    # Push it
    result = _rfm.create_member(payload, db=db)
    if not result.ok:
        raise HTTPException(status_code=502, detail=result.error or "Failed to create in central database")

    # Unwrap the new record
    new_record = result.data
    if isinstance(new_record, dict) and "data" in new_record:
        new_record = new_record["data"]
    new_id = (new_record or {}).get("id")
    if not new_id:
        raise HTTPException(status_code=502, detail="Central database did not return a new member id")

    # Link it locally
    from datetime import datetime as _dt
    member.external_member_id = str(new_id)
    member.external_assembly_id = str(assembly_id)
    member.external_match_status = "manual"
    member.external_synced_at = _dt.utcnow()
    db.commit()

    _log_admin_action(
        request, db, "rfm_sync_create_in_api", "member", member.id,
        f"Created {member.full_name} in central database (external id {new_id})",
    )
    db.commit()

    return {
        "success": True,
        "member_id": member.id,
        "external_member_id": str(new_id),
        "assembly_id": str(assembly_id),
        "pushed_fields": list(payload.keys()),
    }


@router.get("/admin/members/external-search")
def admin_members_external_search(
    request: Request,
    q: str = Query(..., min_length=2, description="Name or phone search query"),
    db: Session = Depends(get_db),
):
    """Search the central rfm-database for members the admin can import into
    the local portal. Excludes anyone already linked locally (their
    external_member_id matches an existing local row) so we don't show
    duplicates. Used by the admin Members page to surface central-only
    matches alongside local search results."""
    _require_committee_or_admin(request, db)
    if not _rfm.is_enabled(db):
        return {"enabled": False, "results": []}
    if not _rfm.is_configured(db):
        return {"enabled": True, "configured": False, "results": []}

    # Search by each significant token of the query (matches the auto-matcher
    # behaviour — phone or name substring matches per-field on the API).
    tokens = _rfm._significant_name_tokens(q)
    # Phone-only searches (digits) get added as-is too
    phone_norm = _rfm.normalise_phone(q)
    if phone_norm and len(phone_norm) >= 6 and phone_norm not in tokens:
        tokens.append(phone_norm)
    if not tokens:
        tokens = [q.strip()]

    seen_ids = set()
    raw_results = []
    for token in tokens:
        api_result = _rfm.search_members(search=token, page=1, size=50, db=db)
        if not api_result.ok or not api_result.data:
            continue
        items = api_result.data
        if isinstance(items, dict):
            items = items.get("data") or []
        for item in items:
            ext_id = item.get("id")
            if ext_id and ext_id not in seen_ids:
                seen_ids.add(ext_id)
                raw_results.append(item)

    # Drop ones already imported locally
    if seen_ids:
        already_linked = {
            row[0] for row in db.query(Member.external_member_id)
            .filter(Member.external_member_id.in_(list(seen_ids))).all()
        }
    else:
        already_linked = set()

    results = []
    for item in raw_results:
        if item.get("id") in already_linked:
            continue
        results.append(_rfm._score_and_strip(item, _rfm.name_match_score(q, item)))
    # Sort: best name-score first, then alphabetical
    results.sort(key=lambda r: (-r.get("_score", 0), (r.get("full_name") or "").lower()))

    return {
        "enabled": True,
        "configured": True,
        "query": q,
        "count": len(results),
        "results": results,
    }


@router.post("/admin/members/import-from-external")
def admin_members_import_from_external(
    request: Request,
    data: dict = Body(...),
    db: Session = Depends(get_db),
):
    """Import a member from the central database into the local portal,
    optionally assigning them to a department in the same call.

    Body:
      external_member_id: UUID (required) — the central record to import
      department_id:      int (optional) — if set, assign immediately
      phone:              str (optional) — required when central record has
                                           no phone (local schema requires it)

    Behaviour:
      * Refuses if a local row already links to this external_id (use the
        admin Members page to find the existing one and assign there)
      * Pulls the central record fresh
      * Builds local fields: full_name from first+last, phone from central
        (canonicalised) or override, email/address joined
      * Creates the local Member with status='approved' source='admin' and
        the link pre-populated
      * If department_id given, creates an approved MemberDepartment row
    """
    actor = _require_committee_or_admin(request, db)

    if not _rfm.is_enabled(db):
        raise HTTPException(status_code=400, detail="rfm-db integration is disabled")
    if not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="rfm-db API not configured")

    external_id = (data.get("external_member_id") or "").strip()
    if not external_id:
        raise HTTPException(status_code=400, detail="external_member_id required")

    # Reject duplicate links
    existing = db.query(Member).filter(Member.external_member_id == external_id).first()
    if existing:
        raise HTTPException(
            status_code=409,
            detail=f"Already imported as local member id {existing.id} ({existing.full_name})",
        )

    # Fetch the canonical record
    api_result = _rfm.get_member(external_id, db=db)
    if not api_result.ok or not isinstance(api_result.data, dict):
        raise HTTPException(
            status_code=502,
            detail=api_result.error or "Could not fetch central member record",
        )
    rec = api_result.data

    full_name = _rfm.fullname_from_member(rec) or "Unknown"
    central_phone = _rfm._clean_phone_for_display(rec.get("phone"))
    override_phone = (data.get("phone") or "").strip()
    if override_phone:
        canonical = _rfm.to_sa_canonical_mobile(override_phone)
        phone = canonical or override_phone
    else:
        phone = central_phone
    if not phone:
        raise HTTPException(
            status_code=400,
            detail="Central record has no phone — provide one in the request body",
        )

    email = (rec.get("email") or "").strip()
    address = _rfm.address_from_member(rec)

    from datetime import datetime as _dt
    member = Member(
        full_name=full_name,
        phone=phone,
        email=email or "",
        address=address or "",
        is_active=True,
        external_member_id=external_id,
        external_assembly_id=rec.get("assembly_id"),
        external_match_status="manual",
        external_synced_at=_dt.utcnow(),
    )
    db.add(member)
    db.commit()
    db.refresh(member)

    assigned_department = None
    department_id = data.get("department_id")
    if department_id:
        try:
            department_id_int = int(department_id)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="department_id must be an integer")
        dept = db.query(Department).options(joinedload(Department.category)).filter(
            Department.id == department_id_int
        ).first()
        if not dept:
            # Don't roll back the import — admin can assign later
            assigned_department = {"error": "Department not found, member imported without assignment"}
        else:
            md = MemberDepartment(
                member_id=member.id,
                department_id=dept.id,
                status="approved",
                source="admin",
                status_changed_at=_dt.utcnow(),
            )
            db.add(md)
            db.commit()
            assigned_department = {"id": dept.id, "name": dept.name}

    _log_admin_action(
        request, db, "import_member_from_external", "member", member.id,
        f"Imported {member.full_name} from central database (external {external_id})"
        + (f"; assigned to {assigned_department['name']}" if assigned_department and assigned_department.get('name') else ""),
    )
    db.commit()

    return {
        "success": True,
        "member_id": member.id,
        "external_member_id": external_id,
        "assigned_department": assigned_department,
    }


@router.post("/admin/rfm-sync/memberships/sync-all")
def admin_rfm_sync_memberships_all(request: Request, db: Session = Depends(get_db)):
    """Push every locally-linked member's approved department assignments
    to the central API. Idempotent — safe to run as many times as needed.
    Members not yet linked, and assignments to departments not yet linked,
    are silently skipped (with counts surfaced in the summary)."""
    _require_committee_or_admin(request, db)
    if not _rfm.is_enabled(db):
        raise HTTPException(status_code=400, detail="rfm-db integration is disabled")
    if not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="rfm-db API not configured")

    summary = {
        "members_scanned": 0,
        "members_synced": 0,
        "members_unlinked": 0,
        "members_no_changes": 0,
        "errors": 0,
        "total_assignments_pushed": 0,
        "skipped_unlinked_depts": 0,
    }
    errors_sample = []

    # Only members with an external_member_id can be synced; others are orphans
    # and must be matched / imported first via the rfm-sync UI.
    linked_members = db.query(Member).filter(Member.external_member_id.isnot(None)).all()
    summary["members_scanned"] = len(linked_members)

    for m in linked_members:
        result = _sync_member_departments_to_central(m.id, db)
        if not result["attempted"]:
            summary["members_unlinked"] += 1
            continue
        if result["ok"]:
            summary["members_synced"] += 1
            summary["total_assignments_pushed"] += result["department_count"]
            summary["skipped_unlinked_depts"] += result["skipped_unlinked"]
        else:
            summary["errors"] += 1
            if len(errors_sample) < 5:
                errors_sample.append({
                    "member_id": m.id,
                    "name": m.full_name,
                    "error": result["error"],
                })

    # Members who weren't linked don't count as errors — they're just not
    # ready yet (need to go through the matcher first).
    summary["members_unlinked"] = summary["members_scanned"] - summary["members_synced"] - summary["errors"]

    _log_admin_action(
        request, db, "rfm_sync_memberships", "members", None,
        f"Synced {summary['members_synced']}/{summary['members_scanned']} members "
        f"({summary['total_assignments_pushed']} dept assignments pushed, "
        f"{summary['errors']} errors, {summary['skipped_unlinked_depts']} dept(s) not yet linked)",
    )
    db.commit()

    return {"summary": summary, "errors_sample": errors_sample}


@router.post("/admin/rfm-sync/departments/sync-all")
def admin_rfm_sync_departments_all(request: Request, db: Session = Depends(get_db)):
    """Push every local department up to the central database, linking by
    name where the central record already exists. Idempotent — running twice
    is safe; already-linked departments are skipped."""
    _require_committee_or_admin(request, db)
    if not _rfm.is_enabled(db):
        raise HTTPException(status_code=400, detail="rfm-db integration is disabled")
    if not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="rfm-db API not configured")

    summary = {"linked": 0, "created": 0, "errors": 0, "skipped": 0}
    errors_sample = []

    departments = db.query(Department).all()
    for d in departments:
        if d.external_department_id:
            summary["skipped"] += 1
            continue
        result = _push_department_to_central(d, request, db, action="create")
        if result["ok"]:
            if "created" in result["fields"]:
                summary["created"] += 1
            else:
                summary["linked"] += 1
        else:
            summary["errors"] += 1
            if len(errors_sample) < 5:
                errors_sample.append({"department_id": d.id, "name": d.name, "error": result["error"]})

    _log_admin_action(
        request, db, "rfm_sync_departments", "departments", None,
        f"Linked {summary['linked']}, created {summary['created']}, "
        f"skipped {summary['skipped']}, errors {summary['errors']}",
    )
    db.commit()
    return {"summary": summary, "errors_sample": errors_sample, "scanned": len(departments)}


@router.post("/admin/rfm-sync/home-churches/sync-all")
def admin_rfm_sync_home_churches_all(request: Request, db: Session = Depends(get_db)):
    """Push every local home church up to the central database. Idempotent —
    already-linked rows are refreshed (PUT) so updates flow up too."""
    _require_committee_or_admin(request, db)
    if not _rfm.is_enabled(db):
        raise HTTPException(status_code=400, detail="rfm-db integration is disabled")
    if not _rfm.is_configured(db):
        raise HTTPException(status_code=400, detail="rfm-db API not configured")

    summary = {"linked": 0, "created": 0, "updated": 0, "errors": 0, "skipped": 0}
    errors_sample = []

    rows = db.query(HomeChurch).all()
    for hc in rows:
        action = "update" if hc.external_home_church_id else "create"
        result = _push_home_church_to_central(hc, request, db, action=action)
        if result["ok"]:
            if "created" in result["fields"]:
                summary["created"] += 1
            elif "linked-by-name" in result["fields"]:
                summary["linked"] += 1
            else:
                summary["updated"] += 1
        elif not result["attempted"]:
            summary["skipped"] += 1
        else:
            summary["errors"] += 1
            if len(errors_sample) < 5:
                errors_sample.append({"home_church_id": hc.id, "name": hc.name, "error": result["error"]})

    _log_admin_action(
        request, db, "rfm_sync_home_churches", "home_churches", None,
        f"Linked {summary['linked']}, created {summary['created']}, "
        f"updated {summary['updated']}, skipped {summary['skipped']}, errors {summary['errors']}",
    )
    db.commit()
    return {"summary": summary, "errors_sample": errors_sample, "scanned": len(rows)}


@router.post("/admin/rfm-sync/unlink")
def admin_rfm_sync_unlink(request: Request, data: dict = Body(...), db: Session = Depends(get_db)):
    """Remove the link for a local member; resets to pending."""
    _require_committee_or_admin(request, db)
    try:
        member_id = int(data["member_id"])
    except (KeyError, ValueError, TypeError):
        raise HTTPException(status_code=400, detail="member_id required")
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    member.external_member_id = None
    member.external_match_status = None
    member.external_synced_at = None
    member.external_assembly_id = None
    db.commit()
    _log_admin_action(
        request, db, "rfm_sync_unlink", "member", member.id,
        f"Unlinked {member.full_name} from external database",
    )
    db.commit()
    return {"success": True}


# ============ SURVEYS ============
#
# Permission model
#   - admin (admin_session cookie)              -> full control of all surveys + creator allowlist
#   - member with can_create_surveys=True       -> CRUD their OWN surveys, view their OWN responses
#   - public (no auth)                          -> view active survey by slug, submit one response
#
# Anonymity guarantee
#   When survey.is_anonymous is True, the response endpoint NEVER stores
#   respondent_member_id and NEVER stores respondent_name. Server-side enforced.

VALID_QUESTION_TYPES = {"text", "long_text", "single_choice", "multi_choice", "rating", "yes_no"}


def _is_admin(request: Request) -> bool:
    from routers.pages import is_authenticated
    return is_authenticated(request)


def _current_member(request: Request, db: Session) -> Optional[Member]:
    from routers.pages import get_current_member
    return get_current_member(request, db)


def _can_manage_survey(request: Request, db: Session, survey: Survey) -> bool:
    if _is_admin(request):
        return True
    member = _current_member(request, db)
    if not member or not getattr(member, "can_create_surveys", False):
        return False
    return survey.created_by_member_id == member.id


def _generate_survey_slug(db: Session) -> str:
    """Short, URL-safe, unguessable slug. Re-tries on the (extremely rare) collision."""
    import secrets
    for _ in range(5):
        slug = secrets.token_urlsafe(8).replace("_", "").replace("-", "")[:12].lower()
        if not slug:
            continue
        exists = db.query(Survey).filter(Survey.slug == slug).first()
        if not exists:
            return slug
    # Last-resort fallback
    return uuid.uuid4().hex[:12]


def _serialize_question(q: SurveyQuestion, *, include_answers: bool = False) -> dict:
    options = []
    if q.options:
        try:
            options = json.loads(q.options) or []
        except (ValueError, TypeError):
            options = []
    return {
        "id": q.id,
        "position": q.position,
        "question_text": q.question_text,
        "question_type": q.question_type,
        "options": options,
        "required": bool(q.required),
    }


def _serialize_survey(s: Survey, *, include_questions: bool = True, include_counts: bool = False, db: Session = None) -> dict:
    out = {
        "id": s.id,
        "title": s.title,
        "description": s.description or "",
        "slug": s.slug,
        "is_anonymous": bool(s.is_anonymous),
        "is_active": bool(s.is_active),
        "allow_multiple_responses": bool(getattr(s, "allow_multiple_responses", False)),
        "created_at": s.created_at.isoformat() if s.created_at else None,
        "closed_at": s.closed_at.isoformat() if s.closed_at else None,
        "created_by": {
            "id": s.created_by.id,
            "full_name": s.created_by.full_name,
        } if s.created_by else None,
    }
    if include_questions:
        out["questions"] = [_serialize_question(q) for q in sorted(s.questions, key=lambda x: x.position)]
    if include_counts and db is not None:
        out["response_count"] = db.query(SurveyResponse).filter(SurveyResponse.survey_id == s.id).count()
    return out


def _validate_question_payload(q: dict) -> Tuple[Optional[dict], Optional[str]]:
    """Returns (clean_dict, error). clean_dict has fields ready for the model."""
    text = (q.get("question_text") or "").strip()
    qtype = (q.get("question_type") or "").strip()
    if not text:
        return None, "Question text is required"
    if qtype not in VALID_QUESTION_TYPES:
        return None, f"Unknown question type: {qtype}"
    options = q.get("options") or []
    if qtype in ("single_choice", "multi_choice"):
        clean_opts = [str(o).strip() for o in options if str(o).strip()]
        if len(clean_opts) < 2:
            return None, f"'{text[:40]}': choice questions need at least 2 options"
        options_json = json.dumps(clean_opts)
    else:
        options_json = None
    return {
        "question_text": text,
        "question_type": qtype,
        "options": options_json,
        "required": bool(q.get("required", False)),
    }, None


@router.get("/surveys")
def list_surveys(request: Request, db: Session = Depends(get_db)):
    """List surveys: admin sees all, creators see their own."""
    if _is_admin(request):
        rows = db.query(Survey).options(joinedload(Survey.created_by)).order_by(Survey.created_at.desc()).all()
        scope = "admin"
    else:
        member = _current_member(request, db)
        if not member or not getattr(member, "can_create_surveys", False):
            raise HTTPException(status_code=403, detail="You don't have permission to manage surveys")
        rows = (
            db.query(Survey)
            .options(joinedload(Survey.created_by))
            .filter(Survey.created_by_member_id == member.id)
            .order_by(Survey.created_at.desc())
            .all()
        )
        scope = "creator"
    return {
        "scope": scope,
        "surveys": [_serialize_survey(s, include_questions=False, include_counts=True, db=db) for s in rows],
    }


@router.post("/surveys")
def create_survey(payload: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    """Create a survey. Admin or any member with can_create_surveys."""
    creator_member_id = None
    if _is_admin(request):
        from routers.pages import get_admin_identity
        identity = get_admin_identity(request)
        creator_member_id = (identity or {}).get("member_id")
    else:
        member = _current_member(request, db)
        if not member or not getattr(member, "can_create_surveys", False):
            raise HTTPException(status_code=403, detail="You don't have permission to create surveys")
        creator_member_id = member.id

    title = (payload.get("title") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title is required")

    questions_payload = payload.get("questions") or []
    if not isinstance(questions_payload, list) or len(questions_payload) == 0:
        raise HTTPException(status_code=400, detail="At least one question is required")

    # Validate questions before any DB writes
    cleaned: List[dict] = []
    for q in questions_payload:
        clean, err = _validate_question_payload(q)
        if err:
            raise HTTPException(status_code=400, detail=err)
        cleaned.append(clean)

    survey = Survey(
        title=title[:300],
        description=(payload.get("description") or "").strip() or None,
        slug=_generate_survey_slug(db),
        is_anonymous=bool(payload.get("is_anonymous", True)),
        is_active=bool(payload.get("is_active", True)),
        allow_multiple_responses=bool(payload.get("allow_multiple_responses", False)),
        created_by_member_id=creator_member_id,
    )
    db.add(survey)
    db.flush()
    for idx, c in enumerate(cleaned):
        db.add(SurveyQuestion(
            survey_id=survey.id,
            position=idx,
            question_text=c["question_text"],
            question_type=c["question_type"],
            options=c["options"],
            required=c["required"],
        ))
    db.commit()
    db.refresh(survey)
    return _serialize_survey(survey, include_questions=True)


@router.get("/surveys/{survey_id}")
def get_survey(survey_id: int, request: Request, db: Session = Depends(get_db)):
    survey = db.query(Survey).options(joinedload(Survey.created_by)).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not _can_manage_survey(request, db, survey):
        raise HTTPException(status_code=403, detail="Not allowed")
    return _serialize_survey(survey, include_questions=True, include_counts=True, db=db)


@router.put("/surveys/{survey_id}")
def update_survey(survey_id: int, payload: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not _can_manage_survey(request, db, survey):
        raise HTTPException(status_code=403, detail="Not allowed")

    # Top-level fields
    if "title" in payload:
        title = (payload.get("title") or "").strip()
        if not title:
            raise HTTPException(status_code=400, detail="Title cannot be empty")
        survey.title = title[:300]
    if "description" in payload:
        survey.description = (payload.get("description") or "").strip() or None
    if "is_anonymous" in payload:
        # Once responses exist, lock the anonymity setting to preserve guarantees.
        existing_response_count = db.query(SurveyResponse).filter(SurveyResponse.survey_id == survey.id).count()
        if existing_response_count > 0 and bool(payload["is_anonymous"]) != bool(survey.is_anonymous):
            raise HTTPException(status_code=400, detail="Cannot change anonymity after responses have been collected")
        survey.is_anonymous = bool(payload["is_anonymous"])
    if "is_active" in payload:
        survey.is_active = bool(payload["is_active"])
        survey.closed_at = None if survey.is_active else datetime.utcnow()
    if "allow_multiple_responses" in payload:
        survey.allow_multiple_responses = bool(payload["allow_multiple_responses"])

    # Questions: full replace if provided
    if "questions" in payload:
        questions_payload = payload.get("questions") or []
        if not isinstance(questions_payload, list) or len(questions_payload) == 0:
            raise HTTPException(status_code=400, detail="At least one question is required")
        cleaned: List[dict] = []
        for q in questions_payload:
            clean, err = _validate_question_payload(q)
            if err:
                raise HTTPException(status_code=400, detail=err)
            cleaned.append(clean)
        # If responses already exist, we keep historical questions intact by
        # blocking question edits — first iteration: keep it simple.
        existing_response_count = db.query(SurveyResponse).filter(SurveyResponse.survey_id == survey.id).count()
        if existing_response_count > 0:
            raise HTTPException(status_code=400, detail="Cannot edit questions after responses have been collected")
        # wipe & recreate
        db.query(SurveyQuestion).filter(SurveyQuestion.survey_id == survey.id).delete()
        for idx, c in enumerate(cleaned):
            db.add(SurveyQuestion(
                survey_id=survey.id,
                position=idx,
                question_text=c["question_text"],
                question_type=c["question_type"],
                options=c["options"],
                required=c["required"],
            ))

    db.commit()
    db.refresh(survey)
    return _serialize_survey(survey, include_questions=True, include_counts=True, db=db)


@router.delete("/surveys/{survey_id}")
def delete_survey(survey_id: int, request: Request, db: Session = Depends(get_db)):
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not _can_manage_survey(request, db, survey):
        raise HTTPException(status_code=403, detail="Not allowed")
    db.delete(survey)
    db.commit()
    return {"success": True}


@router.get("/surveys/{survey_id}/responses")
def get_survey_responses(survey_id: int, request: Request, db: Session = Depends(get_db)):
    survey = (
        db.query(Survey)
        .options(joinedload(Survey.questions))
        .filter(Survey.id == survey_id)
        .first()
    )
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not _can_manage_survey(request, db, survey):
        raise HTTPException(status_code=403, detail="Not allowed")

    responses = (
        db.query(SurveyResponse)
        .options(joinedload(SurveyResponse.answers))
        .filter(SurveyResponse.survey_id == survey_id)
        .order_by(SurveyResponse.submitted_at.desc())
        .all()
    )

    questions_sorted = sorted(survey.questions, key=lambda q: q.position)
    q_lookup = {q.id: q for q in questions_sorted}

    rows = []
    for r in responses:
        answers_by_q = {}
        for a in r.answers:
            opts = []
            if a.answer_options:
                try:
                    opts = json.loads(a.answer_options) or []
                except (ValueError, TypeError):
                    opts = []
            answers_by_q[a.question_id] = {
                "answer_text": a.answer_text or "",
                "answer_options": opts,
            }
        rows.append({
            "id": r.id,
            "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
            # NEVER include identity for anonymous surveys
            "respondent_name": (r.respondent_name if not survey.is_anonymous else None),
            "answers": answers_by_q,
        })

    # Per-question summary stats (counts for choice/yes_no/rating, sample texts otherwise)
    summaries = []
    for q in questions_sorted:
        qsum: dict = {"question_id": q.id, "question_text": q.question_text, "question_type": q.question_type}
        if q.question_type in ("single_choice", "yes_no", "rating"):
            counts: dict = {}
            for r in responses:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_text:
                    counts[a.answer_text] = counts.get(a.answer_text, 0) + 1
            qsum["counts"] = counts
        elif q.question_type == "multi_choice":
            counts = {}
            for r in responses:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_options:
                    try:
                        for opt in json.loads(a.answer_options) or []:
                            counts[opt] = counts.get(opt, 0) + 1
                    except (ValueError, TypeError):
                        pass
            qsum["counts"] = counts
        else:
            samples = []
            for r in responses[:50]:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_text:
                    samples.append(a.answer_text)
            qsum["samples"] = samples
        summaries.append(qsum)

    return {
        "survey": _serialize_survey(survey, include_questions=True, include_counts=False),
        "is_anonymous": bool(survey.is_anonymous),
        "responses": rows,
        "summaries": summaries,
        "total": len(responses),
    }


# --- Public endpoints ---

@router.get("/surveys/public/{slug}")
def get_public_survey(slug: str, db: Session = Depends(get_db)):
    """Return survey + questions for the public response page."""
    survey = db.query(Survey).filter(Survey.slug == slug).first()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not survey.is_active:
        return {
            "closed": True,
            "title": survey.title,
            "description": survey.description or "",
            "is_anonymous": bool(survey.is_anonymous),
        }
    return {
        "closed": False,
        "id": survey.id,
        "title": survey.title,
        "description": survey.description or "",
        "is_anonymous": bool(survey.is_anonymous),
        "allow_multiple_responses": bool(getattr(survey, "allow_multiple_responses", False)),
        "questions": [_serialize_question(q) for q in sorted(survey.questions, key=lambda x: x.position)],
    }


@router.post("/surveys/public/{slug}/respond")
def submit_public_response(slug: str, payload: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    """Public response submission. Anonymity is enforced server-side."""
    survey = (
        db.query(Survey)
        .options(joinedload(Survey.questions))
        .filter(Survey.slug == slug)
        .first()
    )
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not survey.is_active:
        raise HTTPException(status_code=400, detail="This survey is closed")

    answers_payload = payload.get("answers") or {}  # { question_id: {answer_text?, answer_options?} }
    if not isinstance(answers_payload, dict):
        raise HTTPException(status_code=400, detail="Invalid answers payload")

    questions_by_id = {q.id: q for q in survey.questions}

    # Validate required + structure
    for q in survey.questions:
        a = answers_payload.get(str(q.id)) or answers_payload.get(q.id) or {}
        text = (a.get("answer_text") or "").strip() if isinstance(a, dict) else ""
        opts = a.get("answer_options") if isinstance(a, dict) else None
        has_value = bool(text) or (isinstance(opts, list) and len(opts) > 0)
        if q.required and not has_value:
            raise HTTPException(status_code=400, detail=f"'{q.question_text[:60]}' is required")
        if q.question_type in ("single_choice", "yes_no", "rating") and text:
            # accept anything for text in 'rating' (numeric string)
            if q.question_type in ("single_choice",):
                allowed = []
                try:
                    allowed = json.loads(q.options) if q.options else []
                except (ValueError, TypeError):
                    allowed = []
                if allowed and text not in allowed:
                    raise HTTPException(status_code=400, detail=f"Invalid choice for '{q.question_text[:60]}'")
        if q.question_type == "multi_choice" and isinstance(opts, list):
            try:
                allowed = json.loads(q.options) if q.options else []
            except (ValueError, TypeError):
                allowed = []
            for o in opts:
                if allowed and o not in allowed:
                    raise HTTPException(status_code=400, detail=f"Invalid option for '{q.question_text[:60]}'")

    # SERVER-ENFORCED ANONYMITY: ignore identity fields entirely if anonymous.
    respondent_member_id = None
    respondent_name = None
    if not survey.is_anonymous:
        # Allow optional self-identification on non-anonymous surveys.
        member = _current_member(request, db)
        if member:
            respondent_member_id = member.id
            respondent_name = member.full_name
        else:
            # accept a name string from the form (no auth)
            name = (payload.get("respondent_name") or "").strip()
            respondent_name = name[:200] if name else None

    response = SurveyResponse(
        survey_id=survey.id,
        respondent_member_id=respondent_member_id,
        respondent_name=respondent_name,
    )
    db.add(response)
    db.flush()

    for q in survey.questions:
        a = answers_payload.get(str(q.id)) or answers_payload.get(q.id) or {}
        text = (a.get("answer_text") or "").strip() if isinstance(a, dict) else ""
        opts = a.get("answer_options") if isinstance(a, dict) else None
        opts_json = None
        if isinstance(opts, list) and opts:
            cleaned_opts = [str(o) for o in opts]
            opts_json = json.dumps(cleaned_opts)
        if not text and not opts_json:
            continue
        db.add(SurveyAnswer(
            response_id=response.id,
            question_id=q.id,
            answer_text=text or None,
            answer_options=opts_json,
        ))

    db.commit()
    return {"success": True}


# --- Admin: manage who can create surveys ---

@router.delete("/surveys/{survey_id}/responses/{response_id}")
def delete_survey_response(survey_id: int, response_id: int, request: Request, db: Session = Depends(get_db)):
    """Delete a single response. Admin or the survey creator only."""
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not _can_manage_survey(request, db, survey):
        raise HTTPException(status_code=403, detail="Not allowed")
    resp = (
        db.query(SurveyResponse)
        .filter(SurveyResponse.id == response_id, SurveyResponse.survey_id == survey_id)
        .first()
    )
    if not resp:
        raise HTTPException(status_code=404, detail="Response not found")
    db.delete(resp)
    db.commit()
    return {"success": True}


@router.delete("/surveys/{survey_id}/responses")
def delete_all_survey_responses(survey_id: int, request: Request, db: Session = Depends(get_db)):
    """Wipe every response on a survey (keeps the survey & its questions)."""
    survey = db.query(Survey).filter(Survey.id == survey_id).first()
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not _can_manage_survey(request, db, survey):
        raise HTTPException(status_code=403, detail="Not allowed")
    deleted = db.query(SurveyResponse).filter(SurveyResponse.survey_id == survey_id).delete()
    db.commit()
    return {"success": True, "deleted": deleted}


@router.get("/surveys/{survey_id}/export.pdf")
def export_survey_responses_pdf(survey_id: int, request: Request, db: Session = Depends(get_db)):
    """PDF export of all responses with per-question summaries.
    Uses ReportLab if available; falls back to a clean printable HTML otherwise."""
    survey = (
        db.query(Survey)
        .options(joinedload(Survey.questions))
        .filter(Survey.id == survey_id)
        .first()
    )
    if not survey:
        raise HTTPException(status_code=404, detail="Survey not found")
    if not _can_manage_survey(request, db, survey):
        raise HTTPException(status_code=403, detail="Not allowed")

    responses = (
        db.query(SurveyResponse)
        .options(joinedload(SurveyResponse.answers))
        .filter(SurveyResponse.survey_id == survey_id)
        .order_by(SurveyResponse.submitted_at.desc())
        .all()
    )
    questions_sorted = sorted(survey.questions, key=lambda q: q.position)

    # Build summaries
    summaries = []
    for q in questions_sorted:
        s = {"q": q, "counts": None, "samples": None}
        if q.question_type in ("single_choice", "yes_no", "rating"):
            counts: dict = {}
            for r in responses:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_text:
                    counts[a.answer_text] = counts.get(a.answer_text, 0) + 1
            s["counts"] = counts
        elif q.question_type == "multi_choice":
            counts = {}
            for r in responses:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_options:
                    try:
                        for opt in json.loads(a.answer_options) or []:
                            counts[opt] = counts.get(opt, 0) + 1
                    except (ValueError, TypeError):
                        pass
            s["counts"] = counts
        else:
            samples = []
            for r in responses:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_text:
                    samples.append(a.answer_text)
            s["samples"] = samples
        summaries.append(s)

    # Try ReportLab; otherwise fall back to a printable HTML page.
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        )
    except Exception:
        # Fallback: HTML the browser can print to PDF.
        return _survey_pdf_html_fallback(survey, questions_sorted, summaries, responses)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=18 * mm,
        title=f"Survey: {survey.title}",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#4f46e5"), spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#1f2937"), spaceBefore=12, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=13)
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=8, textColor=colors.HexColor("#6b7280"))
    qstyle = ParagraphStyle("q", parent=styles["BodyText"], fontSize=11, textColor=colors.HexColor("#111827"), spaceBefore=8, spaceAfter=2, leading=14)

    flow = []
    flow.append(Paragraph(survey.title, h1))
    flow.append(Paragraph(
        f"{'Anonymous · ' if survey.is_anonymous else ''}"
        f"{len(responses)} response{'s' if len(responses) != 1 else ''} · "
        f"Created {survey.created_at.strftime('%d %b %Y') if survey.created_at else ''}",
        small,
    ))
    if survey.description:
        flow.append(Spacer(1, 4))
        flow.append(Paragraph(survey.description, body))

    # Summary section
    flow.append(Paragraph("Summary", h2))
    for s in summaries:
        q = s["q"]
        flow.append(Paragraph(f"<b>{q.question_text}</b> <font size=8 color='#6b7280'>({q.question_type.replace('_', ' ')})</font>", qstyle))
        if s["counts"] is not None:
            counts = s["counts"]
            total = sum(counts.values()) or 1
            rows = [["Option", "Count", "%"]]
            for opt, n in sorted(counts.items(), key=lambda kv: -kv[1]):
                rows.append([opt, str(n), f"{round((n / total) * 100)}%"])
            if len(rows) == 1:
                flow.append(Paragraph("<i>No answers yet.</i>", small))
            else:
                t = Table(rows, hAlign="LEFT", colWidths=[80 * mm, 25 * mm, 25 * mm])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#3730a3")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
                ]))
                flow.append(t)
        else:
            samples = s["samples"] or []
            if not samples:
                flow.append(Paragraph("<i>No answers yet.</i>", small))
            else:
                for txt in samples[:200]:
                    flow.append(Paragraph(f"• {txt}", body))

    # Per-response detail
    if responses:
        flow.append(PageBreak())
        flow.append(Paragraph("Individual responses", h2))
        for idx, r in enumerate(responses, 1):
            who = "Anonymous" if survey.is_anonymous else (r.respondent_name or "—")
            when = r.submitted_at.strftime("%d %b %Y %H:%M") if r.submitted_at else ""
            flow.append(Paragraph(f"<b>#{idx}</b> &nbsp; {who} &nbsp; <font color='#6b7280'>· {when}</font>", qstyle))
            for q in questions_sorted:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_options:
                    try:
                        ans = ", ".join(json.loads(a.answer_options) or [])
                    except (ValueError, TypeError):
                        ans = "—"
                elif a and a.answer_text:
                    ans = a.answer_text
                else:
                    ans = "—"
                flow.append(Paragraph(f"<font color='#6b7280'>{q.question_text}</font>", small))
                flow.append(Paragraph(ans, body))
            flow.append(Spacer(1, 6))

    doc.build(flow)
    buf.seek(0)
    safe_title = re.sub(r"[^A-Za-z0-9._-]+", "_", survey.title or "survey").strip("_") or "survey"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_title}_responses.pdf"'},
    )


def _survey_pdf_html_fallback(survey: Survey, questions_sorted: list, summaries: list, responses: list) -> Response:
    """Printable HTML fallback when ReportLab isn't installed.
    Browsers' Print-to-PDF produces a clean PDF from this."""
    from html import escape
    rows = []
    rows.append(f"<h1 style='color:#4f46e5;margin:0 0 4px 0'>{escape(survey.title)}</h1>")
    rows.append(f"<p style='color:#6b7280;margin:0 0 12px 0;font-size:12px'>"
                f"{'Anonymous · ' if survey.is_anonymous else ''}{len(responses)} response(s) · "
                f"Created {survey.created_at.strftime('%d %b %Y') if survey.created_at else ''}</p>")
    if survey.description:
        rows.append(f"<p>{escape(survey.description)}</p>")

    rows.append("<h2 style='margin-top:24px;color:#1f2937'>Summary</h2>")
    for s in summaries:
        q = s["q"]
        rows.append(f"<p style='margin:14px 0 4px 0'><b>{escape(q.question_text)}</b> "
                    f"<span style='font-size:11px;color:#6b7280'>({q.question_type.replace('_', ' ')})</span></p>")
        if s["counts"] is not None:
            counts = s["counts"]
            total = sum(counts.values()) or 1
            if not counts:
                rows.append("<p style='color:#9ca3af'><i>No answers yet.</i></p>")
            else:
                rows.append("<table border='1' cellspacing='0' cellpadding='6' style='border-collapse:collapse;font-size:12px;width:100%;border-color:#e5e7eb'>")
                rows.append("<tr style='background:#eef2ff;color:#3730a3'><th align='left'>Option</th><th>Count</th><th>%</th></tr>")
                for opt, n in sorted(counts.items(), key=lambda kv: -kv[1]):
                    rows.append(f"<tr><td>{escape(opt)}</td><td align='center'>{n}</td><td align='center'>{round((n / total) * 100)}%</td></tr>")
                rows.append("</table>")
        else:
            samples = s["samples"] or []
            if not samples:
                rows.append("<p style='color:#9ca3af'><i>No answers yet.</i></p>")
            else:
                rows.append("<ul>" + "".join(f"<li>{escape(t)}</li>" for t in samples[:200]) + "</ul>")

    if responses:
        rows.append("<h2 style='margin-top:24px;color:#1f2937;page-break-before:always'>Individual responses</h2>")
        for idx, r in enumerate(responses, 1):
            who = "Anonymous" if survey.is_anonymous else (r.respondent_name or "—")
            when = r.submitted_at.strftime("%d %b %Y %H:%M") if r.submitted_at else ""
            rows.append(f"<div style='margin-top:12px'><b>#{idx}</b> &nbsp; {escape(who)} "
                        f"<span style='color:#6b7280'>· {when}</span></div>")
            for q in questions_sorted:
                a = next((x for x in r.answers if x.question_id == q.id), None)
                if a and a.answer_options:
                    try:
                        ans = ", ".join(json.loads(a.answer_options) or [])
                    except (ValueError, TypeError):
                        ans = "—"
                elif a and a.answer_text:
                    ans = a.answer_text
                else:
                    ans = "—"
                rows.append(f"<p style='margin:6px 0 2px;color:#6b7280;font-size:11px'>{escape(q.question_text)}</p>"
                            f"<p style='margin:0 0 4px;font-size:13px'>{escape(ans)}</p>")

    html = (
        "<!doctype html><html><head><meta charset='utf-8'>"
        f"<title>{escape(survey.title)} — responses</title>"
        "<style>body{font-family:Inter,Arial,sans-serif;max-width:800px;margin:24px auto;padding:0 16px;color:#111827}"
        "@media print{body{margin:0}}</style>"
        "<script>window.onload=()=>setTimeout(()=>window.print(),300)</script>"
        "</head><body>" + "".join(rows) + "</body></html>"
    )
    return Response(content=html, media_type="text/html")


@router.get("/admin/survey-creators")
def list_survey_creators(request: Request, db: Session = Depends(get_db)):
    if not _is_admin(request):
        raise HTTPException(status_code=403, detail="Admin only")
    rows = db.query(Member).filter(Member.can_create_surveys == True).order_by(Member.full_name).all()
    return [
        {"id": m.id, "full_name": m.full_name, "phone": m.phone, "email": m.email}
        for m in rows
    ]


@router.put("/admin/survey-creators/{member_id}")
def set_survey_creator(member_id: int, payload: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    if not _is_admin(request):
        raise HTTPException(status_code=403, detail="Admin only")
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    can = bool(payload.get("can_create_surveys", False))
    member.can_create_surveys = can
    _log_admin_action(
        request, db, "survey_creator_toggle", "member", member.id,
        f"{'Granted' if can else 'Revoked'} survey-creator for {member.full_name}",
    )
    db.commit()
    return {"id": member.id, "full_name": member.full_name, "can_create_surveys": can}


# ============ MEMBER PORTAL (member-centric) ============
#
# These endpoints power the redesigned portal. Each call requires a logged-in
# member. The portal aggregates local data with a best-effort fetch from the
# central rfm-database for ministries and home church (which live there).

def _require_logged_in_member(request: Request, db: Session) -> Member:
    from routers.pages import get_current_member
    member = get_current_member(request, db)
    if not member:
        raise HTTPException(status_code=401, detail="Please log in")
    return member


def _portal_central_lookup(member: Member, db: Session) -> dict:
    """Best-effort fetch of ministries + home church + assembly from the central
    API. Returns {} on any failure — never raises so the portal always loads."""
    out: dict = {"ministries": [], "home_church": None, "central_synced": False, "assembly_name": None}
    if not member.external_member_id:
        return out
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        return out
    try:
        r = _rfm.get_member(member.external_member_id, db=db)
        if not r.ok or not isinstance(r.data, dict):
            return out
        api_member = r.data
        out["central_synced"] = True
        # Resolve assembly name. The central member API carries assembly_id
        # (a UUID) but no assembly_name, so we look it up against list_assemblies.
        # We match by the member's OWN assembly_id rather than just taking the
        # first assembly in the list — that was the previous bug when an admin
        # key returned multiple assemblies.
        assembly_id = api_member.get("assembly_id") or member.external_assembly_id
        if assembly_id:
            try:
                ar = _rfm.list_assemblies(db=db)
                if ar.ok:
                    items = ar.data if isinstance(ar.data, list) else (ar.data or {}).get("data") or []
                    target = str(assembly_id)
                    match = next((a for a in items if str(a.get("id") or "") == target), None)
                    if match:
                        out["assembly_name"] = match.get("name") or None
                    elif items:
                        # Last-ditch fallback (single-assembly scoped keys)
                        out["assembly_name"] = items[0].get("name") or None
            except Exception:
                pass
        # Persist the assembly_id locally if we just learnt it (helps next time)
        if assembly_id and member.external_assembly_id != str(assembly_id):
            try:
                member.external_assembly_id = str(assembly_id)
                db.commit()
            except Exception:
                db.rollback()
        ministries_raw = api_member.get("ministries") or []
        ministries: List[dict] = []
        for m in ministries_raw:
            if isinstance(m, str):
                ministries.append({"name": m})
            elif isinstance(m, dict):
                ministries.append({"id": m.get("id"), "name": m.get("name") or ""})
        out["ministries"] = ministries
        hc_id = api_member.get("home_church_id")
        hc_name = api_member.get("home_church_name") or api_member.get("home_church")
        if hc_id or hc_name:
            local_hc = None
            if hc_id:
                local_hc = db.query(HomeChurch).filter(HomeChurch.external_home_church_id == str(hc_id)).first()
            if not local_hc and hc_name:
                local_hc = db.query(HomeChurch).filter(HomeChurch.name == hc_name).first()
            out["home_church"] = {
                "id": local_hc.id if local_hc else None,
                "external_id": str(hc_id) if hc_id else None,
                "name": (local_hc.name if local_hc else hc_name) or "",
                "meeting_day": local_hc.meeting_day if local_hc else None,
                "meeting_time": local_hc.meeting_time if local_hc else None,
                "address": local_hc.address if local_hc else None,
                "suburb": local_hc.suburb if local_hc else None,
            }
    except Exception:
        pass
    return out


@router.get("/portal/me")
def portal_me(request: Request, db: Session = Depends(get_db)):
    """Comprehensive member-centric snapshot for the redesigned portal."""
    member = _require_logged_in_member(request, db)

    member = db.query(Member).options(
        joinedload(Member.departments).joinedload(MemberDepartment.department).joinedload(Department.category)
    ).filter(Member.id == member.id).first()

    approved_departments = []
    pending_departments = []
    rejected_departments = []
    for md in member.departments:
        if not md.department:
            continue
        record = {
            "id": md.id,
            "department_id": md.department.id,
            "department_name": md.department.name,
            "category_name": md.department.category.name if md.department.category else None,
            "status": md.status or "pending",
            "source": md.source or "member",
            "admin_note": md.admin_note,
        }
        if record["status"] == "approved":
            approved_departments.append(record)
        elif record["status"] == "rejected":
            rejected_departments.append(record)
        else:
            pending_departments.append(record)

    leadership_roles: List[str] = []
    if member.leadership_roles:
        try:
            leadership_roles = json.loads(member.leadership_roles) if isinstance(member.leadership_roles, str) else list(member.leadership_roles)
        except (ValueError, TypeError):
            leadership_roles = []

    is_hc_leader = db.query(HomeChurch).filter(HomeChurch.leader_member_id == member.id).count() > 0
    led_home_churches = [
        {"id": hc.id, "name": hc.name}
        for hc in db.query(HomeChurch).filter(HomeChurch.leader_member_id == member.id).all()
    ]
    is_committee = _is_committee_member(db, member)
    hod_departments = [
        {"id": d.id, "name": d.name}
        for d in db.query(Department).filter(Department.hod_member_id == member.id).all()
    ]

    central = _portal_central_lookup(member, db)

    open_change_requests = (
        db.query(MemberChangeRequest)
        .filter(MemberChangeRequest.member_id == member.id, MemberChangeRequest.status == "pending")
        .count()
    )

    return {
        "member": {
            "id": member.id,
            "full_name": member.full_name,
            "phone": member.phone,
            "email": member.email or "",
            "address": member.address or "",
            "leadership_roles": leadership_roles,
            "external_synced": bool(member.external_member_id),
        },
        "departments": {
            "approved": approved_departments,
            "pending": pending_departments,
            "rejected": rejected_departments,
        },
        "ministries": central.get("ministries", []),
        "home_church": central.get("home_church"),
        "central_synced": central.get("central_synced", False),
        "assembly_name": central.get("assembly_name") or "",
        "leadership": {
            "is_hc_leader": is_hc_leader,
            "led_home_churches": led_home_churches,
            "is_hc_committee": is_committee,
            "is_hod": len(hod_departments) > 0,
            "hod_departments": hod_departments,
            "can_create_surveys": bool(getattr(member, "can_create_surveys", False)),
        },
        "open_change_requests": open_change_requests,
    }


def _serialize_change_request(cr: MemberChangeRequest) -> dict:
    payload = None
    if cr.payload:
        try:
            payload = json.loads(cr.payload)
        except (ValueError, TypeError):
            payload = None
    return {
        "id": cr.id,
        "member_id": cr.member_id,
        "member_name": cr.member.full_name if cr.member else "",
        "member_phone": cr.member.phone if cr.member else "",
        "change_type": cr.change_type,
        "summary": cr.summary,
        "details": cr.details,
        "payload": payload,
        "status": cr.status,
        "admin_response": cr.admin_response,
        "reviewed_by_id": cr.reviewed_by_id,
        "reviewed_by_name": cr.reviewed_by.full_name if cr.reviewed_by else None,
        "reviewed_at": cr.reviewed_at.isoformat() if cr.reviewed_at else None,
        "created_at": cr.created_at.isoformat() if cr.created_at else None,
    }


@router.get("/portal/change-requests")
def portal_my_change_requests(request: Request, db: Session = Depends(get_db)):
    member = _require_logged_in_member(request, db)
    rows = (
        db.query(MemberChangeRequest)
        .options(joinedload(MemberChangeRequest.reviewed_by))
        .filter(MemberChangeRequest.member_id == member.id)
        .order_by(MemberChangeRequest.created_at.desc())
        .all()
    )
    return [_serialize_change_request(r) for r in rows]


@router.post("/portal/change-requests")
def portal_create_change_request(payload: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    member = _require_logged_in_member(request, db)
    change_type = (payload.get("change_type") or "").strip().lower()
    summary = (payload.get("summary") or "").strip()
    details = (payload.get("details") or "").strip() or None
    structured = payload.get("payload") or None

    valid_types = {"profile", "department", "ministry", "home_church", "other"}
    if change_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Unknown change_type")
    if not summary:
        raise HTTPException(status_code=400, detail="Summary is required")
    if len(summary) > 300:
        summary = summary[:300]

    cr = MemberChangeRequest(
        member_id=member.id,
        change_type=change_type,
        summary=summary,
        details=details,
        payload=json.dumps(structured) if structured else None,
        status="pending",
    )
    db.add(cr)
    db.commit()
    db.refresh(cr)
    _log_member_action(
        request, db, member, "submit_change_request", "change_request", cr.id,
        f"{change_type}: {summary[:120]}",
    )
    db.commit()
    return _serialize_change_request(cr)


@router.get("/admin/change-requests")
def admin_list_change_requests(
    status: Optional[str] = Query(None),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """Admin only: list change requests, optionally filtered by status."""
    from routers.pages import is_authenticated
    if not is_authenticated(request):
        raise HTTPException(status_code=403, detail="Admin only")
    q = db.query(MemberChangeRequest).options(
        joinedload(MemberChangeRequest.member),
        joinedload(MemberChangeRequest.reviewed_by),
    )
    if status:
        q = q.filter(MemberChangeRequest.status == status)
    rows = q.order_by(MemberChangeRequest.created_at.desc()).all()
    return [_serialize_change_request(r) for r in rows]


@router.put("/admin/change-requests/{cr_id}")
def admin_resolve_change_request(
    cr_id: int,
    payload: dict = Body(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    from routers.pages import is_authenticated, get_admin_identity
    if not is_authenticated(request):
        raise HTTPException(status_code=403, detail="Admin only")
    cr = db.query(MemberChangeRequest).filter(MemberChangeRequest.id == cr_id).first()
    if not cr:
        raise HTTPException(status_code=404, detail="Change request not found")

    new_status = (payload.get("status") or "").strip().lower()
    if new_status not in {"approved", "rejected", "applied", "pending"}:
        raise HTTPException(status_code=400, detail="Invalid status")
    cr.status = new_status
    cr.admin_response = (payload.get("admin_response") or "").strip() or None
    identity = get_admin_identity(request) or {}
    cr.reviewed_by_id = identity.get("member_id")
    cr.reviewed_at = datetime.utcnow()
    _log_admin_action(
        request, db, "change_request_resolve", "change_request", cr.id,
        f"{new_status} — {cr.summary[:120]}",
    )
    db.commit()
    db.refresh(cr)
    return _serialize_change_request(cr)


# ============ MEMBER GIVING (portal proxy to central contributions API) ============
#
# The central API's CONTRIBUTION_VIEW_ROLES doesn't include regular members,
# so we cannot hand member credentials directly to it. Instead the portal
# calls the central API using its scoped service key after verifying the
# logged-in member matches the requested member_id. Members only ever see
# their OWN giving here.

def _require_member_with_central_link(request: Request, db: Session) -> Member:
    member = _require_logged_in_member(request, db)
    if not member.external_member_id:
        raise HTTPException(
            status_code=400,
            detail="Your member record isn't linked to the central database yet. Ask an admin to run Member Sync.",
        )
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        raise HTTPException(
            status_code=503,
            detail="Giving information isn't available right now (central database integration is disabled).",
        )
    return member


def _serialize_contribution(c: dict) -> dict:
    """Trim a central contribution record for the portal."""
    if not isinstance(c, dict):
        return {}
    return {
        "id": c.get("id"),
        "category": c.get("category") or "",
        "custom_label": c.get("custom_label") or None,
        "amount": c.get("amount"),
        "currency": c.get("currency") or "ZAR",
        "contribution_date": c.get("contribution_date"),
        "payment_method": c.get("payment_method") or None,
        "reference": c.get("reference") or None,
        "notes": c.get("notes") or None,
        "receipt_sent": bool(c.get("receipt_sent")),
        "receipt_sent_at": c.get("receipt_sent_at"),
    }


@router.get("/portal/giving/summary")
def portal_giving_summary(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """Per-category rollup for the logged-in member."""
    member = _require_member_with_central_link(request, db)
    r = _rfm.member_contribution_summary(
        member.external_member_id, from_date=from_date, to_date=to_date, db=db
    )
    if not r.ok:
        raise HTTPException(status_code=502, detail=f"Central API error: {r.error}")
    rows = r.data if isinstance(r.data, list) else (r.data or {}).get("data") or []
    total = 0.0
    out_rows = []
    for row in rows:
        try:
            amt = float(row.get("total_amount") or 0)
        except (TypeError, ValueError):
            amt = 0.0
        total += amt
        out_rows.append({
            "category": row.get("category") or "",
            "total_amount": amt,
            "count": int(row.get("count") or 0),
        })
    out_rows.sort(key=lambda x: x["total_amount"], reverse=True)
    return {
        "from_date": from_date,
        "to_date": to_date,
        "rows": out_rows,
        "total_amount": total,
        "currency": "ZAR",
    }


@router.get("/portal/giving/recent")
def portal_giving_recent(
    limit: int = Query(20, ge=1, le=100),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """Recent contributions for the logged-in member."""
    member = _require_member_with_central_link(request, db)
    r = _rfm.list_member_contributions(
        member.external_member_id, page=1, size=limit, db=db,
    )
    if not r.ok:
        raise HTTPException(status_code=502, detail=f"Central API error: {r.error}")
    items = []
    if isinstance(r.data, list):
        items = r.data
    elif isinstance(r.data, dict):
        items = r.data.get("data") or []
    return {
        "items": [_serialize_contribution(c) for c in items],
        "count": len(items),
    }


@router.get("/admin/debug/central-member/{member_id}")
def admin_debug_central_member(member_id: int, request: Request, db: Session = Depends(get_db)):
    """Admin diagnostic: dump the raw central API response for a local member.
    Useful when something the portal expects (ministries, home_church_id…)
    isn't surfacing — we can see exactly what the central API returned."""
    from routers.pages import is_authenticated
    if not is_authenticated(request):
        raise HTTPException(status_code=403, detail="Admin only")
    member = db.query(Member).filter(Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="Local member not found")
    out = {
        "local": {
            "id": member.id,
            "full_name": member.full_name,
            "phone": member.phone,
            "external_member_id": member.external_member_id,
            "external_assembly_id": member.external_assembly_id,
            "external_match_status": member.external_match_status,
        },
        "central": None,
        "central_error": None,
    }
    if not member.external_member_id:
        out["central_error"] = "Member is not linked to the central database (run Member Sync)."
        return out
    if not _rfm.is_enabled(db) or not _rfm.is_configured(db):
        out["central_error"] = "Central API integration disabled or unconfigured."
        return out
    r = _rfm.get_member(member.external_member_id, db=db)
    if not r.ok:
        out["central_error"] = r.error
    else:
        out["central"] = r.data
    return out


@router.get("/portal/attendance")
def portal_attendance(year: Optional[int] = Query(None), request: Request = None, db: Session = Depends(get_db)):
    """Attendance dashboard for the logged-in member, sourced from the
    central rfm-database. Returns the API's payload as-is so the portal
    UI is tolerant of fields we add later."""
    member = _require_member_with_central_link(request, db)
    if not year:
        year = datetime.utcnow().year
    r = _rfm.member_attendance_dashboard(member.external_member_id, year=year, db=db)
    if not r.ok:
        raise HTTPException(status_code=502, detail=f"Central API error: {r.error}")
    return {
        "year": year,
        "data": r.data,
    }


@router.get("/portal/giving/banking")
def portal_giving_banking(request: Request, db: Session = Depends(get_db)):
    """Banking details for the church.

    Source-of-truth order (each field falls back independently):
      1. Assembly.metadata in the central rfm-database — keyed by the
         member's external_assembly_id. Recognised keys: bank_name,
         account_holder, account_number, branch_code, account_type,
         swift, reference_hint, online_giving_url.
      2. Local Settings table (the previous behaviour) so existing
         deployments keep working until the central record is filled in.
    """
    member = _require_logged_in_member(request, db)

    def _s(key: str) -> str:
        row = db.query(Settings).filter(Settings.key == key).first()
        return (row.value or "").strip() if row else ""

    # Local fallbacks
    central: dict = {}
    central_source = None

    # Banking is church-wide. ALWAYS use the portal's default assembly id —
    # never the caller's own external_assembly_id — so every member sees the
    # same bank details regardless of whether their record is synced, whether
    # an old central record points at a different assembly, or whether the
    # API key is admin-grade vs scoped. Try caller's id only as a last-ditch
    # fallback in case _resolve_default_assembly_id returns nothing.
    #   assembly.metadata.banking_details — the canonical location
    # Top-level metadata keys are tolerated for backward compatibility.
    candidate_assembly_ids = []
    if _rfm.is_enabled(db) and _rfm.is_configured(db):
        try:
            primary = _resolve_default_assembly_id(db)
            if primary:
                candidate_assembly_ids.append(str(primary))
        except Exception:
            pass
        # Caller's own as a backup if it's different from the default
        if member.external_assembly_id and str(member.external_assembly_id) not in candidate_assembly_ids:
            candidate_assembly_ids.append(str(member.external_assembly_id))

    for assembly_id_for_banking in candidate_assembly_ids:
        try:
            r = _rfm.get_assembly(assembly_id_for_banking, db=db)
            if not (r.ok and isinstance(r.data, dict)):
                continue
            meta = r.data.get("metadata") or r.data.get("extra_metadata") or {}
            if not isinstance(meta, dict):
                continue
            bank_block = meta.get("banking_details")
            if not isinstance(bank_block, dict):
                bank_block = meta  # backward-compat: top-level keys
            norm = {
                str(k).lower(): str(v).strip()
                for k, v in bank_block.items()
                if v not in (None, "")
            }
            if any(norm.get(k) for k in ("bank_name", "account_number", "online_giving_url")):
                central = norm
                central_source = "central"
                break  # first assembly with banking data wins
        except Exception:
            continue

    def pick(*keys, fallback_setting=None):
        # Try central first (any matching key), then a local Settings key
        for k in keys:
            v = central.get(k)
            if v:
                return v
        if fallback_setting:
            return _s(fallback_setting)
        return ""

    # Map central -> response, with local Settings as a fallback per-field
    response = {
        "bank_name":         pick("bank_name", "bank",                 fallback_setting="bank_name"),
        "account_holder":    pick("account_holder", "account_name",    fallback_setting="bank_account_holder"),
        "account_number":    pick("account_number",                    fallback_setting="bank_account_number"),
        "branch_code":       pick("branch_code", "branch",             fallback_setting="bank_branch_code"),
        "account_type":      pick("account_type",                      fallback_setting="bank_account_type"),
        "swift":             pick("swift", "swift_code", "bic",        fallback_setting="bank_swift"),
        "reference_hint":    pick("reference_hint", "reference",       fallback_setting="bank_reference_hint")
                             or "Use your full name as reference",
        "online_giving_url": pick("online_giving_url", "give_url",     fallback_setting="online_giving_url"),
        "source": central_source or ("local" if any([
            _s("bank_name"), _s("bank_account_number"), _s("online_giving_url")
        ]) else None),
    }
    return response


# ============ ANNOUNCEMENTS ============

def _serialize_announcement(a: Announcement) -> dict:
    return {
        "id": a.id,
        "title": a.title,
        "body": a.body or "",
        "pinned": bool(a.pinned),
        "is_active": bool(a.is_active),
        "starts_at": a.starts_at.isoformat() if a.starts_at else None,
        "ends_at": a.ends_at.isoformat() if a.ends_at else None,
        "link_url": a.link_url or None,
        "link_label": a.link_label or None,
        "created_by": (
            {"id": a.created_by.id, "full_name": a.created_by.full_name}
            if a.created_by else None
        ),
        "created_at": a.created_at.isoformat() if a.created_at else None,
        "updated_at": a.updated_at.isoformat() if a.updated_at else None,
    }


def _is_visible_now(a: Announcement, now: datetime) -> bool:
    if not a.is_active:
        return False
    if a.starts_at and a.starts_at > now:
        return False
    if a.ends_at and a.ends_at < now:
        return False
    return True


# ============ VERSE OF THE DAY ============
# Curated rotation, day-of-year indexed so every member sees the same verse
# on the same day and the rotation feels intentional rather than random.

_VERSE_OF_DAY: List[dict] = [
    {"text": "Be still, and know that I am God.", "ref": "Psalm 46:10"},
    {"text": "The Lord is my shepherd, I shall not want.", "ref": "Psalm 23:1"},
    {"text": "I can do all things through Christ who strengthens me.", "ref": "Philippians 4:13"},
    {"text": "For God so loved the world that he gave his one and only Son.", "ref": "John 3:16"},
    {"text": "Trust in the Lord with all your heart, and lean not on your own understanding.", "ref": "Proverbs 3:5"},
    {"text": "The Lord is close to the brokenhearted and saves those who are crushed in spirit.", "ref": "Psalm 34:18"},
    {"text": "Cast all your anxiety on him because he cares for you.", "ref": "1 Peter 5:7"},
    {"text": "But those who hope in the Lord will renew their strength.", "ref": "Isaiah 40:31"},
    {"text": "And we know that in all things God works for the good of those who love him.", "ref": "Romans 8:28"},
    {"text": "Greater love has no one than this: to lay down one's life for one's friends.", "ref": "John 15:13"},
    {"text": "The thief comes only to steal and kill and destroy; I have come that they may have life, and have it to the full.", "ref": "John 10:10"},
    {"text": "Therefore, if anyone is in Christ, the new creation has come: The old has gone, the new is here!", "ref": "2 Corinthians 5:17"},
    {"text": "Come to me, all you who are weary and burdened, and I will give you rest.", "ref": "Matthew 11:28"},
    {"text": "Do not be anxious about anything, but in every situation, by prayer and petition, with thanksgiving, present your requests to God.", "ref": "Philippians 4:6"},
    {"text": "The Lord your God is with you, the Mighty Warrior who saves.", "ref": "Zephaniah 3:17"},
    {"text": "Now faith is confidence in what we hope for and assurance about what we do not see.", "ref": "Hebrews 11:1"},
    {"text": "Above all else, guard your heart, for everything you do flows from it.", "ref": "Proverbs 4:23"},
    {"text": "If we confess our sins, he is faithful and just and will forgive us our sins.", "ref": "1 John 1:9"},
    {"text": "Be strong and courageous. Do not be afraid; do not be discouraged, for the Lord your God will be with you wherever you go.", "ref": "Joshua 1:9"},
    {"text": "Delight yourself in the Lord, and he will give you the desires of your heart.", "ref": "Psalm 37:4"},
    {"text": "Therefore go and make disciples of all nations.", "ref": "Matthew 28:19"},
    {"text": "Whoever dwells in the shelter of the Most High will rest in the shadow of the Almighty.", "ref": "Psalm 91:1"},
    {"text": "Love is patient, love is kind.", "ref": "1 Corinthians 13:4"},
    {"text": "Rejoice in the Lord always. I will say it again: Rejoice!", "ref": "Philippians 4:4"},
    {"text": "The Lord is my light and my salvation — whom shall I fear?", "ref": "Psalm 27:1"},
    {"text": "Weeping may stay for the night, but rejoicing comes in the morning.", "ref": "Psalm 30:5"},
    {"text": "Your word is a lamp for my feet, a light on my path.", "ref": "Psalm 119:105"},
    {"text": "Have I not commanded you? Be strong and courageous.", "ref": "Joshua 1:9"},
    {"text": "I have learned to be content whatever the circumstances.", "ref": "Philippians 4:11"},
    {"text": "Taste and see that the Lord is good; blessed is the one who takes refuge in him.", "ref": "Psalm 34:8"},
    {"text": "The name of the Lord is a fortified tower; the righteous run to it and are safe.", "ref": "Proverbs 18:10"},
    {"text": "Let your light shine before others, that they may see your good deeds and glorify your Father in heaven.", "ref": "Matthew 5:16"},
    {"text": "The fear of the Lord is the beginning of wisdom.", "ref": "Proverbs 9:10"},
    {"text": "Seek first his kingdom and his righteousness, and all these things will be given to you as well.", "ref": "Matthew 6:33"},
    {"text": "Be kind and compassionate to one another, forgiving each other, just as in Christ God forgave you.", "ref": "Ephesians 4:32"},
    {"text": "He gives strength to the weary and increases the power of the weak.", "ref": "Isaiah 40:29"},
    {"text": "Let us hold unswervingly to the hope we profess, for he who promised is faithful.", "ref": "Hebrews 10:23"},
    {"text": "Do not be overcome by evil, but overcome evil with good.", "ref": "Romans 12:21"},
    {"text": "For where two or three gather in my name, there am I with them.", "ref": "Matthew 18:20"},
    {"text": "And without faith it is impossible to please God.", "ref": "Hebrews 11:6"},
]


@router.get("/portal/verse-of-the-day")
def portal_verse_of_the_day():
    """Today's verse. Day-of-year indexed so the rotation is stable and
    every member sees the same verse on the same day."""
    today = date.today()
    idx = (today.toordinal()) % len(_VERSE_OF_DAY)
    v = _VERSE_OF_DAY[idx]
    return {
        "date": today.isoformat(),
        "text": v["text"],
        "reference": v["ref"],
    }


@router.get("/portal/announcements")
def portal_announcements(request: Request, db: Session = Depends(get_db)):
    """Active, in-window announcements for the logged-in member, pinned first."""
    _require_logged_in_member(request, db)
    now = datetime.utcnow()
    rows = (
        db.query(Announcement)
        .order_by(Announcement.pinned.desc(), Announcement.starts_at.desc().nullsfirst(),
                  Announcement.created_at.desc())
        .all()
    )
    visible = [a for a in rows if _is_visible_now(a, now)]
    return [_serialize_announcement(a) for a in visible]


# ---- Admin CRUD ----

def _parse_dt(value) -> Optional[datetime]:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    if not s:
        return None
    # Accept ISO with or without seconds; allow "YYYY-MM-DD" and "YYYY-MM-DDTHH:MM"
    try:
        if len(s) == 10:
            return datetime.fromisoformat(s + "T00:00:00")
        if len(s) == 16:
            return datetime.fromisoformat(s + ":00")
        # Strip trailing Z
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        return datetime.fromisoformat(s)
    except ValueError:
        return None


@router.get("/admin/announcements")
def admin_list_announcements(request: Request, db: Session = Depends(get_db)):
    from routers.pages import is_authenticated
    if not is_authenticated(request):
        raise HTTPException(status_code=403, detail="Admin only")
    rows = (
        db.query(Announcement)
        .options(joinedload(Announcement.created_by))
        .order_by(Announcement.pinned.desc(), Announcement.created_at.desc())
        .all()
    )
    return [_serialize_announcement(a) for a in rows]


@router.post("/admin/announcements")
def admin_create_announcement(payload: dict = Body(...), request: Request = None, db: Session = Depends(get_db)):
    from routers.pages import is_authenticated, get_admin_identity
    if not is_authenticated(request):
        raise HTTPException(status_code=403, detail="Admin only")

    title = (payload.get("title") or "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title is required")
    identity = get_admin_identity(request) or {}

    a = Announcement(
        title=title[:200],
        body=(payload.get("body") or "").strip() or None,
        pinned=bool(payload.get("pinned", False)),
        is_active=bool(payload.get("is_active", True)),
        starts_at=_parse_dt(payload.get("starts_at")),
        ends_at=_parse_dt(payload.get("ends_at")),
        link_url=(payload.get("link_url") or "").strip() or None,
        link_label=(payload.get("link_label") or "").strip()[:80] or None,
        created_by_member_id=identity.get("member_id"),
    )
    db.add(a)
    db.commit()
    db.refresh(a)
    _log_admin_action(request, db, "announcement_create", "announcement", a.id, title[:120])
    db.commit()

    # Best-effort push for pinned + currently-visible announcements
    if a.pinned and _is_visible_now(a, datetime.utcnow()):
        try:
            import push_service
            if push_service.is_configured(db):
                push_service.send_to_all(
                    db,
                    title=a.title,
                    body=(a.body or "")[:200],
                    url=a.link_url or "/portal",
                    tag=f"announce-{a.id}",
                )
        except Exception:
            pass

    return _serialize_announcement(a)


@router.put("/admin/announcements/{announcement_id}")
def admin_update_announcement(
    announcement_id: int,
    payload: dict = Body(...),
    request: Request = None,
    db: Session = Depends(get_db),
):
    from routers.pages import is_authenticated
    if not is_authenticated(request):
        raise HTTPException(status_code=403, detail="Admin only")
    a = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Announcement not found")

    if "title" in payload:
        title = (payload.get("title") or "").strip()
        if not title:
            raise HTTPException(status_code=400, detail="Title cannot be empty")
        a.title = title[:200]
    if "body" in payload:
        a.body = (payload.get("body") or "").strip() or None
    if "pinned" in payload:
        a.pinned = bool(payload["pinned"])
    if "is_active" in payload:
        a.is_active = bool(payload["is_active"])
    if "starts_at" in payload:
        a.starts_at = _parse_dt(payload.get("starts_at"))
    if "ends_at" in payload:
        a.ends_at = _parse_dt(payload.get("ends_at"))
    if "link_url" in payload:
        a.link_url = (payload.get("link_url") or "").strip() or None
    if "link_label" in payload:
        a.link_label = (payload.get("link_label") or "").strip()[:80] or None

    db.commit()
    db.refresh(a)
    _log_admin_action(request, db, "announcement_update", "announcement", a.id, a.title[:120])
    db.commit()
    return _serialize_announcement(a)


@router.delete("/admin/announcements/{announcement_id}")
def admin_delete_announcement(announcement_id: int, request: Request, db: Session = Depends(get_db)):
    from routers.pages import is_authenticated
    if not is_authenticated(request):
        raise HTTPException(status_code=403, detail="Admin only")
    a = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not a:
        raise HTTPException(status_code=404, detail="Announcement not found")
    title = a.title
    db.delete(a)
    db.commit()
    _log_admin_action(request, db, "announcement_delete", "announcement", announcement_id, title[:120])
    db.commit()
    return {"success": True}
